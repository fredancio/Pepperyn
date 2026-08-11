"""
fru_sign_convention_detector.py — FRU v0 STUB: expense sign-convention
detector for Epistemic Dialogue's first executable vertical slice.

Scope, per `docs/Architecture/Cognitive/FRU_EPISTEMIC_FIRST_VERTICAL_SLICE.md`
and the mission that authorized this implementation (2026-08-09): the
smallest deterministic detector that answers exactly one question —

    Does the source strongly suggest that expenses are represented as
    positive absolute values whose economic meaning comes from
    account/post nature rather than raw numeric sign?

THIS IS NOT GENERAL FRU. It is a stub, honestly labeled (mirrors Knowledge
Model v0's own Phidani-loop tests, which used a minimal stub rather than
real FRU). It answers one subject (`EXPENSE_SIGN_CONVENTION`) for one
validated file shape (Belgian PCMN-style account codes). Extending this to
arbitrary financial representation, other subjects, or other chart-of-
accounts conventions is explicitly out of scope — that is the next,
larger, not-yet-authorized mission.

WHAT THIS MODULE NEVER DOES:
  - No LLM call, anywhere, for any reason.
  - No general chart-of-accounts engine — the charge-code prefix set
    below is the minimum needed to interpret Phidani's own codes, not a
    general Belgian PCMN library.
  - No fabricated certainty — degrades honestly to UNKNOWN whenever the
    two signals below disagree or neither is available, exactly as
    `FRU_EPISTEMIC_FIRST_VERTICAL_SLICE.md` §3/§6 specifies. Never
    guesses.

TWO DETERMINISTIC SIGNALS, VALIDATED AGAINST THE REAL PHIDANI.XLSX
(direct inspection, 2026-08-09 — see the implementation mission's own
final report for exact evidence, not asserted from memory):

  1. ACCOUNT-CODE-RANGE SIGNAL — rows whose account code falls in the
     Belgian PCMN charge range (60-66, `_CHARGE_CODE_PREFIXES` below).
     Real Phidani.xlsx: roughly 190 charge-coded rows per period column
     pass the prefix filter, of which a small number (verified: 2 for
     period column C — row 163 code '630' formula '=C162'; row 231 code
     '650' formula '=SUM(C174:C230)', aggregating 57 other rows) are
     NOT independent leaf observations — CORRECTED (independent
     adversarial review, correction 2, 2026-08-09): these are
     formula-derived aggregate/rollup rows that happen to carry an
     in-range account code, and are now structurally excluded (see
     `detect_expense_sign_convention_from_workbook` below) — not by
     hardcoding their codes, but by checking whether the row's own
     period-column cell is itself a formula (derived) rather than a
     literal value (a genuine leaf observation). The prior version of
     this module only special-cased the literal string "60"; this
     generalizes that same principle to any charge-coded row, whatever
     its code. Among the remaining genuine leaf observations, a small
     minority are negative — legitimate documented exceptions (e.g.
     "Utilisations et reprises (-)", "Récupération précompte
     professionnel" — refunds/reversals, whose own label in the source
     already flags them as an adjustment) — NOT evidence of a different
     sign convention for ordinary charges. This is exactly why the
     signal is a MAJORITY/consistency check (`_NONNEG_MAJORITY_THRESHOLD`),
     never an "any negative disqualifies" rule — a single-exception veto
     would be fragile and would misclassify this real file. A genuinely
     mixed distribution (no clear majority either way) still correctly
     degrades to "signal absent" rather than guessing.
     NOTE: this signal's observations are drawn from one sheet/company's
     real leaf rows — they are independent in the sense of "not
     mathematically derived from each other," not in a formal statistical
     sense; no claim of i.i.d. sampling is made or needed for a
     majority/consistency check.
  2. ARITHMETIC/SUBTOTAL SIGNAL — Phidani.xlsx has a row whose account
     code is the bare aggregate `"60"` (charges subtotal, e.g.
     `=SUM(C32:C33)`), and a later "Marge brute" row whose formula
     SUBTRACTS that subtotal's cell from revenue (`=C31-C34`, confirmed
     identical pattern across every period column C through... in the
     real file). Subtraction of a *positive* charges subtotal to compute
     a margin is only arithmetically sound under ABSOLUTE_POSITIVE — an
     ADDITION pattern would instead indicate SIGNED_NATURAL. This module
     searches for whichever pattern actually exists, on the real cell
     reference of the "60" row, in the given period column — never
     hardcoded to one cell address.

Combination rule (mirrors `FRU_EPISTEMIC_FIRST_VERTICAL_SLICE.md` §3
exactly): both signals agree -> STRONG_INFERENCE. Only one signal present
(the other genuinely absent, not contradicting) -> HYPOTHESIS. Both
present but disagree, or neither present -> UNKNOWN, candidate value None.
"""
from __future__ import annotations

import re
from typing import Optional

from .candidate import Candidate

# RELOCATED 2026-08-11 (PERSONNEL_COST_CLASSIFIER_V0_IMPLEMENTATION_CONTRACT.md
# §3, implementation mission §4 — "relocate, do not duplicate"): `Candidate`
# now lives in `services/candidate.py`, imported here rather than redefined,
# because `personnel_cost_classifier.py` is a second real consumer that
# needs the same shape plus a bounded extension (CONTRADICTION tier,
# evidence lists) `Candidate` could not previously represent. Nothing about
# FRU's own behavior changes: this module never sets `tier="CONTRADICTION"`
# and never populates the evidence-list fields (both default to `()`), so
# every existing call site (`Candidate(value=..., tier=...)` below) and
# every existing consumer (`epistemic_dialogue_service.py`,
# `backend/tests/test_epistemic_dialogue_v0.py`, both of which import
# `Candidate` from *this* module) continues to work unchanged — the name
# `Candidate` is re-exported from this module's own namespace by this
# import, so `from backend.services.fru_sign_convention_detector import
# Candidate` still resolves.

# Belgian PCMN charge-account prefixes (classes 60-66) — the minimum
# needed to interpret Phidani's own codes (mission-scoped, not a general
# chart-of-accounts library; see module docstring).
_CHARGE_CODE_PREFIXES = {"60", "61", "62", "63", "64", "65", "66"}

# Majority threshold for the code-range signal: >=90% non-negative ->
# ABSOLUTE_POSITIVE; >=90% negative -> SIGNED_NATURAL; anything genuinely
# mixed in between -> signal absent (never guesses). Chosen because it
# tolerates the real file's own documented exceptions (2.1% negative)
# without being fooled by a genuinely different convention.
#
# Numeric representation (repaired 2026-08-09, "FRU SIGN CONVENTION STUB —
# 90% NUMERIC BOUNDARY REPAIR" mission): the 90% business threshold itself
# is unchanged. What changed is HOW the two symmetric comparisons
# (non-negative majority / negative majority) are computed. The prior
# implementation derived the negative-side bound via floating-point
# subtraction (`1 - _NONNEG_MAJORITY_THRESHOLD`), which is not exactly 0.1
# in IEEE 754 (`1 - 0.9 == 0.09999999999999998`) — this silently broke
# symmetry at exactly the 90% boundary: the ABSOLUTE_POSITIVE side (a
# direct `ratio >= 0.9` comparison) fired correctly at exactly 90%, but
# the SIGNED_NATURAL side (`ratio <= 1 - 0.9`) did not, because
# `0.1 <= 0.09999999999999998` is False. The threshold is now expressed
# as an exact integer fraction (NUM/DEN = 9/10) and both comparisons are
# performed via integer cross-multiplication on exact integer COUNTS
# (`nonneg`, and `neg = total - nonneg`, itself exact integer subtraction)
# — never via a derived floating-point ratio or a floating-point
# subtraction of the threshold. This makes the two comparisons
# mathematically, not just intentionally, symmetric at any exact
# threshold expressible as a small integer fraction.
_NONNEG_MAJORITY_THRESHOLD_NUM = 9
_NONNEG_MAJORITY_THRESHOLD_DEN = 10
_NONNEG_MAJORITY_THRESHOLD = _NONNEG_MAJORITY_THRESHOLD_NUM / _NONNEG_MAJORITY_THRESHOLD_DEN  # 0.9, retained for documentation/introspection only — no comparison uses this float

ABSOLUTE_POSITIVE = "ABSOLUTE_POSITIVE"
SIGNED_NATURAL = "SIGNED_NATURAL"


def _code_range_signal(charge_values: list[float]) -> Optional[str]:
    """
    Majority/consistency signal (candidate signal 1, module docstring).
    Returns None (signal absent) if there is no data, or if the
    distribution is genuinely mixed with no clear majority either way —
    never guesses.

    Numeric method (repaired 2026-08-09 — see
    `_NONNEG_MAJORITY_THRESHOLD_NUM`/`_DEN`'s own comment for the full
    root-cause explanation): both branches compare exact integer COUNTS
    via cross-multiplication against the exact integer threshold
    fraction, never a derived floating-point ratio and never a
    floating-point subtraction of the threshold. `neg` is computed as
    `total - nonneg` — exact integer subtraction, not `1 - <float ratio>`
    — which is what makes the two branches provably symmetric at exactly
    the threshold, not merely symmetric by intention.
    """
    if not charge_values:
        return None
    total = len(charge_values)
    nonneg = sum(1 for v in charge_values if v >= 0)
    neg = total - nonneg
    # nonneg/total >= NUM/DEN  <=>  nonneg*DEN >= NUM*total (exact, integer).
    if nonneg * _NONNEG_MAJORITY_THRESHOLD_DEN >= _NONNEG_MAJORITY_THRESHOLD_NUM * total:
        return ABSOLUTE_POSITIVE
    # neg/total >= NUM/DEN  <=>  neg*DEN >= NUM*total (exact, integer) —
    # the mirror-image comparison, structurally identical in form to the
    # one above, on `neg` rather than `nonneg`. This is the fix: previously
    # this branch used `ratio <= (1 - _NONNEG_MAJORITY_THRESHOLD)`.
    if neg * _NONNEG_MAJORITY_THRESHOLD_DEN >= _NONNEG_MAJORITY_THRESHOLD_NUM * total:
        return SIGNED_NATURAL
    return None


def detect_expense_sign_convention(
    charge_values: list[float],
    charges_subtracted_in_margin_formula: Optional[bool],
) -> Candidate:
    """
    Pure, deterministic core (unit-testable with synthetic data,
    Phase 15's adversarial matrix) — no file I/O here.

    Args:
        charge_values: observed numeric values for account-coded charge
            rows (candidate signal 1), for a single period column.
        charges_subtracted_in_margin_formula: True if a formula subtracts
            the charges-subtotal cell (signal -> ABSOLUTE_POSITIVE), False
            if a formula adds it instead (signal -> SIGNED_NATURAL), None
            if no such formula relationship was found (signal absent).

    Returns:
        Candidate — never fabricates a value when the two signals
        disagree or neither is available (module docstring).
    """
    codes_signal = _code_range_signal(charge_values)
    if charges_subtracted_in_margin_formula is True:
        arithmetic_signal = ABSOLUTE_POSITIVE
    elif charges_subtracted_in_margin_formula is False:
        arithmetic_signal = SIGNED_NATURAL
    else:
        arithmetic_signal = None

    if codes_signal is not None and arithmetic_signal is not None:
        if codes_signal == arithmetic_signal:
            return Candidate(value=codes_signal, tier="STRONG_INFERENCE")
        return Candidate(value=None, tier="UNKNOWN")  # disagree -> never guess

    if codes_signal is not None:
        return Candidate(value=codes_signal, tier="HYPOTHESIS")
    if arithmetic_signal is not None:
        return Candidate(value=arithmetic_signal, tier="HYPOTHESIS")
    return Candidate(value=None, tier="UNKNOWN")


def detect_expense_sign_convention_from_workbook(
    path: str, period_column: int, sheet_name: Optional[str] = None,
) -> Candidate:
    """
    Real-file adapter (Phase 14) — the only place this module touches
    openpyxl. Extracts the two signals from an actual workbook shaped
    like Phidani.xlsx and delegates to the pure core above. Never used
    by the synthetic adversarial matrix tests (Phase 15), which call
    `detect_expense_sign_convention` directly.

    `period_column` is a 1-indexed column number (e.g. 3 for column C),
    matching openpyxl's own convention.

    STRUCTURAL, NOT SEMANTIC, ROLLUP EXCLUSION (correction 2, independent
    adversarial review, 2026-08-09): a row is treated as a genuine leaf
    observation only if its own period-column cell holds a literal
    numeric value in the FORMULAS-mode workbook (i.e. the cell is not
    itself a formula). A row whose period-column cell IS a formula
    (`=C162`, `=SUM(C174:C230)`, ...) is, by definition, derived from
    other cells and is excluded from `charge_values` regardless of what
    its own account code happens to be — this is a structural test on
    the cell's own nature (formula vs. literal), never a hardcoded list
    of "known" rollup codes, and never an attempt to parse or understand
    what the formula computes (no general Excel semantic engine). The
    bare "60" aggregate row is still special-cased separately below
    because it additionally supplies the cell reference the arithmetic
    signal searches for — that role is independent of leaf-exclusion and
    is unchanged by this correction.
    """
    import openpyxl  # local import: this module's pure core has zero
    # dependency on openpyxl; only this adapter function does.

    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    ws_f = wb_formulas[sheet_name] if sheet_name else wb_formulas.active
    ws_v = wb_values[sheet_name] if sheet_name else wb_values.active

    charge_values: list[float] = []
    charges_subtotal_row: Optional[int] = None

    for row in ws_f.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        code = cell.value
        if code is None:
            continue
        if isinstance(code, str) and code.strip() == "60":
            charges_subtotal_row = cell.row
            continue
        try:
            icode = int(code)
        except (TypeError, ValueError):
            continue
        if icode <= 99:
            # Excludes bare aggregate codes (e.g. a hypothetical "70"
            # parsed as int) — real account lines in this file are far
            # more than two digits (604000, 611001, ...).
            continue
        prefix = str(icode)[:2]
        if prefix not in _CHARGE_CODE_PREFIXES:
            continue
        # Structural rollup exclusion (correction 2): a formula in this
        # row's own period-column cell means the value is derived from
        # other rows, not an independent leaf observation — exclude it
        # regardless of its account code, never by hardcoding which
        # codes are "known" rollups.
        period_cell_formula = ws_f.cell(row=cell.row, column=period_column).value
        if isinstance(period_cell_formula, str) and period_cell_formula.startswith("="):
            continue
        v = ws_v.cell(row=cell.row, column=period_column).value
        if isinstance(v, (int, float)):
            charge_values.append(float(v))

    charges_subtracted: Optional[bool] = None
    if charges_subtotal_row is not None:
        col_letter = openpyxl.utils.get_column_letter(period_column)
        target_ref = f"{col_letter}{charges_subtotal_row}"
        minus_pattern = re.compile(rf"-\s*{re.escape(target_ref)}\b")
        plus_pattern = re.compile(rf"\+\s*{re.escape(target_ref)}\b")
        for row in ws_f.iter_rows(min_col=period_column, max_col=period_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if minus_pattern.search(cell.value):
                        charges_subtracted = True
                        break
                    if plus_pattern.search(cell.value):
                        charges_subtracted = False
                        break
            if charges_subtracted is not None:
                break

    return detect_expense_sign_convention(charge_values, charges_subtracted)
