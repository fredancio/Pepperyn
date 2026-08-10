"""
formula_evidence.py — Observation Structure v0: dedicated formula-evidence reader.

Canonical contract: docs/Architecture/Cognitive/OBSERVATION_STRUCTURE_V0_IMPLEMENTATION_CONTRACT.md
Section §5 (resolved 2026-08-10, "Observation Structure v0 — Formula Access
Arbitration" mission): the shared production Excel parser
(services/file_parser.py) reads workbooks with openpyxl(data_only=True) as
its primary strategy — it sees cached formula RESULTS, never formula TEXT.
Formula presence/text is therefore not observable from that pipeline today.

This module is the decided, additive answer: exactly ONE additional
openpyxl.load_workbook(..., data_only=False) read, performed only by a
consumer that actually needs formula evidence. file_parser.py is not
imported, not modified, and not duplicated by this module.

WHAT THIS MODULE NEVER DOES (contract §5, §9):
  - Never produces or exposes a numeric value. file_parser.py's
    data_only=True cached-value output remains the sole canonical numeric
    source ("no second source of truth"). This module exposes only
    booleans and raw formula text.
  - Never evaluates, recalculates, or interprets a formula. `formula_text`
    is exactly the string openpyxl returns for the cell's raw content.
  - Never claims a cached result is "fresh" merely because a formula is
    present — formula presence proves derivation *mechanism*, not
    currency of any cached result (contract §5 and §3.2's honesty note).
    This module does not open a data_only=True companion view at all, so
    it never even sees a cached result to make claims about.
  - Never opens macros, never follows external links, never performs
    network access. openpyxl's default `keep_vba=False` load is used;
    macros are not executed by openpyxl regardless.

SAME-ARTIFACT GUARANTEE — CORRECTED FRAMING (2026-08-10, independent
adversarial pre-merge review): this module's contribution is an INTERFACE
that ENABLES the guarantee, not an end-to-end guarantee by itself.
`load_formula_workbook` takes `file_bytes: bytes` — the exact same
parameter shape as `file_parser.parse_file(file_bytes, filename)`
(verified against backend/services/file_parser.py:35 and
backend/connectors/file_connector.py's FileConnector, which reads
`UploadFile` bytes exactly once in backend/routers/analyze.py and reuses
that same `bytes` object for both the quality gate and the parse). IF a
caller holds that same `bytes` object and passes it to both the canonical
numeric parse and this module, THEN there is no re-download, no alternate
path, and no regenerated copy — both reads are `io.BytesIO(file_bytes)`-
wrapped views of the identical byte string, exactly mirroring
file_parser.py's own internal `io.BytesIO(file_bytes)` usage
(file_parser.py:61). This module never accepts or opens a filesystem path,
which removes one entire class of drift (no second file-path resolution
that could ever point at a different file).

What this does NOT yet prove: that any production orchestration actually
does this. As of this commit, `observation_structure`/`formula_evidence`
are not called from any router — nothing in the live request path passes
`file_bytes` to this module at all (confirmed by `git diff`/`grep`: zero
references from `backend/routers/*`). The guarantee is therefore
structural / interface-level / proven only by this module's own test
suite (which constructs the "same bytes, two reads" scenario directly) —
it is NOT yet a verified production guarantee. That becomes an actual
obligation, requiring its own dedicated integration test asserting the
same `bytes` variable reaches both call sites, only when a future
consumer (e.g. the deferred FRU refactor, contract §15) is actually wired
in. Do not cite this module as proof that the guarantee holds end-to-end
today — only that nothing about its own design would break it.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Optional

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class CellFormulaEvidence:
    """Formula evidence for exactly one (row, column) cell.

    Never a numeric value — see module docstring, "no second source of
    truth" (contract §5, §9).
    """

    is_formula: bool
    formula_text: Optional[str]  # raw, unparsed, unevaluated; None unless is_formula is True.
    has_literal_value: bool  # True iff the raw cell content (data_only=False) is non-None and not a formula.


def load_formula_workbook(file_bytes: bytes) -> Workbook:
    """Perform the one additional data_only=False read (contract §5).

    Pure I/O boundary: this is the only function in this module (and, per
    the contract, the only *additional* file read introduced by this
    slice) that touches openpyxl.load_workbook. Callers should load once
    and reuse the returned Workbook for every cell they need — do not call
    this per-cell.
    """
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)


def get_worksheet(wb: Workbook, sheet_name: Optional[str] = None) -> Worksheet:
    """Select a worksheet by name, or the active sheet if none given.

    Mirrors the selection convention already used by
    fru_sign_convention_detector.py::detect_expense_sign_convention_from_workbook
    (sheet_name optional, defaults to the workbook's active sheet) — no new
    convention introduced.
    """
    if sheet_name is not None:
        return wb[sheet_name]
    return wb.active


def cell_formula_evidence(ws: Worksheet, row: int, column: int) -> CellFormulaEvidence:
    """Formula evidence for one cell, given an already-loaded formula-mode
    worksheet (from load_formula_workbook + get_worksheet).

    Pure function: no I/O, no interpretation beyond "does this cell's raw
    content look like a formula, and if not, is there a literal value at
    all." Coordinates are real, unshifted openpyxl (row, column) — 1-indexed,
    matching openpyxl's own convention, never the position-remapped index
    file_parser.py's DataFrame pipeline produces (contract §5).
    """
    raw = ws.cell(row=row, column=column).value
    is_formula = isinstance(raw, str) and raw.startswith("=")
    has_literal_value = (raw is not None) and not is_formula
    return CellFormulaEvidence(
        is_formula=is_formula,
        formula_text=raw if is_formula else None,
        has_literal_value=has_literal_value,
    )
