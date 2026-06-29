"""
excel_export.py — Executive Financial Model™ v0.9
Pepperyn — Copilote Financier Exécutif

Le dirigeant ne doit jamais avoir l'impression d'ouvrir un classeur Excel.
Il doit avoir l'impression d'ouvrir un logiciel financier professionnel.

RÈGLE ABSOLUE : ce fichier ne contient que de la présentation et de l'ergonomie.
Aucune logique métier. Aucun calcul. Aucune formule modifiée.

Structure (9 feuilles) :
  0. 🏠 Accueil          — Porte d'entrée du modèle
  1. 📊 Dashboard         — KPI en temps réel
  2. ⚙ Hypothèses        — Paramètres modifiables (cellules bleues)
  3. 🎯 Decision Lab      — Simulateur de décisions
  4. 📈 Sensibilité       — Leviers d'influence
  5. 📉 Scénarios         — Prudent / Central / Ambitieux / Personnalisé
  6. 🗺 Roadmap           — Pilotage visuel des décisions
  7. 🕒 Historique        — Comparaison inter-périodes
  8. EDM (masquée)        — Source technique unique, ne pas modifier

Convention couleurs (standard modèles financiers) :
  Bleu  = saisie utilisateur modifiable
  Noir  = formule calculée automatiquement
  Vert  = lien inter-feuille
  Orange = alerte
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.excel_layout import (
    ROW_H_HEADER, ROW_H_SECTION, ROW_H_COL_HDR, ROW_H_NAV,
    ROW_H_INSTRUCTION, ROW_H_INSTRUCTION_WIDE, ROW_H_KPI_PROJECTED,
    ROW_H_GUIDE_NOTE,
    COL_END_EDM, COL_END_DASHBOARD, COL_END_HYPOTHESES, COL_END_DECISION_LAB,
    COL_END_SENSIBILITE, COL_END_SCENARIOS, COL_END_ROADMAP, COL_END_HISTORIQUE,
    COLS_EDM, COLS_DASHBOARD, COLS_HYPOTHESES, COLS_DECISION_LAB,
    COLS_SENSIBILITE, COLS_SCENARIOS, COLS_ROADMAP, COLS_HISTORIQUE,
)
from models.schemas import AnalysisResult
from models.executive_case import ExecutiveCaseJSON

# ─── Palette Pepperyn ────────────────────────────────────────────────────────
P_NAVY    = "0A2540"
P_NAVY2   = "071A2E"   # légèrement plus sombre pour le nav bar
P_BLUE    = "1B73E8"
P_DARK    = "1A1A2E"
P_GRAY    = "5F6368"
P_SLATE   = "94A3B8"
P_LINE    = "E2E8F0"
P_WHITE   = "FFFFFF"
P_LIGHT   = "F8FAFF"
P_GREEN   = "158038"
P_RED     = "DC2626"
P_AMBER   = "B45309"

# Convention de couleur standard (modèles financiers)
FONT_INPUT   = "0000FF"   # Bleu  = cellule saisie utilisateur
FONT_FORMULA = "000000"   # Noir  = formule automatique
FONT_LINK    = "008000"   # Vert  = lien inter-feuille

# ─── Noms des feuilles ───────────────────────────────────────────────────────
SN_ACCUEIL = "🏠 Accueil"
SN_DASH    = "📊 Dashboard"
SN_HYPO    = "⚙ Hypothèses"
SN_LAB     = "🎯 Decision Lab"
SN_SENSI   = "📈 Sensibilité"
SN_SCEN    = "📉 Scénarios"
SN_ROAD    = "🗺 Roadmap"
SN_HIST    = "🕒 Historique"
SN_EDM     = "EDM"

# Ordre de navigation (affiché dans le bandeau)
NAV_SHEETS = [SN_ACCUEIL, SN_DASH, SN_HYPO, SN_LAB, SN_SENSI, SN_SCEN, SN_ROAD, SN_HIST]

# Labels courts pour le bandeau de navigation
NAV_LABELS = {
    SN_ACCUEIL: "Accueil",
    SN_DASH:    "Dashboard",
    SN_HYPO:    "Hypothèses",
    SN_LAB:     "Decision Lab",
    SN_SENSI:   "Sensibilité",
    SN_SCEN:    "Scénarios",
    SN_ROAD:    "Roadmap",
    SN_HIST:    "Historique",
}

# ─── Lignes clés dans la feuille EDM ─────────────────────────────────────────
EDM_R_HEALTH    = 3
EDM_R_CONFID    = 4
EDM_R_EBITDA    = 5
EDM_R_CASH      = 6
EDM_R_COI_YEAR  = 7
EDM_R_COI_MONTH = 8
EDM_R_COI_WEEK  = 9
EDM_R_COI_DAY   = 10
EDM_R_COI_HOUR  = 11
EDM_R_IMPACT    = 12
EDM_R_DEC_START = 29    # Première décision : ligne 29, col B = annual_impact
MAX_DEC = 10

# ─── Lignes clés dans Decision Lab ───────────────────────────────────────────
LAB_R_DEC_START = 6
LAB_R_DEC_END   = LAB_R_DEC_START + MAX_DEC - 1   # = 15
LAB_R_TOTAL     = LAB_R_DEC_END + 1               # = 16
LAB_R_EBITDA    = LAB_R_TOTAL + 4                 # = 20
LAB_R_IMPACT    = LAB_R_TOTAL + 5                 # = 21
LAB_R_PROJ      = LAB_R_TOTAL + 6                 # = 22
LAB_R_VAR_PCT   = LAB_R_TOTAL + 7                 # = 23
LAB_COL_IMPACT  = 5                               # Colonne E = impact projeté

# ─── Lignes clés dans Hypothèses ─────────────────────────────────────────────
HYPO_R_EBITDA  = 5
HYPO_R_CASH    = 6
HYPO_R_COI     = 7
HYPO_R_CA      = 11
HYPO_R_CROISS  = 12
HYPO_R_PRIX    = 13
HYPO_R_VOLUME  = 14
HYPO_R_MARGE_B = 15
HYPO_R_MASA    = 18
HYPO_R_CHARGES = 19
HYPO_R_SOUSTRX = 20
HYPO_R_INFLAT  = 21
HYPO_R_MARGE_N = 22


# ─── Utilitaires de formatage ────────────────────────────────────────────────

def _parse_eur(s: Optional[str]) -> float:
    """'1 800 000 €' → 1_800_000.0  /  '1.2M €' → 1_200_000.0"""
    if not s:
        return 0.0
    c = re.sub(r"[€\s]", "", str(s)).replace(",", ".").strip()
    try:
        if c.upper().endswith("M"):
            return float(c[:-1]) * 1_000_000
        if c.upper().endswith("K"):
            return float(c[:-1]) * 1_000
        return float(c)
    except (ValueError, TypeError):
        return 0.0


# ─── Primitives de style ──────────────────────────────────────────────────────

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _font(size: int = 10, bold: bool = False, color: str = P_DARK,
          italic: bool = False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=0)

def _align_indent(h: str = "left", v: str = "center",
                  wrap: bool = False, indent: int = 1) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _border_thin() -> Border:
    s = Side(style="thin", color=P_LINE)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_accent() -> Border:
    return Border(left=Side(style="medium", color=P_BLUE),
                  right=Side(style="thin", color=P_LINE),
                  top=Side(style="thin", color=P_LINE),
                  bottom=Side(style="thin", color=P_LINE))

def _header_bar(ws, row: int, text: str, col_start: int = 1,
                col_end: int = 18, size: int = 13, bg: str = P_NAVY) -> None:
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=f"  {text}")
    c.font = _font(size=size, bold=True, color=P_WHITE)
    c.fill = _fill(bg)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = ROW_H_HEADER

def _section_bar(ws, row: int, text: str, col_start: int = 1,
                 col_end: int = 18, bg: str = P_BLUE) -> None:
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=f"  {text}")
    c.font = _font(size=9, bold=True, color=P_WHITE)
    c.fill = _fill(bg)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = ROW_H_SECTION

def _col_headers(ws, row: int, headers: list, start_col: int = 1,
                 bg: str = P_NAVY) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = _font(size=9, bold=True, color=P_WHITE)
        c.fill = _fill(bg)
        c.alignment = _align("center")
        c.border = _border_thin()
    ws.row_dimensions[row].height = ROW_H_COL_HDR

def _instruction(ws, row: int, text: str, col: int = 1, col_end: int = 18,
                 height: float = ROW_H_INSTRUCTION) -> None:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col, value=f"  ℹ  {text}")
    c.font = _font(size=9, italic=True, color=P_SLATE)
    c.fill = _fill("FFFBEB")
    c.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[row].height = height

def _kpi_block(ws, row: int, col: int, label: str, value, *,
               width: int = 3, height: int = 3, fill_c: str = P_LIGHT,
               val_size: int = 16, val_color: str = P_DARK,
               num_format: str = None) -> None:
    end_col = col + width - 1
    for r in range(row, row + height):
        for ci in range(col, end_col + 1):
            ws.cell(row=r, column=ci).fill = _fill(fill_c)
    for ci in range(col, end_col + 1):
        ws.cell(row=row, column=ci).border = Border(
            top=Side(style="medium", color=P_BLUE))
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=end_col)
    lc = ws.cell(row=row, column=col, value=label.upper())
    lc.font = _font(size=8, color=P_SLATE, bold=True)
    lc.alignment = _align("left")
    ws.merge_cells(start_row=row + 1, start_column=col,
                   end_row=row + height - 1, end_column=end_col)
    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = _font(size=val_size, bold=True, color=val_color)
    vc.alignment = _align("left", "center")
    if num_format:
        vc.number_format = num_format

def _widths(ws, mapping: dict) -> None:
    for k, w in mapping.items():
        col = get_column_letter(k) if isinstance(k, int) else k
        ws.column_dimensions[col].width = w

def _freeze(ws, cell: str = "A4") -> None:
    ws.freeze_panes = cell


def _row_height_for_text(text: str, col_width: float,
                         font_size: int = 10,
                         min_height: float = 30.0,
                         padding: float = 8.0) -> float:
    """
    PEPPERYN_EXECUTIVE_RENDERING_SPECIFICATION — Rows:
    'Dynamic height. Renderer computes row height. Never fixed.'

    Compute the row height in points required to display `text` fully
    when wrapped in a column of `col_width` character units (openpyxl unit).

    At Calibri 10pt, one openpyxl character unit ≈ 0.85 visible characters.
    One line height ≈ font_size × 1.5 (standard leading).
    A safety factor of 0.85 ensures text never clips at the boundary.
    """
    if not text:
        return min_height
    chars_per_line = max(int(col_width * 0.85), 1)
    num_lines = -(-len(str(text)) // chars_per_line)   # ceiling division
    height = num_lines * font_size * 1.5 + padding
    return max(height, min_height)


# ─── Helpers UX ──────────────────────────────────────────────────────────────

def _setup_sheet(ws) -> None:
    """Masque quadrillage et en-têtes ligne/colonne : l'utilisateur voit une application."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

def _nav_bar(ws, active: str, row: int = 3, col_end: int = 18) -> None:
    """Bandeau de navigation breadcrumb — feuille active entre crochets."""
    parts = []
    for sn in NAV_SHEETS:
        label = NAV_LABELS.get(sn, sn)
        if sn == active:
            parts.append(f"[ {label} ]")
        else:
            parts.append(label)
    nav_text = "   ›   ".join(parts)
    prefix = "  Vous êtes ici :   "

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=prefix + nav_text)
    c.font = _font(size=7, color="8899BB")
    c.fill = _fill(P_NAVY2)
    c.alignment = _align("left", "center")
    ws.row_dimensions[row].height = ROW_H_NAV


# ─── FEUILLE EDM (masquée — source unique) ────────────────────────────────────

def _build_edm(wb: Workbook, edm, raw: dict) -> None:
    ws = wb.create_sheet(SN_EDM)
    ws.tab_color = P_SLATE
    _setup_sheet(ws)

    _header_bar(ws, 1, "EXECUTIVE DECISION MODEL — Source technique (ne jamais modifier)", col_end=8)
    ws.cell(row=2, column=1,
            value="Ce feuillet alimente l'ensemble du classeur. Toutes ses valeurs proviennent de l'EDM Pepperyn."
            ).font = _font(italic=True, color=P_SLATE)

    coi = edm.cost_of_inaction

    # ── Extraire EBITDA et Cash depuis le kpi_dashboard (raw dict) ────────────
    # edm.ebitda / edm.available_cash sont None dans le pipeline V2 :
    # case_to_edm() ne les propage pas (champs absents du modèle EDM legacy).
    # RULE 004 — le renderer AFFICHE. Il lit depuis la source : raw.ceo_dashboard.
    # RULE 001 — aucune valeur ne peut rester à 0 si la donnée est disponible.
    _ebitda_val = 0.0
    _cash_val   = 0.0
    for _card in ((raw or {}).get("ceo_dashboard") or []):
        _lbl = (_card.get("label", "") if isinstance(_card, dict) else "").lower()
        _val = (_card.get("value", "") if isinstance(_card, dict) else "")
        if "ebitda" in _lbl:
            _ebitda_val = _parse_eur(_val)
        elif any(k in _lbl for k in ("trésor", "tresor", "cash", "liquid")):
            _cash_val = _parse_eur(_val)

    global_rows = [
        (EDM_R_HEALTH,    "Score Santé",                edm.health_score or 0),
        (EDM_R_CONFID,    "Niveau de confiance",        edm.executive_confidence or 0),
        (EDM_R_EBITDA,    "EBITDA (numérique €)",       _ebitda_val),
        (EDM_R_CASH,      "Cash disponible (€)",        _cash_val),
        (EDM_R_COI_YEAR,  "Coût de l'inaction / an",    (coi.per_year or 0) if coi else 0),
        (EDM_R_COI_MONTH, "Coût de l'inaction / mois",  (coi.per_month or 0) if coi else 0),
        (EDM_R_COI_WEEK,  "Coût de l'inaction / sem.",  (coi.per_week or 0) if coi else 0),
        (EDM_R_COI_DAY,   "Coût de l'inaction / jour",  (coi.per_day or 0) if coi else 0),
        (EDM_R_COI_HOUR,  "Coût de l'inaction / heure", (coi.per_hour or 0) if coi else 0),
        (EDM_R_IMPACT,    "Impact total identifié",
         -sum(abs(d.annual_impact or 0) for d in edm.value_destroyers)),
    ]
    for r, label, val in global_rows:
        ws.cell(row=r, column=1, value=label).font = _font(color=P_GRAY)
        ws.cell(row=r, column=2, value=val)

    # Value destroyers
    dh = 15
    _col_headers(ws, dh, ["Levier de destruction", "Impact annuel (€)", "Impact mensuel (€)", "Tendance", "Commentaire"])
    for i, d in enumerate(edm.value_destroyers[:MAX_DEC]):
        r = dh + 1 + i
        ws.cell(row=r, column=1, value=d.name)
        ws.cell(row=r, column=2, value=d.annual_impact or 0)
        ws.cell(row=r, column=3, value=d.monthly_impact or 0)
        ws.cell(row=r, column=4, value=d.trend or "stable")
        ws.cell(row=r, column=5, value=(d.comment or "")[:100])

    # Executive decisions — col B = annual_impact (référencé par Decision Lab)
    dech = EDM_R_DEC_START - 1
    _col_headers(ws, dech, ["Décision", "Impact annuel (€)", "Difficulté",
                             "Horizon", "Responsable", "Priorité", "ROI score"])
    for i, dec in enumerate(edm.executive_decisions[:MAX_DEC]):
        r = EDM_R_DEC_START + i
        ws.cell(row=r, column=1, value=dec.decision)
        ws.cell(row=r, column=2, value=dec.annual_impact or 0)
        ws.cell(row=r, column=3, value=dec.difficulty or "—")
        ws.cell(row=r, column=4, value=dec.timeline or "—")
        ws.cell(row=r, column=5, value=dec.owner or "—")
        ws.cell(row=r, column=6, value=dec.priority)
        ws.cell(row=r, column=7, value=dec.roi_score)

    # CEO dashboard cards
    raw_cards = (raw or {}).get("ceo_dashboard") or []
    r = 42
    ws.cell(row=r - 1, column=1, value="CEO Dashboard brut").font = _font(bold=True, color=P_NAVY)
    for i, card in enumerate(raw_cards[:8]):
        if isinstance(card, dict):
            ws.cell(row=r + i, column=1, value=card.get("label", ""))
            ws.cell(row=r + i, column=2, value=card.get("value", "—"))

    # Séries temporelles (12 mois)
    r = 53
    ws.cell(row=r - 1, column=1, value="Séries temporelles").font = _font(bold=True, color=P_NAVY)
    _col_headers(ws, r, ["Série"] + [f"M{j+1}" for j in range(12)])
    ws.cell(row=r + 1, column=1, value="Action (avec décisions)")
    ws.cell(row=r + 2, column=1, value="Inaction (sans décision)")
    ws.cell(row=r + 3, column=1, value="Trajectoire cible")
    for j, v in enumerate(edm.action_series[:12]):
        ws.cell(row=r + 1, column=2 + j, value=v)
    for j, v in enumerate(edm.do_nothing_series[:12]):
        ws.cell(row=r + 2, column=2 + j, value=v)
    for j, v in enumerate(edm.monthly_projection[:12]):
        ws.cell(row=r + 3, column=2 + j, value=v)

    _widths(ws, COLS_EDM)


# ─── FEUILLE 0 : ACCUEIL ─────────────────────────────────────────────────────

def _build_accueil(wb: Workbook) -> None:
    ws = wb.create_sheet(SN_ACCUEIL)
    ws.tab_color = P_NAVY
    _setup_sheet(ws)

    COLS = 18

    # ── Row 1 : bandeau de marque ──
    _header_bar(ws, 1,
                "PEPPERYN  ·  EXECUTIVE FINANCIAL MODEL™  ·  CONFIDENTIEL",
                col_end=COLS, size=12)

    # ── Row 2 : navigation (Accueil actif) ──
    _nav_bar(ws, SN_ACCUEIL, row=2, col_end=COLS)

    # ── Row 3 : espace ──
    ws.row_dimensions[3].height = 28

    # ── Rows 4-6 : titre principal ──
    ws.merge_cells(start_row=4, start_column=2, end_row=6, end_column=COLS - 1)
    c = ws.cell(row=4, column=2, value="Executive Financial Model™")
    c.font = Font(name="Calibri", size=26, bold=True, color=P_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="bottom")
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20

    # ── Row 7-8 : sous-titre ──
    ws.merge_cells(start_row=7, start_column=2, end_row=8, end_column=COLS - 1)
    c = ws.cell(row=7, column=2,
                value="Votre environnement interactif d'aide à la décision.")
    c.font = Font(name="Calibri", size=13, italic=True, color=P_GRAY)
    c.alignment = Alignment(horizontal="left", vertical="top")
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 16

    # Séparateur fin
    for ci in range(2, COLS):
        cell = ws.cell(row=8, column=ci)
        cell.border = Border(bottom=Side(style="medium", color=P_BLUE))

    ws.row_dimensions[9].height = 16

    # ── BLOC 1 : À quoi sert ce modèle ? ──
    _section_bar(ws, 10, "À QUOI SERT CE MODÈLE ?",
                 col_start=2, col_end=COLS - 1)

    b1_text = (
        "Ce modèle vous permet de tester différents scénarios financiers "
        "avant de prendre une décision.\n"
        "Toutes les simulations reposent directement sur l'analyse réalisée par Pepperyn."
    )
    ws.merge_cells(start_row=11, start_column=2, end_row=13, end_column=COLS - 1)
    c = ws.cell(row=11, column=2, value=b1_text)
    c.font = _font(size=11, color=P_DARK)
    c.fill = _fill(P_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[11].height = 20
    ws.row_dimensions[12].height = 20
    ws.row_dimensions[13].height = 20

    ws.row_dimensions[14].height = 14

    # ── BLOC 2 : Comment utiliser ce modèle ? ──
    _section_bar(ws, 15, "COMMENT UTILISER CE MODÈLE ?",
                 col_start=2, col_end=COLS - 1)

    steps = [
        "1.   Lire l'Executive Report™ afin de comprendre la situation.",
        "2.   Présenter l'Executive Board Deck™ afin d'aligner les décideurs.",
        "3.   Explorer différents scénarios dans ce modèle.",
        "4.   Modifier uniquement les cellules bleues.",
        "5.   Comparer les impacts et choisir les meilleures décisions.",
    ]

    r = 16
    for i, step in enumerate(steps):
        # Ligne étape
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=COLS - 1)
        c = ws.cell(row=r, column=2, value=f"   {step}")
        is_this_model = (i in (2, 3, 4))   # étapes 3, 4, 5 = cœur du modèle
        c.font = _font(size=11, bold=is_this_model,
                       color=P_NAVY if is_this_model else P_DARK)
        c.fill = _fill("EFF6FF" if is_this_model else P_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        # Flèche (sauf après la dernière)
        if i < len(steps) - 1:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=COLS - 1)
            c = ws.cell(row=r, column=2, value="              ↓")
            c.font = _font(size=10, color=P_SLATE)
            c.fill = _fill(P_WHITE)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 14
            r += 1

    ws.row_dimensions[r].height = 14
    r += 1

    # ── BLOC 3 : Règles d'utilisation ──
    _section_bar(ws, r, "RÈGLES D'UTILISATION",
                 col_start=2, col_end=COLS - 1)
    r += 1

    rules = [
        ("●   Les cellules bleues sont modifiables.",      "EFF6FF",  FONT_INPUT),
        ("●   Toutes les autres cellules sont calculées automatiquement.", P_LIGHT, P_DARK),
        ("●   Chaque modification met instantanément à jour les projections.", P_LIGHT, P_DARK),
    ]
    for rule_text, bg, fc in rules:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=COLS - 1)
        c = ws.cell(row=r, column=2, value=f"   {rule_text}")
        c.font = _font(size=11, color=fc, bold=(fc == FONT_INPUT))
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

    ws.row_dimensions[r].height = 24
    r += 1

    # ── Bouton ▶ Commencer ──
    btn_row = r
    btn_col_start = 5
    btn_col_end = 14

    ws.merge_cells(start_row=btn_row, start_column=btn_col_start,
                   end_row=btn_row + 1, end_column=btn_col_end)
    c = ws.cell(row=btn_row, column=btn_col_start,
                value="▶   Commencer  →  Dashboard")
    c.font = Font(name="Calibri", size=15, bold=True, color=P_WHITE)
    c.fill = _fill(P_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.hyperlink = f"#'{SN_DASH}'!A1"
    ws.row_dimensions[btn_row].height = 26
    ws.row_dimensions[btn_row + 1].height = 26

    # Ombre portée (cellules sous le bouton)
    shadow_row = btn_row + 2
    ws.merge_cells(start_row=shadow_row, start_column=btn_col_start + 1,
                   end_row=shadow_row, end_column=btn_col_end + 1)
    for ci in range(btn_col_start + 1, btn_col_end + 2):
        ws.cell(row=shadow_row, column=ci).fill = _fill("C8D8F0")
    ws.row_dimensions[shadow_row].height = 4

    ws.row_dimensions[shadow_row + 1].height = 20

    # Légende du bouton
    ws.merge_cells(start_row=shadow_row + 2, start_column=btn_col_start,
                   end_row=shadow_row + 2, end_column=btn_col_end)
    c = ws.cell(row=shadow_row + 2, column=btn_col_start,
                value="Cliquez pour accéder au tableau de bord")
    c.font = _font(size=9, italic=True, color=P_SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[shadow_row + 2].height = 16

    # Pied de page discret
    r_footer = shadow_row + 4
    ws.merge_cells(start_row=r_footer, start_column=1,
                   end_row=r_footer, end_column=COLS)
    c = ws.cell(row=r_footer, column=1,
                value="  Pepperyn Executive Financial Model™  ·  Document confidentiel  ·  Généré automatiquement depuis votre analyse Pepperyn")
    c.font = _font(size=8, italic=True, color=P_SLATE)
    c.fill = _fill(P_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r_footer].height = 20

    # Largeurs colonnes Accueil
    _widths(ws, {"A": 2})
    for col_i in range(2, COLS):
        ws.column_dimensions[get_column_letter(col_i)].width = 10
    ws.column_dimensions[get_column_letter(COLS)].width = 2


# ─── FEUILLE 1 : DASHBOARD ───────────────────────────────────────────────────

def _build_dashboard(wb: Workbook, edm, raw: dict) -> None:
    ws = wb.create_sheet(SN_DASH)
    ws.tab_color = P_NAVY
    _setup_sheet(ws)

    date_str = datetime.now().strftime("%d %B %Y")
    _header_bar(ws, 1, f"EXECUTIVE DASHBOARD  ·  {date_str}  ·  CONFIDENTIEL", col_end=17)
    _instruction(ws, 2,
                 "Tableau de bord en temps réel — toutes les valeurs se mettent à jour automatiquement "
                 "lorsque vous modifiez vos hypothèses ou simulez des décisions dans Decision Lab.",
                 col_end=17, height=ROW_H_INSTRUCTION_WIDE)
    _nav_bar(ws, SN_DASH, row=3, col_end=17)

    # Ligne 1 de KPI
    kpi_cols = [2, 6, 10, 14]
    kpi_w = 3
    # Colonnes espaceurs (A=1, E=5, I=9, M=13, Q=17) — doivent partager le fond
    # du bloc pour former un rectangle homogène entre les barres.
    spacer_cols = [1, 5, 9, 13, 17]

    kpi_r1 = [
        ("EBITDA actuel",       f"=EDM!B{EDM_R_EBITDA}",  P_DARK,  "#,##0 €;(#,##0 €);-"),
        ("Cash disponible",     f"=EDM!B{EDM_R_CASH}",    P_DARK,  "#,##0 €;(#,##0 €);-"),
        ("Score santé",         f"=EDM!B{EDM_R_HEALTH}",  P_DARK,  "0\"/10\""),
        ("Niveau de confiance", f"=EDM!B{EDM_R_CONFID}",  P_DARK,  "0\"/10\""),
    ]
    r1_start = 4
    # Fond homogène sur toute la largeur avant de poser les blocs KPI
    for sc in spacer_cols:
        for r in range(r1_start, r1_start + 4):
            ws.cell(row=r, column=sc).fill = _fill(P_LIGHT)
    for (label, val, vc, nf), col in zip(kpi_r1, kpi_cols):
        _kpi_block(ws, r1_start, col, label, val, width=kpi_w, height=4,
                   val_size=18, val_color=vc, num_format=nf)

    kpi_r2 = [
        ("Coût inaction / an",  f"=EDM!B{EDM_R_COI_YEAR}",            P_RED,   "#,##0 €;(#,##0 €);-"),
        ("Impact identifié",    f"=EDM!B{EDM_R_IMPACT}",               P_AMBER, "#,##0 €;(#,##0 €);-"),
        ("EBITDA projeté",      f"='{SN_LAB}'!C{LAB_R_PROJ}",          P_GREEN, "#,##0 €;(#,##0 €);-"),
        ("Variation projetée",  f"='{SN_LAB}'!C{LAB_R_VAR_PCT}",       P_BLUE,  "0.0%"),
    ]
    r2_start = r1_start + 5
    for sc in spacer_cols:
        for r in range(r2_start, r2_start + 4):
            ws.cell(row=r, column=sc).fill = _fill(P_LIGHT)
    for (label, val, vc, nf), col in zip(kpi_r2, kpi_cols):
        _kpi_block(ws, r2_start, col, label, val, width=kpi_w, height=4,
                   val_size=18, val_color=vc, num_format=nf)

    # Coût de l'inaction détaillé
    r_coi_hdr = r2_start + 5
    ws.row_dimensions[r_coi_hdr].height = 8
    _section_bar(ws, r_coi_hdr + 1, "COÛT DE L'INACTION — Chaque heure compte", col_end=17)
    coi_items = [
        ("Par heure",   f"=EDM!B{EDM_R_COI_HOUR}", 2),
        ("Par jour",    f"=EDM!B{EDM_R_COI_DAY}",  6),
        ("Par semaine", f"=EDM!B{EDM_R_COI_WEEK}", 10),
        ("Par mois",    f"=EDM!B{EDM_R_COI_MONTH}",14),
    ]
    cr = r_coi_hdr + 2
    for sc in spacer_cols:
        for r in range(cr, cr + 3):
            ws.cell(row=r, column=sc).fill = _fill("FFF8F8")
    for label, val, col in coi_items:
        _kpi_block(ws, cr, col, label, val, width=kpi_w, height=3,
                   fill_c="FFF8F8", val_size=14, val_color=P_RED,
                   num_format="#,##0 €;(#,##0 €);-")

    # Navigation rapide
    nav_r = cr + 4
    ws.row_dimensions[nav_r].height = 8
    _section_bar(ws, nav_r + 1, "NAVIGATION RAPIDE", col_end=17)
    nav_items = [
        (2,  "→  Hypothèses",    "Modifier les paramètres"),
        (6,  "→  Decision Lab",  "Simuler des décisions"),
        (10, "→  Sensibilité",   "Identifier les leviers"),
        (14, "→  Scénarios",     "Comparer les visions"),
    ]
    # Fond homogène sur toute la largeur des 2 lignes de nav
    for ci in range(1, 18):
        ws.cell(row=nav_r + 2, column=ci).fill = _fill(P_LIGHT)
        ws.cell(row=nav_r + 3, column=ci).fill = _fill(P_LIGHT)
    for col, name, desc in nav_items:
        c = ws.cell(row=nav_r + 2, column=col, value=name)
        c.font = _font(size=9, color=P_BLUE, bold=True)
        c.fill = _fill(P_LIGHT)
        ws.merge_cells(start_row=nav_r + 2, start_column=col,
                       end_row=nav_r + 2, end_column=col + kpi_w - 1)
        c2 = ws.cell(row=nav_r + 3, column=col, value=desc)
        c2.font = _font(size=8, color=P_GRAY)
        c2.fill = _fill(P_LIGHT)
        ws.merge_cells(start_row=nav_r + 3, start_column=col,
                       end_row=nav_r + 3, end_column=col + kpi_w - 1)

    _widths(ws, COLS_DASHBOARD)


# ─── FEUILLE 2 : HYPOTHÈSES ───────────────────────────────────────────────────

def _build_hypotheses(wb: Workbook, edm) -> None:
    ws = wb.create_sheet(SN_HYPO)
    ws.tab_color = P_BLUE
    _setup_sheet(ws)

    _header_bar(ws, 1, "HYPOTHÈSES — Paramètres de votre modèle financier", col_end=4)
    _instruction(ws, 2,
                 "Cellules BLEUES = vous pouvez les modifier. "
                 "Toutes les autres feuilles se recalculent immédiatement.",
                 col_end=4)
    _nav_bar(ws, SN_HYPO, row=3, col_end=4)
    _col_headers(ws, 4, ["Paramètre", "Valeur", "Unité", "Note / Source"], bg=P_NAVY)

    # Section A — Données de référence EDM
    _section_bar(ws, HYPO_R_EBITDA - 1,
                 "A — DONNÉES DE RÉFÉRENCE (issues de l'EDM — liens automatiques)",
                 col_end=4)
    ref_rows = [
        (HYPO_R_EBITDA, "EBITDA de référence",     f"=EDM!B{EDM_R_EBITDA}",   "€",    "Lien EDM Pepperyn"),
        (HYPO_R_CASH,   "Cash disponible",          f"=EDM!B{EDM_R_CASH}",     "€",    "Lien EDM Pepperyn"),
        (HYPO_R_COI,    "Coût de l'inaction / an", f"=EDM!B{EDM_R_COI_YEAR}", "€/an", "Lien EDM Pepperyn"),
    ]
    for r, label, val, unit, note in ref_rows:
        ws.cell(row=r, column=1, value=label).font = _font(color=P_GRAY)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _font(color=FONT_LINK, bold=True)
        c.number_format = "#,##0 €;(#,##0 €);-"
        c.border = _border_thin()
        ws.cell(row=r, column=3, value=unit).font = _font(color=P_SLATE)
        ws.cell(row=r, column=4, value=note).font = _font(italic=True, color=P_SLATE)
    ws.row_dimensions[HYPO_R_COI + 1].height = 8

    # Section B — Hypothèses de revenus
    ebitda_num = _parse_eur(edm.ebitda)
    _section_bar(ws, HYPO_R_CA - 1,
                 "B — HYPOTHÈSES DE REVENUS  ↓ cellules bleues modifiables",
                 col_end=4)
    rev_rows = [
        (HYPO_R_CA,     "Chiffre d'affaires estimé", max(ebitda_num * 2.5, 0), "#,##0 €;(#,##0 €);-", "Votre estimation CA annuel"),
        (HYPO_R_CROISS, "Taux de croissance annuel",  0.05,                     "0.0%",                "Objectif de croissance"),
        (HYPO_R_PRIX,   "Prix moyen de vente (€)",    0.0,                      "#,##0 €;(#,##0 €);-", "Prix unitaire moyen"),
        (HYPO_R_VOLUME, "Volume mensuel (unités)",     0.0,                      "#,##0",               "Nombre d'unités / mois"),
        (HYPO_R_MARGE_B,"Marge brute (%)",             0.40,                     "0.0%",                "Marge avant charges de structure"),
    ]
    for r, label, val, nf, note in rev_rows:
        ws.cell(row=r, column=1, value=label).font = _font(color=P_DARK)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _font(color=FONT_INPUT, bold=True)
        c.fill = _fill("EFF6FF")
        c.number_format = nf
        c.border = _border_thin()
        ws.cell(row=r, column=4, value=note).font = _font(italic=True, color=P_SLATE)
    ws.row_dimensions[HYPO_R_MARGE_B + 1].height = 8

    # Section C — Hypothèses de coûts
    masa_init  = abs(next((d.annual_impact for d in edm.value_destroyers
                           if "salarial" in (d.name or "").lower()), None) or 0)
    charg_init = abs(next((d.annual_impact for d in edm.value_destroyers
                           if "charge" in (d.name or "").lower()), None) or 0)
    sous_init  = abs(next((d.annual_impact for d in edm.value_destroyers
                           if "sous" in (d.name or "").lower()), None) or 0)
    _section_bar(ws, HYPO_R_MASA - 1,
                 "C — HYPOTHÈSES DE COÛTS  ↓ cellules bleues modifiables",
                 col_end=4)
    cost_rows = [
        (HYPO_R_MASA,    "Masse salariale annuelle", masa_init,  "#,##0 €;(#,##0 €);-", "Initialisation EDM"),
        (HYPO_R_CHARGES, "Charges fixes annuelles",  charg_init, "#,##0 €;(#,##0 €);-", "Initialisation EDM"),
        (HYPO_R_SOUSTRX, "Coût de sous-traitance",   sous_init,  "#,##0 €;(#,##0 €);-", "Initialisation EDM"),
        (HYPO_R_INFLAT,  "Taux d'inflation (%)",     0.025,      "0.0%",                "Hypothèse macro"),
        (HYPO_R_MARGE_N, "Marge nette cible (%)",    0.10,       "0.0%",                "Objectif de profitabilité"),
    ]
    for r, label, val, nf, note in cost_rows:
        ws.cell(row=r, column=1, value=label).font = _font(color=P_DARK)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _font(color=FONT_INPUT, bold=True)
        c.fill = _fill("EFF6FF")
        c.number_format = nf
        c.border = _border_thin()
        ws.cell(row=r, column=4, value=note).font = _font(italic=True, color=P_SLATE)
    ws.row_dimensions[HYPO_R_MARGE_N + 1].height = 8

    # Section D — Paramètres de simulation
    sim_base = HYPO_R_MARGE_N + 2
    _section_bar(ws, sim_base, "D — PARAMÈTRES DE SIMULATION", col_end=4)
    sim_rows = [
        (sim_base + 1, "Horizon de projection (mois)",         12,    "#,##0",  "Période couverte par la simulation"),
        (sim_base + 2, "Probabilité de réussite des décisions", 0.80, "0.0%",   "Taux de succès moyen estimé"),
        (sim_base + 3, "Taux d'actualisation / WACC",          0.08,  "0.0%",   "Coût d'opportunité du capital"),
    ]
    for r, label, val, nf, note in sim_rows:
        ws.cell(row=r, column=1, value=label).font = _font(color=P_DARK)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _font(color=FONT_INPUT, bold=True)
        c.fill = _fill("EFF6FF")
        c.number_format = nf
        c.border = _border_thin()
        ws.cell(row=r, column=4, value=note).font = _font(italic=True, color=P_SLATE)

    _widths(ws, COLS_HYPOTHESES)


# ─── FEUILLE 3 : DECISION LAB ─────────────────────────────────────────────────

def _build_decision_lab(wb: Workbook, edm, result_dict: dict | None = None) -> None:
    ws = wb.create_sheet(SN_LAB)
    ws.tab_color = P_GREEN
    _setup_sheet(ws)

    _header_bar(ws, 1, "DECISION LAB — Simulateur stratégique", col_end=8)
    _instruction(ws, 2,
                 "Modifiez la colonne EXÉCUTION % (cellules bleues) pour simuler vos scénarios. "
                 "Tous les indicateurs se recalculent instantanément. "
                 "0% = décision non prise. 100% = décision totalement exécutée.",
                 col_end=8)
    _nav_bar(ws, SN_LAB, row=3, col_end=8)
    _section_bar(ws, 4, "SIMULATION — Ajustez le % d'exécution pour chaque décision", col_end=8)

    headers = ["#", "Décision", "Impact de base (€)", "Exécution %",
               "Impact projeté (€)", "Responsable", "Horizon", "Priorité"]
    _col_headers(ws, 5, headers, bg=P_NAVY)

    decisions = edm.executive_decisions[:MAX_DEC]
    lc = get_column_letter(LAB_COL_IMPACT)

    for i in range(MAX_DEC):
        r = LAB_R_DEC_START + i
        alt = P_LIGHT if i % 2 == 0 else P_WHITE
        dec = decisions[i] if i < len(decisions) else None

        c = ws.cell(row=r, column=1, value=i + 1)
        c.font = _font(color=P_SLATE, bold=True)
        c.fill = _fill(alt)
        c.border = _border_thin()
        c.alignment = _align("center")

        dec_text = dec.decision if dec else "[Décision disponible]"
        c = ws.cell(row=r, column=2, value=dec_text)
        c.font = _font(color=P_DARK if dec else P_SLATE,
                       italic=not dec, bold=bool(dec))
        c.fill = _fill(alt)
        c.border = _border_thin()
        c.alignment = _align("left", wrap=True)
        # SPEC: Rows — Dynamic height. Renderer computes row height. Never fixed.
        ws.row_dimensions[r].height = _row_height_for_text(dec_text, 90)

        edm_ref = f"EDM!B{EDM_R_DEC_START + i}"
        c = ws.cell(row=r, column=3, value=f"={edm_ref}")
        c.font = _font(color=FONT_LINK, bold=True)
        c.fill = _fill(alt)
        c.border = _border_thin()
        c.number_format = "#,##0 €;(#,##0 €);-"
        c.alignment = _align("right")

        exec_val = 1.0 if dec else 0.0
        c = ws.cell(row=r, column=4, value=exec_val)
        c.font = _font(color=FONT_INPUT, bold=True)
        c.fill = _fill("DBEAFE")
        c.border = _border_accent()
        c.number_format = "0%"
        c.alignment = _align("center")

        c = ws.cell(row=r, column=LAB_COL_IMPACT, value=f"=C{r}*D{r}")
        c.font = _font(color=FONT_FORMULA, bold=True)
        c.fill = _fill(alt)
        c.border = _border_thin()
        c.number_format = "#,##0 €;(#,##0 €);-"
        c.alignment = _align("right")

        c = ws.cell(row=r, column=6, value=dec.owner if dec else "—")
        c.font = _font(color=FONT_INPUT)
        c.fill = _fill(alt)
        c.border = _border_thin()
        c.alignment = _align("left", wrap=True)   # SPEC: Left for descriptions, Wrap mandatory

        c = ws.cell(row=r, column=7, value=dec.timeline if dec else "—")
        c.font = _font(color=FONT_INPUT)
        c.fill = _fill(alt)
        c.border = _border_thin()
        c.alignment = _align("left", wrap=True)   # SPEC: Left for descriptions, Wrap mandatory

        if dec:
            p_lower = (dec.priority or "").lower()
            p_color = P_RED if "high" in p_lower else (P_AMBER if "medium" in p_lower else P_GRAY)
            c = ws.cell(row=r, column=8, value=dec.priority)
            c.font = _font(color=p_color, bold=True)
        else:
            ws.cell(row=r, column=8, value="—").font = _font(color=P_SLATE)
        ws.cell(row=r, column=8).fill = _fill(alt)
        ws.cell(row=r, column=8).border = _border_thin()
        ws.cell(row=r, column=8).alignment = _align("center")

    # Ligne totaux
    r_tot = LAB_R_TOTAL
    ws.merge_cells(start_row=r_tot, start_column=1, end_row=r_tot, end_column=2)
    c = ws.cell(row=r_tot, column=1, value="TOTAL IMPACT SIMULÉ")
    c.font = _font(size=10, bold=True, color=P_WHITE)
    c.fill = _fill(P_NAVY)
    c.alignment = _align("right")

    c3 = ws.cell(row=r_tot, column=3, value=f"=SUM(C{LAB_R_DEC_START}:C{LAB_R_DEC_END})")
    c3.font = _font(bold=True, color=P_WHITE)
    c3.fill = _fill(P_NAVY)
    c3.number_format = "#,##0 €;(#,##0 €);-"
    c3.alignment = _align("right")

    c5 = ws.cell(row=r_tot, column=LAB_COL_IMPACT,
                 value=f"=SUM({lc}{LAB_R_DEC_START}:{lc}{LAB_R_DEC_END})")
    c5.font = _font(bold=True, color=P_WHITE)
    c5.fill = _fill(P_NAVY)
    c5.number_format = "#,##0 €;(#,##0 €);-"
    c5.alignment = _align("right")

    # Tableau de bord projeté
    ws.row_dimensions[r_tot + 1].height = 10
    _section_bar(ws, r_tot + 2, "TABLEAU DE BORD PROJETÉ — Recalcul automatique", col_end=8)

    proj_def = [
        (LAB_R_EBITDA,   "EBITDA de base (€)",
         f"='{SN_HYPO}'!B{HYPO_R_EBITDA}", "#,##0 €;(#,##0 €);-", FONT_LINK),
        (LAB_R_IMPACT,   "Impact simulé total (€)",
         f"={lc}{r_tot}", "#,##0 €;(#,##0 €);-", FONT_FORMULA),
        (LAB_R_PROJ,     "EBITDA projeté (€)",
         f"=C{LAB_R_EBITDA}+C{LAB_R_IMPACT}", "#,##0 €;(#,##0 €);-", FONT_FORMULA),
        (LAB_R_VAR_PCT,  "Variation vs actuel (%)",
         f"=IFERROR(C{LAB_R_IMPACT}/ABS(C{LAB_R_EBITDA}),0)", "0.0%", FONT_FORMULA),
        (LAB_R_VAR_PCT+1,"Coût de l'inaction (rappel)",
         f"=EDM!B{EDM_R_COI_YEAR}", "#,##0 €;(#,##0 €);-", FONT_LINK),
        (LAB_R_VAR_PCT+2,"Gain net vs inaction",
         f"=C{LAB_R_PROJ}+ABS(EDM!B{EDM_R_COI_YEAR})",
         "#,##0 €;(#,##0 €);-", FONT_FORMULA),
    ]
    for r, label, val, nf, fc in proj_def:
        # Label in col B (90u) — no merge needed, right-aligned for clean layout
        lbl = ws.cell(row=r, column=2, value=label)
        lbl.font = _font(color=P_GRAY, bold=True)
        lbl.alignment = _align("right", "center")
        ws.row_dimensions[r].height = ROW_H_KPI_PROJECTED
        # Value in col C (20u) — wide enough for any formatted monetary value
        c = ws.cell(row=r, column=3, value=val)
        c.font = _font(bold=True, color=fc, size=12)
        c.number_format = nf
        c.border = _border_thin()
        c.alignment = _align("right")

    # ── EDX-001 : Chaîne décisionnelle ──────────────────────────────────────
    # Section de raisonnement sous le Decision Lab, même feuille, sans nouvelle sheet.
    reasoning_list = (result_dict or {}).get("decision_reasoning", []) or []
    has_reasoning  = any(
        r.get("why_this_decision") or r.get("problem_source")
        for r in reasoning_list
    )

    if has_reasoning:
        reasoning_by_idx = {r["decision_index"]: r for r in reasoning_list}
        decisions_list   = edm.executive_decisions[:MAX_DEC]

        # Ligne de séparation visuelle
        EDX_R_START = LAB_R_VAR_PCT + 4   # = 27 (2 lignes de marge après bord)
        ws.row_dimensions[EDX_R_START - 1].height = 14   # espace

        # En-tête de section
        _section_bar(ws, EDX_R_START,
                     "POURQUOI PEPPERYN RECOMMANDE CES DÉCISIONS — Chaîne de raisonnement",
                     col_end=8)
        ws.row_dimensions[EDX_R_START].height = 20

        # En-têtes colonnes
        _col_headers(ws, EDX_R_START + 1,
                     ["#", "Décision", "Problème résolu", "Raisonnement Pepperyn",
                      "Si vous n'agissez pas (90j)", "Confiance", "Correspondance"],
                     bg=P_NAVY)
        ws.row_dimensions[EDX_R_START + 1].height = 16

        r_edx = EDX_R_START + 2
        for i, dec in enumerate(decisions_list):
            rz = reasoning_by_idx.get(i, {})
            why        = rz.get("why_this_decision") or ""
            prob       = rz.get("problem_source") or ""
            risk       = rz.get("inaction_risk") or ""
            conf       = rz.get("decision_confidence")
            match_conf = rz.get("matching_confidence") or ""

            if not why and not prob:
                continue   # n'affiche que les décisions avec raisonnement

            alt = P_LIGHT if i % 2 == 0 else P_WHITE

            # Col 1 : numéro
            c = ws.cell(row=r_edx, column=1, value=i + 1)
            c.font = _font(color=P_SLATE, bold=True)
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.alignment = _align("center")

            # Col 2 : libellé décision (tronqué)
            c = ws.cell(row=r_edx, column=2, value=dec.decision)
            c.font = _font(color=P_DARK, bold=True)
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.alignment = _align("left", wrap=True)

            # Col 3 : problème résolu (problem_source)
            conf_suffix = ""
            if match_conf == "LOW":
                conf_suffix = " (≈)"
            elif match_conf == "FALLBACK_INDEX":
                conf_suffix = " (~)"
            c = ws.cell(row=r_edx, column=3, value=(prob + conf_suffix) if prob else "—")
            c.font = _font(color=P_BLUE if prob else P_SLATE, bold=bool(prob))
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.alignment = _align("left", wrap=True)

            # Col 4 : why_this_decision
            c = ws.cell(row=r_edx, column=4, value=why or "—")
            c.font = _font(color=P_DARK)
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.alignment = _align("left", wrap=True)

            # Col 5 : inaction_risk
            c = ws.cell(row=r_edx, column=5, value=risk or "—")
            c.font = _font(color=P_RED if risk else P_SLATE, italic=bool(risk))
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.alignment = _align("left", wrap=True)

            # Col 6 : decision_confidence
            conf_val = f"{conf}%" if conf is not None else "—"
            conf_color = (P_GREEN if (conf or 0) >= 75
                          else (P_AMBER if (conf or 0) >= 55 else P_RED))
            c = ws.cell(row=r_edx, column=6, value=conf_val)
            c.font = _font(color=conf_color, bold=True)
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.alignment = _align("center")

            # Col 7 : matching_confidence label
            match_label = {"HIGH": "Haute", "LOW": "Approximative",
                           "FALLBACK_INDEX": "Estimée"}.get(match_conf, "—")
            c = ws.cell(row=r_edx, column=7, value=match_label)
            c.font = _font(color=P_GRAY, italic=True)
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.alignment = _align("center")

            # Hauteur de ligne dynamique basée sur la colonne why (la plus longue)
            ws.row_dimensions[r_edx].height = _row_height_for_text(why or dec.decision, 70)
            r_edx += 1

    # SPEC: Widths are fixed. Renderer must apply predefined widths.
    # Décision (B) = primary content column — 51 % de la largeur totale (174u).
    # Colonnes secondaires comprimées à leur minimum fonctionnel.
    # Cible : décisions standard (≤ 100 chars) sur 2 lignes maximum.
    _widths(ws, COLS_DECISION_LAB)


# ─── FEUILLE 4 : SENSIBILITÉ ──────────────────────────────────────────────────

def _build_sensitivity(wb: Workbook, edm) -> None:
    ws = wb.create_sheet(SN_SENSI)
    ws.tab_color = P_AMBER
    _setup_sheet(ws)

    _header_bar(ws, 1, "ANALYSE DE SENSIBILITÉ — Qu'est-ce qui influence le plus notre résultat ?", col_end=8)
    _instruction(ws, 2,
                 "Ce tableau montre l'impact € d'une variation de ±X% de chaque variable sur votre résultat. "
                 "Les valeurs se recalculent automatiquement lorsque vous modifiez la feuille Hypothèses.",
                 col_end=8)
    _nav_bar(ws, SN_SENSI, row=3, col_end=8)

    pcts = [-0.20, -0.10, -0.05, 0.05, 0.10, 0.20]
    pct_labels = [f"{int(p*100):+d}%" for p in pcts]
    _col_headers(ws, 4, ["Variable"] + pct_labels + ["Levier clé ?"], bg=P_NAVY)

    sens_vars = [
        ("Chiffre d'affaires",        f"'{SN_HYPO}'!B{HYPO_R_CA}",      True),
        ("Masse salariale",           f"'{SN_HYPO}'!B{HYPO_R_MASA}",    False),
        ("Charges fixes",             f"'{SN_HYPO}'!B{HYPO_R_CHARGES}", False),
        ("Sous-traitance",            f"'{SN_HYPO}'!B{HYPO_R_SOUSTRX}", False),
        ("Prix moyen de vente",       f"'{SN_HYPO}'!B{HYPO_R_PRIX}",    True),
        ("Volume (unités/mois × 12)", f"'{SN_HYPO}'!B{HYPO_R_VOLUME}",  True),
        ("Marge brute (%×CA)",        f"'{SN_HYPO}'!B{HYPO_R_CA}",      True),
        ("Taux d'inflation",          f"'{SN_HYPO}'!B{HYPO_R_MASA}",    False),
    ]

    for i, (label, ref, is_lever) in enumerate(sens_vars):
        r = 5 + i
        alt = P_LIGHT if i % 2 == 0 else P_WHITE

        c = ws.cell(row=r, column=1, value=label)
        c.font = _font(color=P_DARK, bold=True)
        c.fill = _fill(alt)
        c.border = _border_thin()

        for j, pct in enumerate(pcts):
            formula = f"=IFERROR({ref}*{pct},0)"
            c = ws.cell(row=r, column=2 + j, value=formula)
            c.font = _font(color=P_GREEN if pct > 0 else P_RED, bold=True)
            c.fill = _fill(alt)
            c.border = _border_thin()
            c.number_format = "#,##0 €;(#,##0 €);-"
            c.alignment = _align("right")

        lev = ws.cell(row=r, column=2 + len(pcts), value="★ Levier clé" if is_lever else "—")
        lev.font = _font(color=P_AMBER if is_lever else P_SLATE, bold=is_lever)
        lev.fill = _fill(alt)
        lev.border = _border_thin()

    r_note = 5 + len(sens_vars) + 2
    _section_bar(ws, r_note, "COMMENT LIRE CE TABLEAU", col_end=8, bg=P_NAVY)
    guide = [
        ("Valeurs vertes (+)", "  Signifie : la variable aide le résultat quand elle augmente (ex: CA, prix, volume)"),
        ("Valeurs rouges (−)", "  Signifie : la variable pèse sur le résultat quand elle augmente (ex: coûts)"),
        ("★ Levier clé",      "  Variable avec le plus fort potentiel d'impact sur la performance"),
        ("Pour agir",         "  → Allez dans la feuille Hypothèses (onglet bleu) et modifiez les cellules bleues"),
    ]
    for k, (label, val) in enumerate(guide):
        r = r_note + 1 + k
        ws.cell(row=r, column=1, value=label).font = _font(bold=True, color=P_DARK)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _font(color=P_GRAY)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

    _widths(ws, COLS_SENSIBILITE)


# ─── FEUILLE 5 : SCÉNARIOS ───────────────────────────────────────────────────

def _build_scenarios(wb: Workbook, edm) -> None:
    ws = wb.create_sheet(SN_SCEN)
    ws.tab_color = P_BLUE
    _setup_sheet(ws)

    _header_bar(ws, 1, "COMPARAISON DE SCÉNARIOS — Prudent · Central · Ambitieux · Personnalisé", col_end=5)
    _instruction(ws, 2,
                 "Le scénario Personnalisé reprend vos simulations du Decision Lab en temps réel. "
                 "Les trois autres appliquent des multiplicateurs au même impact de base.",
                 col_end=5)
    _nav_bar(ws, SN_SCEN, row=3, col_end=5)

    _col_headers(ws, 4, ["Indicateur", "Prudent (75%)", "Central (100%)",
                          "Ambitieux (130%)", "Personnalisé (Decision Lab)"], bg=P_NAVY)

    lc = get_column_letter(LAB_COL_IMPACT)
    lab_total_ref   = f"='{SN_LAB}'!{lc}{LAB_R_TOTAL}"
    lab_ebitda_ref  = f"='{SN_LAB}'!C{LAB_R_EBITDA}"

    row_base   = 5   # ligne EBITDA de base
    row_impact = 6   # ligne Impact décisions
    row_proj   = 7   # ligne EBITDA projeté

    metrics = [
        ("EBITDA de base (€)",
         f"={lab_ebitda_ref}", f"={lab_ebitda_ref}", f"={lab_ebitda_ref}", f"={lab_ebitda_ref}",
         "#,##0 €;(#,##0 €);-"),
        ("Impact décisions (€)",
         f"={lab_total_ref}*0.75", f"={lab_total_ref}", f"={lab_total_ref}*1.30", f"={lab_total_ref}",
         "#,##0 €;(#,##0 €);-"),
        ("EBITDA projeté (€)",
         f"=B{row_base}+B{row_impact}", f"=C{row_base}+C{row_impact}",
         f"=D{row_base}+D{row_impact}", f"=E{row_base}+E{row_impact}",
         "#,##0 €;(#,##0 €);-"),
        ("Variation vs actuel (%)",
         f"=IFERROR(B{row_impact}/ABS(B{row_base}),0)",
         f"=IFERROR(C{row_impact}/ABS(C{row_base}),0)",
         f"=IFERROR(D{row_impact}/ABS(D{row_base}),0)",
         f"=IFERROR(E{row_impact}/ABS(E{row_base}),0)",
         "0.0%"),
        ("Coût inaction évité (€)",
         f"=ABS(EDM!B{EDM_R_COI_YEAR})*0.75",
         f"=ABS(EDM!B{EDM_R_COI_YEAR})",
         f"=ABS(EDM!B{EDM_R_COI_YEAR})*1.30",
         f"=ABS(EDM!B{EDM_R_COI_YEAR})",
         "#,##0 €;(#,##0 €);-"),
        ("Score santé (rappel)",
         f"=EDM!B{EDM_R_HEALTH}", f"=EDM!B{EDM_R_HEALTH}",
         f"=EDM!B{EDM_R_HEALTH}", f"=EDM!B{EDM_R_HEALTH}",
         "0\"/10\""),
    ]

    scen_colors = [P_AMBER, P_NAVY, P_GREEN, P_BLUE]
    row_base = 5
    for i, (label, v1, v2, v3, v4, nf) in enumerate(metrics):
        r = row_base + i
        alt = P_LIGHT if i % 2 == 0 else P_WHITE

        ws.cell(row=r, column=1, value=label).font = _font(bold=True, color=P_DARK)
        ws.cell(row=r, column=1).fill = _fill(alt)
        ws.cell(row=r, column=1).border = _border_thin()

        is_hero = (r == row_base + 2)
        for j, val in enumerate([v1, v2, v3, v4]):
            c = ws.cell(row=r, column=2 + j, value=val)
            c.font = _font(bold=True, color=P_WHITE if is_hero else P_DARK,
                           size=14 if is_hero else 10)
            c.fill = _fill(scen_colors[j] if is_hero else alt)
            c.number_format = nf
            c.border = _border_thin()
            c.alignment = _align("right")

    # Scénarios qualitatifs LLM
    if edm.scenarios:
        r_sc = row_base + len(metrics) + 2
        _section_bar(ws, r_sc, "SCÉNARIOS IDENTIFIÉS PAR LE COPILOTE", col_end=5)
        _col_headers(ws, r_sc + 1, ["Scénario", "Label", "Description", "", ""], bg=P_NAVY)
        # Fusionner le header "Description" pour qu'il couvre C-E comme le contenu
        ws.merge_cells(start_row=r_sc + 1, start_column=3,
                       end_row=r_sc + 1, end_column=5)
        for k, sc in enumerate(edm.scenarios[:5]):
            r = r_sc + 2 + k
            ws.cell(row=r, column=1, value=getattr(sc, "nom", f"S{k+1}")).font = _font(bold=True)
            ws.cell(row=r, column=2, value=getattr(sc, "label", "—")).font = _font(color=P_BLUE)
            desc = getattr(sc, "description", "—")
            c = ws.cell(row=r, column=3, value=(desc or "")[:140])
            c.font = _font(color=P_GRAY)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

    _widths(ws, COLS_SCENARIOS)


# ─── FEUILLE 6 : ROADMAP ─────────────────────────────────────────────────────

def _build_roadmap(wb: Workbook, edm) -> None:
    ws = wb.create_sheet(SN_ROAD)
    ws.tab_color = P_NAVY
    _setup_sheet(ws)

    _header_bar(ws, 1, "ROADMAP D'EXÉCUTION — Pilotage des décisions", col_end=9)
    _instruction(ws, 2,
                 "Modifiez le Statut (colonne D, cellule bleue) pour suivre l'avancement. "
                 "Valeurs : To launch · In progress · Completed · Blocked",
                 col_end=9)
    _nav_bar(ws, SN_ROAD, row=3, col_end=9)

    _col_headers(ws, 4, ["#", "Décision / Action", "Responsable", "Statut",
                          "Impact (€)", "Priorité", "Horizon", "Phase 90j", "ROI score"])

    statuses_color = {
        "To launch": P_SLATE, "In progress": P_BLUE,
        "Completed": P_GREEN, "Blocked": P_RED,
    }

    for i in range(MAX_DEC):
        r = 5 + i
        alt = P_LIGHT if i % 2 == 0 else P_WHITE
        dec = edm.executive_decisions[i] if i < len(edm.executive_decisions) else None

        for col in range(1, 10):
            ws.cell(row=r, column=col).fill = _fill(alt)
            ws.cell(row=r, column=col).border = _border_thin()

        ws.cell(row=r, column=1, value=i + 1).alignment = _align("center")
        ws.cell(row=r, column=1).font = _font(color=P_SLATE, bold=True)

        if dec:
            c = ws.cell(row=r, column=2, value=dec.decision)
            c.font = _font(color=P_DARK, bold=True)
            c.alignment = _align("left", wrap=True)
            # SPEC: Rows — Dynamic height. Renderer computes row height. Never fixed.
            ws.row_dimensions[r].height = _row_height_for_text(dec.decision, 76)

            c = ws.cell(row=r, column=3, value=dec.owner or "À définir")
            c.font = _font(color=FONT_INPUT)
            c.alignment = _align("left", wrap=True)   # SPEC: Left for descriptions, Wrap mandatory

            c = ws.cell(row=r, column=4, value=dec.status)
            c.font = _font(color=statuses_color.get(dec.status, P_SLATE), bold=True)
            c.fill = _fill("DBEAFE")
            c.border = _border_accent()

            c = ws.cell(row=r, column=5, value=f"=EDM!B{EDM_R_DEC_START + i}")
            c.font = _font(color=FONT_LINK, bold=True)
            c.number_format = "#,##0 €;(#,##0 €);-"

            p_lower = (dec.priority or "").lower()
            p_color = P_RED if "high" in p_lower else (P_AMBER if "medium" in p_lower else P_GRAY)
            ws.cell(row=r, column=6, value=dec.priority).font = _font(color=p_color, bold=True)
            c7 = ws.cell(row=r, column=7, value=dec.timeline or "—")
            c7.font = _font(color=P_GRAY)
            c7.alignment = _align("left", wrap=True)   # SPEC: Left for descriptions, Wrap mandatory
            phase_label = "0–30j" if i < 3 else ("30–60j" if i < 6 else "60–90j")
            ws.cell(row=r, column=8, value=phase_label).font = _font(color=P_SLATE)

            c = ws.cell(row=r, column=9, value=f"=EDM!G{EDM_R_DEC_START + i}")
            c.font = _font(color=FONT_LINK)
            c.number_format = "0.0"
            c.alignment = _align("center")
        else:
            ws.cell(row=r, column=2, value="[Décision à définir]").font = _font(color=P_SLATE, italic=True)

    # Phases roadmap
    if edm.roadmap_90_days:
        r_ph = 5 + MAX_DEC + 2
        _section_bar(ws, r_ph, "DÉTAIL PHASES 30 / 60 / 90 JOURS (source EDM)", col_end=9)
        r_ph += 1
        for phase in edm.roadmap_90_days[:3]:
            _section_bar(ws, r_ph,
                         f"Phase {phase.horizon} jours — {phase.phase_label}",
                         col_end=9, bg=P_SLATE)
            r_ph += 1
            for action in (phase.actions or [])[:6]:
                ws.cell(row=r_ph, column=1, value="›").font = _font(color=P_BLUE, bold=True)
                c = ws.cell(row=r_ph, column=2, value=action.decision)
                c.font = _font(color=P_DARK)
                ws.merge_cells(start_row=r_ph, start_column=2, end_row=r_ph, end_column=5)
                ws.cell(row=r_ph, column=6, value=action.owner or "—").font = _font(color=P_GRAY)
                ws.cell(row=r_ph, column=7, value=action.due_date or "—").font = _font(color=P_GRAY)
                if action.impact is not None:
                    c = ws.cell(row=r_ph, column=8, value=action.impact)
                    c.font = _font(color=P_GREEN if action.impact > 0 else P_RED, bold=True)
                    c.number_format = "#,##0 €;(#,##0 €);-"
                r_ph += 1

    _widths(ws, COLS_ROADMAP)


# ─── FEUILLE 7 : HISTORIQUE ───────────────────────────────────────────────────

def _build_historique(wb: Workbook, edm) -> None:
    ws = wb.create_sheet(SN_HIST)
    ws.tab_color = P_SLATE
    _setup_sheet(ws)

    date_now = datetime.now().strftime("%b %Y")
    _header_bar(ws, 1, "HISTORIQUE — Comparaison inter-périodes", col_end=6)
    _instruction(ws, 2,
                 "Colonne actuelle = données EDM en temps réel. "
                 "Renseignez les colonnes bleues manuellement à chaque nouvelle analyse.",
                 col_end=6)
    _nav_bar(ws, SN_HIST, row=3, col_end=6)

    _col_headers(ws, 4, ["Indicateur", date_now, "Mois N-1",
                          "Trim. N-1", "An N-1", "Évolution (vs An N-1)"], bg=P_NAVY)

    metrics = [
        ("EBITDA (€)",               f"=EDM!B{EDM_R_EBITDA}",    "#,##0 €;(#,##0 €);-"),
        ("Cash disponible (€)",      f"=EDM!B{EDM_R_CASH}",      "#,##0 €;(#,##0 €);-"),
        ("Coût de l'inaction / an",  f"=EDM!B{EDM_R_COI_YEAR}",  "#,##0 €;(#,##0 €);-"),
        ("Impact total identifié",   f"=EDM!B{EDM_R_IMPACT}",    "#,##0 €;(#,##0 €);-"),
        ("Score santé",              f"=EDM!B{EDM_R_HEALTH}",    "0\"/10\""),
        ("Niveau de confiance",      f"=EDM!B{EDM_R_CONFID}",    "0\"/10\""),
        ("EBITDA projeté (Decision Lab)", f"='{SN_LAB}'!C{LAB_R_PROJ}", "#,##0 €;(#,##0 €);-"),
        ("Variation projetée (%)",   f"='{SN_LAB}'!C{LAB_R_VAR_PCT}", "0.0%"),
    ]

    for i, (label, cur_val, nf) in enumerate(metrics):
        r = 5 + i
        alt = P_LIGHT if i % 2 == 0 else P_WHITE

        c = ws.cell(row=r, column=1, value=label)
        c.font = _font(bold=True, color=P_DARK)
        c.fill = _fill(alt)
        c.border = _border_thin()

        c = ws.cell(row=r, column=2, value=cur_val)
        c.font = _font(bold=True, color=FONT_LINK)
        c.fill = _fill(alt)
        c.number_format = nf
        c.border = _border_thin()
        c.alignment = _align("right")

        for j in range(1, 4):
            c = ws.cell(row=r, column=2 + j)
            c.font = _font(color=FONT_INPUT)
            c.fill = _fill("EFF6FF" if j % 2 == 0 else alt)
            c.number_format = nf
            c.border = _border_thin()
            c.alignment = _align("right")

        c = ws.cell(row=r, column=6, value=f"=IFERROR((B{r}-E{r})/ABS(E{r}),\"—\")")
        c.font = _font(color=FONT_FORMULA, bold=True)
        c.fill = _fill(alt)
        c.number_format = "0.0%"
        c.border = _border_thin()
        c.alignment = _align("right")

    # Guide d'utilisation
    r_guide = 5 + len(metrics) + 2
    _section_bar(ws, r_guide, "MODE D'EMPLOI", col_end=6, bg=P_SLATE)
    guide = [
        "Après chaque nouvelle analyse Pepperyn, copiez-collez les valeurs actuelles dans 'Mois N-1'.",
        "Renseignez 'Trim. N-1' et 'An N-1' manuellement depuis vos données de gestion.",
        "La colonne Évolution se calcule automatiquement dès que An N-1 est renseignée.",
        "Ce tableau vous permet de suivre la progression de votre santé financière dans le temps.",
    ]
    for k, note in enumerate(guide):
        r = r_guide + 1 + k
        c = ws.cell(row=r, column=1, value=f"  {k+1}.  {note}")
        c.font = _font(italic=True, color=P_GRAY)
        c.fill = _fill(P_LIGHT)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = ROW_H_GUIDE_NOTE

    _widths(ws, COLS_HISTORIQUE)


# ─── POINT D'ENTRÉE ───────────────────────────────────────────────────────────

def generate_excel_report(
    analysis: AnalysisResult,
    original_data: "dict[str, Any] | ExecutiveCaseJSON",
    filename: str = "analyse",
) -> bytes:
    """
    Génère l'Executive Financial Model™ v0.9 (9 feuilles).

    Args:
        analysis     : AnalysisResult (inchangé — non utilisé en V2).
        original_data: ExecutiveCaseJSON (V2 — source unique de vérité)
                       OU dict brut de l'analyse LLM (legacy).
        filename     : nom du fichier (couverture Accueil).

    Pipeline V2 :
        - Si ExecutiveCaseJSON → adapté via case_to_edm() + case_to_result_dict()
          Garantie : Excel affiche exactement les mêmes chiffres que PDF et PPTX.
        - Si dict → comportement legacy inchangé.

    Ordre des feuilles :
      0. 🏠 Accueil   1. 📊 Dashboard  2. ⚙ Hypothèses  3. 🎯 Decision Lab
      4. 📈 Sensibilité  5. 📉 Scénarios  6. 🗺 Roadmap  7. 🕒 Historique
      8. EDM (masquée)
    """
    if isinstance(original_data, ExecutiveCaseJSON):
        # ── V2 : source unique de vérité ─────────────────────────────────────
        from services.executive_case_builder import case_to_edm, case_to_result_dict
        edm          = case_to_edm(original_data)
        original_data = case_to_result_dict(original_data)
    else:
        # ── Legacy : comportement existant ───────────────────────────────────
        from services.executive_decision_model import build_executive_decision_model
        edm = build_executive_decision_model(original_data or {})

    wb = Workbook()
    wb.remove(wb.active)

    # RULE 003 — Renderer Responsibility
    # Each sheet builder is isolated. A crash on one sheet never kills the workbook.
    # The renderer owns its presentation problems — they never reach the caller.

    def _safe_sheet(sheet_name, builder_fn):
        try:
            builder_fn()
        except Exception:
            # Graceful fallback: minimal sheet with error label
            if sheet_name not in wb.sheetnames:
                ws_err = wb.create_sheet(sheet_name)
            else:
                ws_err = wb[sheet_name]
            ws_err["A1"] = f"{sheet_name} — données insuffisantes"

    # EDM en premier (toutes les autres feuilles le référencent)
    _safe_sheet(SN_EDM, lambda: _build_edm(wb, edm, original_data))

    # 8 feuilles visibles — Accueil en tête
    _safe_sheet(SN_ACCUEIL,  lambda: _build_accueil(wb))
    _safe_sheet(SN_DASH,     lambda: _build_dashboard(wb, edm, original_data))
    _safe_sheet(SN_HYPO,     lambda: _build_hypotheses(wb, edm))
    _safe_sheet(SN_LAB,      lambda: _build_decision_lab(wb, edm, original_data))
    _safe_sheet(SN_SENSI,    lambda: _build_sensitivity(wb, edm))
    _safe_sheet(SN_SCEN,     lambda: _build_scenarios(wb, edm))
    _safe_sheet(SN_ROAD,     lambda: _build_roadmap(wb, edm))
    _safe_sheet(SN_HIST,     lambda: _build_historique(wb, edm))

    # Masquer la feuille EDM
    if SN_EDM in wb.sheetnames:
        wb[SN_EDM].sheet_state = "hidden"

    # Accueil actif à l'ouverture
    if SN_ACCUEIL in wb.sheetnames:
        wb.active = wb[SN_ACCUEIL]

    # Recalcul automatique à l'ouverture
    wb.calculation.calcMode = "auto"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
