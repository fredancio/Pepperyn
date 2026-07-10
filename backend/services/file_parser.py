"""
File parser service — extracts structured data from Excel, PDF, CSV files.
No LLM calls — pure pandas/openpyxl/pdfplumber.

Formula handling strategy for Excel files:
  1. PRIMARY  — openpyxl data_only=True  → reads cached formula results
               Covers ~95% of real user files (saved by Excel / LibreOffice).
  2. FALLBACK — LibreOffice headless subprocess re-opens and re-saves the file,
               forcing formula evaluation.  Used when the primary read yields
               >40% null cells (file was generated programmatically without cache).
"""
import io
import logging
import os
import subprocess
import tempfile
import json
from typing import Any

import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

MAX_ROWS = 10_000
MAX_COLS = 50

# Null-density threshold: if >40% of cells that should have data are null,
# assume the file has no formula cache and trigger the LibreOffice fallback.
NULL_DENSITY_THRESHOLD = 0.40


def parse_file(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """
    Parse a financial file and return a structured summary dict.
    Keeps total under ~2000 tokens for LLM input.
    """
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext in ('xlsx', 'xls'):
        return _parse_excel(file_bytes, filename)
    elif ext == 'csv':
        return _parse_csv(file_bytes, filename)
    elif ext == 'pdf':
        return _parse_pdf(file_bytes, filename)
    else:
        raise ValueError(f"Format non supporté: .{ext}")


# ── Excel ──────────────────────────────────────────────────────────────────────

def _parse_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """
    Parse Excel with formula-aware reading.
    Tries openpyxl data_only first; falls back to LibreOffice if needed.
    """
    # ── Step 1: try data_only (reads cached formula results) ──────────────────
    try:
        wb_data = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheets_data, null_ratio = _extract_sheets_openpyxl(wb_data)

        if null_ratio <= NULL_DENSITY_THRESHOLD:
            logger.info(f"[parser] Excel read via openpyxl data_only (null ratio={null_ratio:.0%})")
            return _build_excel_result(filename, wb_data.sheetnames, sheets_data)

        logger.info(
            f"[parser] High null ratio ({null_ratio:.0%}) — formulas not cached. "
            "Trying LibreOffice fallback…"
        )
    except Exception as e:
        logger.warning(f"[parser] openpyxl data_only failed: {e}")

    # ── Step 2: LibreOffice fallback ──────────────────────────────────────────
    try:
        evaluated_bytes = _convert_with_libreoffice(file_bytes, filename)
        wb_eval = openpyxl.load_workbook(io.BytesIO(evaluated_bytes), data_only=True)
        sheets_data, null_ratio = _extract_sheets_openpyxl(wb_eval)
        logger.info(f"[parser] LibreOffice fallback succeeded (null ratio={null_ratio:.0%})")
        return _build_excel_result(filename, wb_eval.sheetnames, sheets_data)
    except Exception as e:
        logger.warning(f"[parser] LibreOffice fallback failed: {e}. Falling back to pandas.")

    # ── Step 3: Last resort — plain pandas (formulas read as strings) ─────────
    return _parse_excel_pandas(file_bytes, filename)


def _extract_sheets_openpyxl(
    wb: openpyxl.Workbook,
) -> tuple[list[dict[str, Any]], float]:
    """
    Convert openpyxl workbook sheets to DataFrames and analyse them.
    Returns (sheets_data, overall_null_ratio).
    """
    sheets_data: list[dict[str, Any]] = []
    total_cells = 0
    null_cells = 0

    # RÈGLE ABSOLUE N°10 : lire jusqu'à 8 feuilles en détail (vs 5 avant).
    # Le manifeste ALL_SHEETS_MANIFEST listera les feuilles non parsées pour
    # que le LLM ne déclare jamais une donnée absente sans les avoir vérifiées.
    for sheet_name in wb.sheetnames[:8]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            sheets_data.append({"sheet_name": sheet_name, "rows": 0, "columns": []})
            continue

        # Detect best header row (skip merged title rows)
        header_idx = _detect_header_row_from_rows(rows)
        header = [str(c) if c is not None else f"Col_{i}" for i, c in enumerate(rows[header_idx])]

        data_rows = rows[header_idx + 1 : header_idx + 1 + MAX_ROWS]
        df = pd.DataFrame(data_rows, columns=header)
        df = df.loc[:, df.columns[:MAX_COLS]]
        df = df.dropna(how='all').dropna(axis=1, how='all')

        # Null density tracking
        numeric_df = df.select_dtypes(include=['number', 'object'])
        total_cells += max(df.size, 1)
        null_cells += int(df.isna().sum().sum())

        sheets_data.append(_analyze_dataframe(df, sheet_name))

    overall_null_ratio = null_cells / max(total_cells, 1)
    return sheets_data, overall_null_ratio


def _extract_bfr_summary(sheets_data: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Proactively extract BFR / liquidity indicators from all parsed sheets.

    Placed at the TOP of the LLM JSON payload so these critical values are
    never lost due to the 14 000-char truncation applied in llm_service.py.
    Scans every row of every sheet for cells whose string value contains
    DSO / DPO / DIO / BFR / trésorerie keywords and picks adjacent numerics.
    """
    BFR_PATTERNS = {
        "dso_jours":     ["dso", "délai de recouvrement", "jours clients", "délai clients",
                          "days sales outstanding"],
        "dpo_jours":     ["dpo", "délai fournisseurs", "délai paiement fournisseur",
                          "days payable outstanding"],
        "dio_jours":     ["dio", "jours de stock", "rotation des stocks", "days inventory"],
        "bfr_jours":     ["bfr (jours)", "bfr en jours", "besoin en fonds de roulement (j",
                          "bfr normatif (j"],
        "bfr_eur":       ["bfr (€", "bfr total", "besoin en fonds de roulement (€",
                          "surfinancement", "excédent de bfr"],
        "tresorerie_eur": ["trésorerie", "tresorerie", "solde de trésorerie", "cash disponible"],
    }

    found: dict[str, Any] = {}

    for sheet in sheets_data:
        rows = sheet.get("full_table") or sheet.get("sample_rows") or []
        sheet_name = sheet.get("sheet_name", "")
        for row in rows:
            for cell_val in row.values():
                if not isinstance(cell_val, str):
                    continue
                cell_lower = cell_val.lower()
                for kpi, patterns in BFR_PATTERNS.items():
                    if kpi in found:
                        continue
                    if any(pat in cell_lower for pat in patterns):
                        # Pick the first non-zero numeric value in the same row
                        nums = [
                            v for v in row.values()
                            if isinstance(v, (int, float)) and v not in (0, 0.0)
                        ]
                        if nums:
                            found[kpi] = {
                                "label": cell_val[:80],
                                "value": nums[0],
                                "sheet": sheet_name,
                            }

    return found


def _extract_bilan_summary(sheets_data: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Proactively extract key balance-sheet (bilan) indicators from all parsed sheets.

    Placed at the TOP of the LLM JSON payload (like bfr_summary) so bilan data
    is never lost due to the 14 000-char truncation in llm_service.py.
    Scans every row of every sheet for actif / passif / capitaux keywords and
    picks the first adjacent non-zero numeric value.
    """
    BILAN_PATTERNS = {
        "total_actif":           ["total actif", "total de l'actif", "total assets"],
        "actifs_immobilises":    ["actifs immobilisés", "actif immobilisé", "immobilisations nettes",
                                  "total immobilisations"],
        "actifs_circulants":     ["actif circulant", "actifs circulants", "total actif circulant",
                                  "current assets"],
        "creances_clients":      ["créances clients", "creances clients", "clients et comptes",
                                  "accounts receivable"],
        "stocks":                ["stocks", "stock total", "inventaires"],
        "tresorerie_actif":      ["trésorerie actif", "tresorerie actif", "disponibilités",
                                  "valeurs mobilières", "cash and cash equivalents"],
        "total_passif":          ["total passif", "total du passif", "total liabilities"],
        "capitaux_propres":      ["capitaux propres", "fonds propres", "total capitaux propres",
                                  "shareholders equity", "equity"],
        "dettes_financieres_lt": ["dettes financières", "emprunts", "dettes à long", "dettes lt",
                                  "long term debt"],
        "dettes_fournisseurs":   ["dettes fournisseurs", "fournisseurs et comptes",
                                  "accounts payable"],
        "autres_dettes_ct":      ["autres dettes court terme", "dettes à court terme",
                                  "current liabilities"],
    }

    found: dict[str, Any] = {}

    for sheet in sheets_data:
        rows = sheet.get("full_table") or sheet.get("sample_rows") or []
        sheet_name = sheet.get("sheet_name", "")
        for row in rows:
            for cell_val in row.values():
                if not isinstance(cell_val, str):
                    continue
                cell_lower = cell_val.lower().strip()
                for kpi, patterns in BILAN_PATTERNS.items():
                    if kpi in found:
                        continue
                    if any(pat in cell_lower for pat in patterns):
                        nums = [
                            v for v in row.values()
                            if isinstance(v, (int, float)) and v not in (0, 0.0)
                        ]
                        if nums:
                            found[kpi] = {
                                "label": cell_val[:80],
                                "value": nums[0],
                                "sheet": sheet_name,
                            }

    return found


def _build_excel_result(
    filename: str, sheet_names: list[str], sheets_data: list[dict[str, Any]]
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "filename": filename,
        "format": "excel",
        "sheets_count": len(sheet_names),
        "total_rows": sum(s.get("rows", 0) for s in sheets_data),
    }
    # BFR summary injected FIRST so it survives the 14 000-char LLM truncation
    bfr_summary = _extract_bfr_summary(sheets_data)
    if bfr_summary:
        result["bfr_summary"] = bfr_summary

    # Bilan summary injected at the TOP — key balance-sheet lines are
    # extracted proactively and placed before the full sheets JSON so they
    # survive the 14 000-char truncation even if the bilan sheet appears later.
    bilan_summary = _extract_bilan_summary(sheets_data)
    if bilan_summary:
        result["bilan_summary"] = bilan_summary

    # ── RÈGLE ABSOLUE N°10 : manifeste complet de toutes les feuilles ────────
    # Injecté dans le contexte LLM AVANT les données.
    # Interdit de déclarer une donnée absente sans avoir vérifié que la feuille
    # qui pourrait la contenir figure bien dans "sheets_parsed_in_detail".
    parsed_names: set[str] = {s.get("sheet_name", "") for s in sheets_data}
    not_parsed: list[str] = [n for n in sheet_names if n not in parsed_names]
    result["all_sheets_manifest"] = {
        "total_sheets_in_workbook": len(sheet_names),
        "sheets_parsed_in_detail": [n for n in sheet_names if n in parsed_names],
        "sheets_present_but_not_parsed": not_parsed,
        "audit_note": (
            f"ATTENTION : {len(not_parsed)} feuille(s) présente(s) dans le classeur "
            f"n'ont PAS été analysées en détail : {not_parsed}. "
            "Avant de conclure qu'une donnée est absente (bilan, trésorerie, DSO…), "
            "vérifier si elle pourrait se trouver dans ces feuilles non parsées."
        ) if not_parsed else "Toutes les feuilles du classeur ont été analysées en détail.",
    }

    result["sheets"] = sheets_data
    return result


def _detect_header_row_from_rows(rows: list[tuple], max_scan: int = 6) -> int:
    """
    Find the row index with the most non-null values (≥35% fill).
    Works on raw openpyxl rows (list of tuples).
    """
    best_row = 0
    best_score = 0
    total_cols = max(len(r) for r in rows[:max_scan]) if rows else 1

    for idx, row in enumerate(rows[:max_scan]):
        filled = sum(1 for c in row if c is not None)
        if filled > best_score and filled / max(total_cols, 1) >= 0.35:
            best_score = filled
            best_row = idx
    return best_row


def _convert_with_libreoffice(file_bytes: bytes, filename: str) -> bytes:
    """
    Use LibreOffice headless to open the Excel file and re-save it,
    which evaluates all formulas and populates the formula cache.
    Returns the evaluated .xlsx as bytes.
    Raises RuntimeError if LibreOffice is not available or conversion fails.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        # Write input file
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'xlsx'
        in_path = os.path.join(tmpdir, f"input.{ext}")
        out_dir = os.path.join(tmpdir, "out")
        os.makedirs(out_dir, exist_ok=True)

        with open(in_path, 'wb') as f:
            f.write(file_bytes)

        # Run LibreOffice to convert to xlsx (evaluates formulas)
        result = subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--norestore",
                "--calc",
                "--convert-to", "xlsx",
                "--outdir", out_dir,
                in_path,
            ],
            capture_output=True,
            timeout=60,
        )

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice exited with code {result.returncode}: "
                f"{result.stderr.decode('utf-8', errors='replace')[:500]}"
            )

        # Find output file
        out_files = [f for f in os.listdir(out_dir) if f.endswith('.xlsx')]
        if not out_files:
            raise RuntimeError("LibreOffice produced no output file")

        out_path = os.path.join(out_dir, out_files[0])
        with open(out_path, 'rb') as f:
            return f.read()


def _parse_excel_pandas(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """Last-resort Excel parsing using pandas (formulas appear as None)."""
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    all_sheet_names = xl.sheet_names
    sheets_data = []
    parsed_names: set[str] = set()

    for sheet_name in all_sheet_names[:8]:  # RÈGLE N°10 : 8 feuilles max en détail
        header_row = _detect_header_row(xl, sheet_name)
        df = xl.parse(sheet_name, nrows=MAX_ROWS, header=header_row)
        df = df.loc[:, df.columns[:MAX_COLS]]
        sheet_summary = _analyze_dataframe(df, sheet_name)
        sheets_data.append(sheet_summary)
        parsed_names.add(sheet_name)

    not_parsed = [n for n in all_sheet_names if n not in parsed_names]
    result_pandas: dict[str, Any] = {
        "filename": filename,
        "format": "excel",
        "sheets_count": len(all_sheet_names),
        "total_rows": sum(s.get("rows", 0) for s in sheets_data),
    }
    # BFR + bilan summaries injected FIRST (before sheets) to survive 14 000-char truncation
    bfr_s = _extract_bfr_summary(sheets_data)
    if bfr_s:
        result_pandas["bfr_summary"] = bfr_s
    bilan_s = _extract_bilan_summary(sheets_data)
    if bilan_s:
        result_pandas["bilan_summary"] = bilan_s
    result_pandas["all_sheets_manifest"] = {
        "total_sheets_in_workbook": len(all_sheet_names),
        "sheets_parsed_in_detail": list(parsed_names),
        "sheets_present_but_not_parsed": not_parsed,
        "audit_note": (
            f"ATTENTION : {len(not_parsed)} feuille(s) non analysée(s) : {not_parsed}. "
            "Vérifier avant de déclarer une donnée absente."
        ) if not_parsed else "Toutes les feuilles ont été analysées.",
    }
    result_pandas["sheets"] = sheets_data
    return result_pandas


def _detect_header_row(xl: pd.ExcelFile, sheet_name: str, max_scan: int = 6) -> int:
    """
    Scan the first few rows to find the best header row.
    Returns the row index (0-based) where most columns have non-null values.
    Skips merged title rows which usually have only 1-2 filled cells.
    """
    try:
        df_raw = xl.parse(sheet_name, header=None, nrows=max_scan)
        best_row = 0
        best_score = 0
        total_cols = len(df_raw.columns)
        for row_idx in range(min(max_scan, len(df_raw))):
            row = df_raw.iloc[row_idx]
            filled = row.notna().sum()
            if filled > best_score and filled / max(total_cols, 1) >= 0.35:
                best_score = filled
                best_row = row_idx
        return best_row
    except Exception:
        return 0


# ── CSV ────────────────────────────────────────────────────────────────────────

def _parse_csv(file_bytes: bytes, filename: str) -> dict[str, Any]:
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), nrows=MAX_ROWS)
    except Exception:
        df = pd.read_csv(io.BytesIO(file_bytes), nrows=MAX_ROWS, sep=';')

    df = df.loc[:, df.columns[:MAX_COLS]]
    return {
        "filename": filename,
        "format": "csv",
        "sheets_count": 1,
        "sheets": [_analyze_dataframe(df, "Data")],
        "total_rows": len(df),
    }


# ── PDF ────────────────────────────────────────────────────────────────────────

def _parse_pdf(file_bytes: bytes, filename: str) -> dict[str, Any]:
    try:
        import pdfplumber
        text_content = []
        tables_data = []
        pages_count = 0

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            pages_count = len(pdf.pages)
            for i, page in enumerate(pdf.pages[:10]):
                text = page.extract_text()
                if text:
                    text_content.append(f"[Page {i+1}]\n{text[:1000]}")

                tables = page.extract_tables()
                for table in tables[:3]:
                    if table and len(table) > 1:
                        headers = [str(h or '').strip() for h in table[0]]
                        rows_sample = []
                        for row in table[1:11]:
                            rows_sample.append([str(c or '').strip() for c in row])
                        tables_data.append({
                            "headers": headers,
                            "rows_count": len(table) - 1,
                            "sample_rows": rows_sample
                        })

        return {
            "filename": filename,
            "format": "pdf",
            "pages": pages_count,
            "text_content": "\n\n".join(text_content)[:5000],
            "tables": tables_data,
        }
    except ImportError:
        raise ValueError("pdfplumber non installé. Installez: pip install pdfplumber")


# ── DataFrame analysis ─────────────────────────────────────────────────────────

def _analyze_dataframe(df: pd.DataFrame, sheet_name: str) -> dict[str, Any]:
    """
    Extract structured data from a DataFrame for LLM analysis.
    For small datasets (P&L, budgets): transmit full table.
    For large datasets: transmit stats + sample rows.
    """
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how='all').dropna(axis=1, how='all')

    # Convert object columns that look numeric (handles formula strings like "=SUM(...)")
    for col in df.select_dtypes(include=['object']).columns:
        converted = pd.to_numeric(df[col], errors='coerce')
        # Only replace if we gained significant numeric data
        if converted.notna().sum() > df[col].notna().sum() * 0.5:
            df[col] = converted

    summary: dict[str, Any] = {
        "sheet_name": sheet_name,
        "rows": len(df),
        "columns": list(df.columns[:20]),
    }

    numeric_cols = df.select_dtypes(include=['number']).columns
    is_pl_format = _is_pl_format(df)
    is_bilan_format = _is_bilan_format(df)

    # ── Full table for P&L / bilan / small structured datasets ───────────────
    # Les feuilles bilan (≤120 lignes) sont transmises en intégralité pour que le
    # LLM dispose de toutes les lignes actif/passif/capitaux — même si elles
    # dépassent les 60 lignes du seuil générique.
    send_full = (
        is_pl_format
        or (is_bilan_format and len(df) <= 120)
        or (len(df) <= 60 and len(df.columns) <= 20)
    )
    if send_full:
        full_rows = []
        for _, row in df.iterrows():
            row_dict = {str(k): _serialize_value(v) for k, v in row.items()}
            if any(v is not None for v in row_dict.values()):
                full_rows.append(row_dict)
        summary["full_table"] = full_rows
        if is_pl_format:
            summary["format_hint"] = "pl_mensuel"
        elif is_bilan_format:
            summary["format_hint"] = "bilan"
        else:
            summary["format_hint"] = "structured"
    else:
        # ── Stats for larger datasets ──────────────────────────────────────────
        if len(numeric_cols) > 0:
            numeric_summary: dict[str, Any] = {}
            for col in numeric_cols[:15]:
                col_data = df[col].dropna()
                if len(col_data) == 0:
                    continue
                stats: dict[str, Any] = {
                    "sum":   round(float(col_data.sum()), 2),
                    "mean":  round(float(col_data.mean()), 2),
                    "min":   round(float(col_data.min()), 2),
                    "max":   round(float(col_data.max()), 2),
                    "count": int(len(col_data)),
                }
                if len(col_data) > 5:
                    std = col_data.std()
                    if std > 0:
                        outliers = col_data[abs(col_data - col_data.mean()) > 3 * std]
                        if len(outliers) > 0:
                            stats["anomalies"] = [round(float(v), 2) for v in outliers.values[:3]]
                numeric_summary[col] = stats

            summary["numeric_analysis"] = numeric_summary

        # Sample rows
        sample_rows = []
        for _, row in df.head(8).iterrows():
            sample_rows.append({str(k): _serialize_value(v) for k, v in row.items()})
        if len(df) > 8:
            for _, row in df.tail(3).iterrows():
                sample_rows.append({str(k): _serialize_value(v) for k, v in row.items()})
        summary["sample_rows"] = sample_rows

    # ── Anomaly detection on numeric columns (all cases) ─────────────────────
    if len(numeric_cols) > 0:
        anomalies_detected = []
        for col in numeric_cols[:20]:
            col_data = df[col].dropna()
            if len(col_data) < 4:
                continue
            mean, std = col_data.mean(), col_data.std()
            if std > 0:
                for idx, val in col_data.items():
                    z = abs(val - mean) / std
                    if z > 2.5:
                        try:
                            row_label = str(df.loc[idx, df.columns[0]]) if isinstance(idx, int) and len(df.columns) > 0 else str(idx)
                        except (KeyError, IndexError):
                            row_label = str(idx)
                        anomalies_detected.append({
                            "colonne": str(col),
                            "valeur": round(float(val), 2),
                            "moyenne_col": round(float(mean), 2),
                            "ecart_sigma": round(float(z), 1),
                            "ligne": row_label[:50],
                        })
        if anomalies_detected:
            summary["anomalies_statistiques"] = anomalies_detected[:10]

    detected_categories = _detect_financial_columns(list(numeric_cols))
    if detected_categories:
        summary["detected_categories"] = detected_categories

    return summary


def _is_pl_format(df: pd.DataFrame) -> bool:
    """
    Detect if the dataframe looks like a monthly P&L.
    Heuristic: has columns named like months, or first column has financial keywords.
    """
    mois_fr = {"jan", "fév", "fev", "mar", "avr", "mai", "jui", "jul", "aoû", "aou",
               "sep", "oct", "nov", "déc", "dec"}
    mois_en = {"jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"}
    pl_keywords = {"ca", "chiffre", "ebitda", "ebit", "marge", "résultat", "resultat",
                   "charges", "salaire", "amortissement", "revenu", "produit"}

    col_names_lower = {str(c).lower()[:5] for c in df.columns}
    has_month_cols = bool(col_names_lower & (mois_fr | mois_en))

    first_col_values = set()
    if len(df.columns) > 0:
        first_col_values = {str(v).lower() for v in df.iloc[:, 0].dropna().head(20)}
    has_pl_rows = any(any(kw in v for kw in pl_keywords) for v in first_col_values)

    return has_month_cols or has_pl_rows


def _is_bilan_format(df: pd.DataFrame) -> bool:
    """
    Detect if the dataframe looks like a balance sheet (bilan).
    Heuristic: first column contains bilan-specific keywords.
    """
    bilan_keywords = {
        "actif", "passif", "capitaux", "immobilisation", "créance", "creance",
        "dettes", "dette", "fonds propres", "patrimoine", "bfr",
        "fonds de roulement", "total actif", "total passif", "trésorerie",
        "tresorerie", "stock", "disponibilit",
    }
    first_col_values: set[str] = set()
    if len(df.columns) > 0:
        first_col_values = {str(v).lower() for v in df.iloc[:, 0].dropna().head(30)}
    return any(any(kw in v for kw in bilan_keywords) for v in first_col_values)


def _detect_financial_columns(columns: list[str]) -> dict[str, list[str]]:
    """Detect financial column types from column names."""
    categories: dict[str, list[str]] = {"revenus": [], "couts": [], "marges": [], "autres": []}

    revenue_keywords = ['ca', 'chiffre', 'vente', 'revenu', 'recette', 'produit', 'revenue', 'sales', 'turnover']
    cost_keywords = ['cout', 'charge', 'depense', 'achat', 'frais', 'cost', 'expense']
    margin_keywords = ['marge', 'resultat', 'benefice', 'profit', 'margin', 'ebitda', 'ebit', 'rex']

    for col in columns:
        col_lower = col.lower().replace('é', 'e').replace('û', 'u').replace('â', 'a').replace('ô', 'o')
        if any(kw in col_lower for kw in revenue_keywords):
            categories["revenus"].append(col)
        elif any(kw in col_lower for kw in cost_keywords):
            categories["couts"].append(col)
        elif any(kw in col_lower for kw in margin_keywords):
            categories["marges"].append(col)
        else:
            categories["autres"].append(col)

    return {k: v for k, v in categories.items() if v}


def assess_file_quality(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """
    Pre-analysis quality check. Evaluates the file before sending to LLM.
    Returns:
      - score (0-100)
      - issues: list of detected problems
      - recommendation: "ok" | "warn" | "block"
      - user_message: human-readable explanation in French
    """
    issues: list[str] = []
    score = 100
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    try:
        if ext in ('xlsx', 'xls'):
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet_names = wb.sheetnames
            n_sheets = len(sheet_names)

            # Too many sheets
            if n_sheets > 15:
                issues.append(f"Fichier très complexe : {n_sheets} onglets détectés")
                score -= 25
            elif n_sheets > 8:
                issues.append(f"{n_sheets} onglets détectés — seuls les 5 premiers seront analysés")
                score -= 10

            total_cells = 0
            null_cells = 0
            has_numeric = False
            header_issues = 0
            all_columns: list[str] = []

            for sheet_name in sheet_names[:5]:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # Check null density
                for row in rows[:200]:
                    for cell in row:
                        total_cells += 1
                        if cell is None or str(cell).strip() == '':
                            null_cells += 1
                        elif isinstance(cell, (int, float)):
                            has_numeric = True

                # Check headers
                if rows:
                    first_row = [str(c) if c is not None else '' for c in rows[0]]
                    unnamed = sum(1 for c in first_row if not c or c.startswith('Col_') or c.isdigit())
                    if len(first_row) > 0 and unnamed / max(len(first_row), 1) > 0.6:
                        header_issues += 1
                    all_columns.extend([c for c in first_row if c])

            null_ratio = null_cells / max(total_cells, 1)

            if null_ratio > 0.75:
                issues.append(f"Fichier très creux : {null_ratio:.0%} de cellules vides")
                score -= 35
            elif null_ratio > 0.55:
                issues.append(f"Nombreuses cellules vides ({null_ratio:.0%})")
                score -= 15

            if not has_numeric:
                issues.append("Aucune donnée numérique détectée — impossible d'analyser des chiffres")
                score -= 40

            if header_issues >= 2:
                issues.append("En-têtes de colonnes manquants ou illisibles")
                score -= 20

            # Check for cryptic column names (mostly codes/numbers)
            cryptic = sum(1 for c in all_columns if c and (c.isdigit() or len(c) <= 2))
            if len(all_columns) > 3 and cryptic / max(len(all_columns), 1) > 0.5:
                issues.append("Noms de colonnes cryptiques ou codifiés (non lisibles)")
                score -= 15

        elif ext == 'csv':
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), nrows=100)
            except Exception:
                df = pd.read_csv(io.BytesIO(file_bytes), nrows=100, sep=';')

            null_ratio = df.isna().sum().sum() / max(df.size, 1)
            numeric_cols = df.select_dtypes(include=['number']).columns
            if null_ratio > 0.7:
                issues.append(f"Fichier CSV très creux ({null_ratio:.0%} de valeurs manquantes)")
                score -= 30
            if len(numeric_cols) == 0:
                issues.append("Aucune colonne numérique détectée dans le CSV")
                score -= 35

        elif ext == 'pdf':
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                n_pages = len(pdf.pages)
                text_total = sum(len(p.extract_text() or '') for p in pdf.pages[:5])
                if n_pages > 30:
                    issues.append(f"PDF très long ({n_pages} pages) — seules les 10 premières seront analysées")
                    score -= 10
                if text_total < 200:
                    issues.append("PDF scanné ou illisible — le texte ne peut pas être extrait")
                    score -= 50

    except Exception as e:
        logger.warning(f"[quality_check] Error: {e}")
        return {"score": 50, "issues": [], "recommendation": "warn",
                "user_message": "Impossible d'évaluer la qualité du fichier à l'avance."}

    score = max(0, min(100, score))

    if score >= 65:
        recommendation = "ok"
        user_message = None
    elif score >= 35:
        recommendation = "warn"
        user_message = (
            "⚠️ **Ce fichier présente des limitations qui pourraient affecter la qualité de l'analyse :**\n\n"
            + "\n".join(f"• {issue}" for issue in issues)
            + "\n\n💡 *Pour une analyse optimale, consultez notre [guide de préparation des données](/guide-donnees).*"
        )
    else:
        recommendation = "block"
        user_message = (
            "❌ **Ce fichier est trop complexe ou trop incomplet pour être analysé correctement.**\n\n"
            + "\n".join(f"• {issue}" for issue in issues)
            + "\n\n**Que faire ?**\n"
            + "• Nettoyez le fichier en supprimant les onglets inutiles et en ajoutant des en-têtes claires\n"
            + "• Utilisez **Microsoft Copilot** ou **ChatGPT** pour une première passe de nettoyage\n"
            + "• Consultez notre [guide de préparation des données](/guide-donnees)\n\n"
            + "*Astuce : un fichier Excel propre avec 1-3 onglets structurés donne les meilleures analyses.*"
        )

    return {
        "score": score,
        "issues": issues,
        "recommendation": recommendation,
        "user_message": user_message,
    }


def _serialize_value(v: Any) -> Any:
    """Convert pandas values to JSON-serializable types."""
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (int, float)):
        if abs(v) > 1e15:
            return None
        return round(float(v), 4) if isinstance(v, float) else int(v)
    return str(v)[:100]
