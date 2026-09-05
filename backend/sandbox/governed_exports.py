"""Deterministic, synthetic-only exports for a persisted V1 governed envelope.

These renderers deliberately consume the immutable governed envelope rather
than the legacy AnalysisResult projection.  They perform no provider call and
do not create decisions or facts that are absent from the envelope.
"""

from __future__ import annotations

from io import BytesIO
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from services.v1_analysis_contract import GovernedAnalysisEnvelope

_NAVY = "183B56"
_BLUE = "2A6F97"
_PALE = "EAF2F8"
_TEXT = "243746"


def _neutralize_spreadsheet_formula(value: object) -> object:
    """Keep provider-authored text inert when Excel opens the workbook."""

    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _fact_map(envelope: GovernedAnalysisEnvelope) -> dict[str, str]:
    return {
        fact.fact_id: f"{fact.metric} ({fact.period}, {fact.source_sheet_ref}/{fact.source_field})"
        for fact in envelope.source_facts.facts
    }


def _refs(ids: tuple[str, ...], fact_map: dict[str, str]) -> str:
    return "; ".join(f"{fact_id}: {fact_map[fact_id]}" for fact_id in ids)


def generate_governed_excel(envelope: GovernedAnalysisEnvelope) -> bytes:
    """Render an auditable workbook from the validated envelope only."""

    envelope = GovernedAnalysisEnvelope.model_validate(envelope)
    analysis, source = envelope.governed_analysis, envelope.source_facts
    fact_map = _fact_map(envelope)
    wb = Workbook()
    summary = wb.active
    summary.title = "Synthese"
    summary.sheet_view.showGridLines = False
    summary.append(["Pepperyn - Analyse financiere gouvernee"])
    summary.append(["Perimetre", "Donnees synthetiques uniquement"])
    summary.append(["Periode", source.current_period or "UNKNOWN"])
    summary.append(["Statut de comprehension", source.status])
    summary.append(["Empreinte source SHA-256", source.source_representation_sha256])
    summary.append([])
    summary.append(["Diagnostic (inference)", analysis.executive_diagnosis])
    summary.append(["Faits cites", _refs(analysis.diagnosis_fact_ids, fact_map)])
    summary.append([])
    summary.append(["Limite", "Les recommandations IA ne constituent pas des decisions confirmees."])

    facts = wb.create_sheet("Faits sources")
    facts.append(["Fact ID", "Metrique", "Valeur", "Unite", "Periode", "Feuille source", "Champ source"])
    for fact in source.facts:
        facts.append([fact.fact_id, fact.metric, fact.value, fact.unit, fact.period,
                      fact.source_sheet_ref, fact.source_field])

    reasoning = wb.create_sheet("Inferences")
    reasoning.append(["Type", "Contenu", "Confiance", "Faits cites", "Validations requises"])
    for item in analysis.observations:
        reasoning.append([f"Observation source-matched - severite inferentielle {item.severity}",
                          f"{item.metric} = {item.observed_value}", None,
                          _refs((item.fact_id,), fact_map), ""])
    for item in analysis.dimension_assessments:
        reasoning.append([f"Dimension {item.scope} - score inferentiel {item.score}/10", item.rationale,
                          item.confidence, _refs(item.fact_ids, fact_map), "; ".join(item.validation_required)])
    for item in analysis.inferences:
        reasoning.append(["Inference", item.statement, item.confidence,
                          _refs(item.fact_ids, fact_map), "; ".join(item.validation_required)])
    for item in analysis.contradictions:
        reasoning.append(["Contradiction", item.statement, None, _refs(item.fact_ids, fact_map), ""])

    recommendations = wb.create_sheet("Recommandations")
    recommendations.append(["Priorite proposee", "Action proposee", "Rationale", "Faits cites", "Prerequis"])
    for item in analysis.recommendations:
        recommendations.append([item.priority, item.action, item.rationale,
                                _refs(item.fact_ids, fact_map), "; ".join(item.prerequisite_validation)])

    unknowns = wb.create_sheet("UNKNOWN")
    unknowns.append(["Materialite", "Question non resolue"])
    for item in analysis.unknowns:
        unknowns.append([item.materiality, item.question])
    for item in source.unknowns:
        unknowns.append(["SOURCE", item])

    for ws in wb.worksheets:
        ws.freeze_panes = "A2" if ws.max_row > 1 else None
        ws.auto_filter.ref = ws.dimensions if ws.max_column > 1 and ws.max_row > 1 else None
        for row in ws.iter_rows():
            for cell in row:
                cell.value = _neutralize_spreadsheet_formula(cell.value)
                cell.font = Font(name="Arial", size=10, color=_TEXT, bold=cell.font.bold)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in ws[1]:
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_NAVY)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in ws.columns:
            letter = column[0].column_letter
            longest = min(max((len(str(cell.value or "")) for cell in column), default=10), 65)
            ws.column_dimensions[letter].width = max(12, longest + 2)
    summary.freeze_panes = None
    summary.auto_filter.ref = None
    summary.merge_cells("A1:B1")
    summary["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    summary["A7"].fill = PatternFill("solid", fgColor=_PALE)
    summary["A7"].font = Font(name="Arial", bold=True, color=_NAVY)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_governed_pdf(envelope: GovernedAnalysisEnvelope) -> bytes:
    """Render a professional, bounded PDF from the validated envelope only."""

    envelope = GovernedAnalysisEnvelope.model_validate(envelope)
    analysis, source = envelope.governed_analysis, envelope.source_facts
    fact_map = _fact_map(envelope)
    output = BytesIO()
    styles = getSampleStyleSheet()
    title = ParagraphStyle("PepperynTitle", parent=styles["Title"], fontName="Helvetica-Bold",
                           fontSize=18, leading=22, textColor=colors.HexColor(f"#{_NAVY}"), alignment=TA_LEFT)
    heading = ParagraphStyle("PepperynHeading", parent=styles["Heading2"], fontName="Helvetica-Bold",
                             fontSize=12, leading=15, spaceBefore=8, spaceAfter=5,
                             textColor=colors.HexColor(f"#{_BLUE}"))
    body = ParagraphStyle("PepperynBody", parent=styles["BodyText"], fontName="Helvetica",
                          fontSize=9, leading=13, textColor=colors.HexColor(f"#{_TEXT}"))
    small = ParagraphStyle("PepperynSmall", parent=body, fontSize=7.5, leading=10)
    table_header = ParagraphStyle("PepperynTableHeader", parent=small, fontName="Helvetica-Bold",
                                  textColor=colors.white)

    def p(value: object, style=body) -> Paragraph:
        return Paragraph(escape(str(value)).replace("\n", "<br/>"), style)

    def section(label: str) -> list[object]:
        return [Paragraph(label, heading)]

    story: list[object] = [Paragraph("Pepperyn - Analyse financiere gouvernee", title), Spacer(1, 3 * mm)]
    metadata = [
        [p("Perimetre", small), p("Donnees synthetiques uniquement", small)],
        [p("Periode", small), p(source.current_period or "UNKNOWN", small)],
        [p("Statut de comprehension", small), p(source.status, small)],
        [p("Empreinte source SHA-256", small), p(source.source_representation_sha256, small)],
    ]
    table = Table(metadata, colWidths=[43 * mm, 137 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_PALE}")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(f"#{_TEXT}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CAD6DF")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([table, *section("Diagnostic - inference"), p(analysis.executive_diagnosis),
                  p("Faits cites: " + _refs(analysis.diagnosis_fact_ids, fact_map), small)])

    story.extend(section("Faits sources"))
    fact_rows = [[p(x, table_header) for x in ("Fact ID", "Metrique", "Valeur", "Unite", "Periode", "Source")]]
    for fact in source.facts:
        fact_rows.append([p(fact.fact_id, small), p(fact.metric, small), p(fact.value, small), p(fact.unit, small),
                          p(fact.period, small), p(f"{fact.source_sheet_ref}/{fact.source_field}", small)])
    facts_table = Table(fact_rows, repeatRows=1, colWidths=[25*mm, 31*mm, 23*mm, 16*mm, 19*mm, 52*mm])
    facts_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CAD6DF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F9FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.extend([facts_table, *section("Observations gouvernees")])
    for item in analysis.observations:
        story.extend([p(f"Observation source-matched: {item.metric} = {item.observed_value}"),
                      p(f"Severite inferentielle: {item.severity}"),
                      p("Fait cite: " + _refs((item.fact_id,), fact_map), small), Spacer(1, 2*mm)])
    story.extend(section("Inferences et validations"))
    for item in analysis.dimension_assessments:
        story.extend([p(f"Dimension {item.scope} - score inferentiel {item.score}/10 (confiance {item.confidence}%)"),
                      p(item.rationale), p("Faits cites: " + _refs(item.fact_ids, fact_map), small),
                      p("Validations requises: " + "; ".join(item.validation_required), small), Spacer(1, 2*mm)])
    for item in analysis.inferences:
        story.extend([p(f"Inference (confiance {item.confidence}%): {item.statement}"),
                      p("Faits cites: " + _refs(item.fact_ids, fact_map), small),
                      p("Validations requises: " + "; ".join(item.validation_required), small), Spacer(1, 2*mm)])

    story.extend(section("UNKNOWN et contradictions"))
    if not analysis.unknowns and not source.unknowns and not analysis.contradictions:
        story.append(p("Aucun element declare."))
    for item in analysis.unknowns:
        story.append(p(f"UNKNOWN ({item.materiality}): {item.question}"))
    for item in source.unknowns:
        story.append(p(f"UNDERSTANDING UNKNOWN: {item}"))
    for item in analysis.contradictions:
        story.extend([p("CONTRADICTION: " + item.statement), p("Faits cites: " + _refs(item.fact_ids, fact_map), small)])

    story.extend(section("Recommandations proposees"))
    for item in analysis.recommendations:
        story.extend([p(f"{item.priority} - {item.action}"), p(item.rationale),
                      p("Faits cites: " + (_refs(item.fact_ids, fact_map) if item.fact_ids else "Aucun"), small),
                      p("Prerequis: " + ("; ".join(item.prerequisite_validation) or "Aucun"), small), Spacer(1, 2*mm)])
    story.extend([Spacer(1, 3*mm), p("Les recommandations IA ne constituent pas des decisions confirmees.", small)])

    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm,
                            topMargin=14*mm, bottomMargin=14*mm,
                            title="Pepperyn - Analyse financiere gouvernee")
    doc.build(story)
    return output.getvalue()
