"""
PEPPERYN QUALITY CONTRACT — LAYOUT BASELINE v1.0

Ce test garantit que le layout du Financial Model Excel reste identique
entre les versions. Toute déviation par rapport aux constantes de
config/excel_layout.py constitue une régression.

RÈGLE : Ces tests ne peuvent être modifiés que par décision CTO explicite.
"""
import io
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, "/sessions/modest-sleepy-hypatia/mnt/Pepperyn/backend")

from models.schemas import AnalysisResult
from tests.test_rule_001_zero_manual_intervention import OPTILUX
from services.excel_export import generate_excel_report
from config.excel_layout import (
    ROW_H_HEADER, ROW_H_SECTION, ROW_H_COL_HDR, ROW_H_NAV,
    COL_END_DASHBOARD, COL_END_HYPOTHESES, COL_END_DECISION_LAB,
    COL_END_SENSIBILITE, COL_END_SCENARIOS, COL_END_ROADMAP, COL_END_HISTORIQUE,
    COLS_DASHBOARD, COLS_HYPOTHESES, COLS_DECISION_LAB,
    COLS_SENSIBILITE, COLS_SCENARIOS, COLS_ROADMAP, COLS_HISTORIQUE,
)


# ─── FIXTURE ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def wb():
    """Génère le workbook une seule fois pour toute la suite."""
    data = generate_excel_report(AnalysisResult(), OPTILUX, "Optilux SAS")
    return load_workbook(io.BytesIO(data))


def _col_width(ws, col_letter: str) -> float:
    """Retourne la largeur de colonne telle que définie par le renderer."""
    dim = ws.column_dimensions.get(col_letter)
    return dim.width if dim else 8.0  # fallback Excel default


# ─── HELPER ───────────────────────────────────────────────────────────────────

def assert_cols(ws, expected: dict, sheet_name: str):
    """Vérifie que chaque colonne a exactement la largeur attendue."""
    for col_key, expected_w in expected.items():
        letter = col_key if isinstance(col_key, str) else get_column_letter(col_key)
        actual_w = _col_width(ws, letter)
        assert actual_w == pytest.approx(expected_w, abs=0.01), (
            f"[{sheet_name}] Colonne {letter} : "
            f"attendu {expected_w}u, obtenu {actual_w}u — "
            f"layout déviation détectée. Modifier config/excel_layout.py pour "
            f"officialiser ce changement."
        )


# ─── TESTS LARGEURS DE COLONNES ───────────────────────────────────────────────

class TestColumnWidthsBaseline:

    def test_dashboard_column_widths(self, wb):
        ws = wb["📊 Dashboard"]
        assert_cols(ws, COLS_DASHBOARD, "Dashboard")

    def test_hypotheses_column_widths(self, wb):
        ws = wb["⚙ Hypothèses"]
        assert_cols(ws, COLS_HYPOTHESES, "Hypothèses")

    def test_decision_lab_column_widths(self, wb):
        ws = wb["🎯 Decision Lab"]
        assert_cols(ws, COLS_DECISION_LAB, "Decision Lab")

    def test_sensibilite_column_widths(self, wb):
        ws = wb["📈 Sensibilité"]
        assert_cols(ws, COLS_SENSIBILITE, "Sensibilité")

    def test_scenarios_column_widths(self, wb):
        ws = wb["📉 Scénarios"]
        assert_cols(ws, COLS_SCENARIOS, "Scénarios")

    def test_roadmap_column_widths(self, wb):
        ws = wb["🗺 Roadmap"]
        assert_cols(ws, COLS_ROADMAP, "Roadmap")

    def test_historique_column_widths(self, wb):
        ws = wb["🕒 Historique"]
        assert_cols(ws, COLS_HISTORIQUE, "Historique")


# ─── TESTS HAUTEURS DE LIGNES STANDARD ────────────────────────────────────────

class TestRowHeightsBaseline:

    def _row_h(self, ws, row: int) -> float:
        dim = ws.row_dimensions.get(row)
        return dim.height if dim else 15.0

    def test_header_row_height_dashboard(self, wb):
        ws = wb["📊 Dashboard"]
        assert self._row_h(ws, 1) == pytest.approx(ROW_H_HEADER, abs=0.5), \
            f"Dashboard row 1 (header) doit être {ROW_H_HEADER}pt"

    def test_nav_row_height_dashboard(self, wb):
        ws = wb["📊 Dashboard"]
        assert self._row_h(ws, 3) == pytest.approx(ROW_H_NAV, abs=0.5), \
            f"Dashboard row 3 (nav) doit être {ROW_H_NAV}pt"

    def test_header_row_height_roadmap(self, wb):
        ws = wb["🗺 Roadmap"]
        assert self._row_h(ws, 1) == pytest.approx(ROW_H_HEADER, abs=0.5), \
            f"Roadmap row 1 (header) doit être {ROW_H_HEADER}pt"

    def test_col_headers_row_height_roadmap(self, wb):
        ws = wb["🗺 Roadmap"]
        assert self._row_h(ws, 4) == pytest.approx(ROW_H_COL_HDR, abs=0.5), \
            f"Roadmap row 4 (col headers) doit être {ROW_H_COL_HDR}pt"

    def test_header_row_height_historique(self, wb):
        ws = wb["🕒 Historique"]
        assert self._row_h(ws, 1) == pytest.approx(ROW_H_HEADER, abs=0.5), \
            f"Historique row 1 (header) doit être {ROW_H_HEADER}pt"


# ─── TESTS COL_END (ALIGNEMENT DES BARRES) ────────────────────────────────────

class TestColEndBaseline:
    """
    Vérifie que les barres (header, nav, section) sont fusionnées exactement
    jusqu'à la dernière colonne de contenu — ni plus, ni moins.
    """

    def _merged_end_col(self, ws, row: int) -> int:
        """Retourne le numéro de la dernière colonne de la plage fusionnée sur cette ligne."""
        for merge in ws.merged_cells.ranges:
            if merge.min_row == row and merge.min_col == 1:
                return merge.max_col
        return 1

    def test_dashboard_header_col_end(self, wb):
        ws = wb["📊 Dashboard"]
        assert self._merged_end_col(ws, 1) == COL_END_DASHBOARD, \
            f"Dashboard header doit fusionner jusqu'à col {COL_END_DASHBOARD}"

    def test_hypotheses_header_col_end(self, wb):
        ws = wb["⚙ Hypothèses"]
        assert self._merged_end_col(ws, 1) == COL_END_HYPOTHESES, \
            f"Hypothèses header doit fusionner jusqu'à col {COL_END_HYPOTHESES}"

    def test_decision_lab_header_col_end(self, wb):
        ws = wb["🎯 Decision Lab"]
        assert self._merged_end_col(ws, 1) == COL_END_DECISION_LAB, \
            f"Decision Lab header doit fusionner jusqu'à col {COL_END_DECISION_LAB}"

    def test_sensibilite_header_col_end(self, wb):
        ws = wb["📈 Sensibilité"]
        assert self._merged_end_col(ws, 1) == COL_END_SENSIBILITE, \
            f"Sensibilité header doit fusionner jusqu'à col {COL_END_SENSIBILITE}"

    def test_scenarios_header_col_end(self, wb):
        ws = wb["📉 Scénarios"]
        assert self._merged_end_col(ws, 1) == COL_END_SCENARIOS, \
            f"Scénarios header doit fusionner jusqu'à col {COL_END_SCENARIOS}"

    def test_roadmap_header_col_end(self, wb):
        ws = wb["🗺 Roadmap"]
        assert self._merged_end_col(ws, 1) == COL_END_ROADMAP, \
            f"Roadmap header doit fusionner jusqu'à col {COL_END_ROADMAP}"

    def test_historique_header_col_end(self, wb):
        ws = wb["🕒 Historique"]
        assert self._merged_end_col(ws, 1) == COL_END_HISTORIQUE, \
            f"Historique header doit fusionner jusqu'à col {COL_END_HISTORIQUE}"
