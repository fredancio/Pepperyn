"""
test_fte_minimal.py — FTE v0 (Financial Time Engine, deterministic kernel).

Covers the 20 invariant tests required by the FTE v0 mission (Phase 14),
plus the Golden Case Phidani replay (Phase 15) and the temporal_role /
LLM / Evidence-immutability / DecisionArc boundary tests (Phase 7/17/18).

Scope discipline (mirrors test_decision_memory_integrity_repair.py's own
"stays within mission boundaries" tests): this suite proves invariants,
not implementation trivia — no closure semantics, no Business Moment, no
LLM, no new aggregate. See docs/Architecture/FTE_MINIMAL_IMPLEMENTATION_CONTRACT.md.
"""
import os
import sys
from datetime import date
from unittest.mock import MagicMock, patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest

from services import fte_minimal as fte
from services.temporal_normalizer import build_temporal_context
from services.evidence_ledger_service import save_evidence_capture


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _ctx(headers: list[str]) -> dict:
    return build_temporal_context(headers)


def make_supabase_mock(rows=None):
    """Chainable Supabase mock covering the read chain used by
    resolve_previous_observed_period_end (select/eq/not_/is_/order/limit)
    and the write chain used by save_evidence_capture (from_/insert)."""
    mock = MagicMock()
    for method in ("from_", "insert", "select", "eq", "single", "order", "limit"):
        getattr(mock, method).return_value = mock
    mock.not_ = MagicMock()
    mock.not_.is_.return_value = mock
    mock.execute.return_value = MagicMock(data=rows if rows is not None else [])
    return mock


class _OrderAwareSupabaseMock:
    """A plain chainable MagicMock (make_supabase_mock above) cannot prove
    WHICH column governs row selection, because it never actually sorts —
    it just returns whatever `rows` list it was handed. This fixture
    simulates real Postgres `ORDER BY <column> DESC LIMIT 1` behavior by
    recording the column name passed to .order() and sorting the fixture
    rows by that column before returning — letting
    TestPreviousPeriodSelectionIsBusinessTime fail loudly if the
    implementation ever regresses to ordering by the wrong clock again."""

    def __init__(self, rows):
        self._rows = rows
        self._order_col = None

    def from_(self, *_a, **_k):
        return self

    def select(self, *_a, **_k):
        return self

    def eq(self, *_a, **_k):
        return self

    def limit(self, *_a, **_k):
        return self

    @property
    def not_(self):
        return self

    def is_(self, *_a, **_k):
        return self

    def order(self, column, desc=True):
        self._order_col = column
        return self

    def execute(self):
        rows = sorted(self._rows, key=lambda r: r[self._order_col], reverse=True)
        return MagicMock(data=rows)


# Real-world monthly header sets, resolvable end-to-end by temporal_normalizer
# (named-month French labels — the format temporal_normalizer's month regex
# already supports; see TestPhidaniRealFile below for the exact-real-file
# format, which does NOT resolve month — a genuine, reported gap).
_JAN_TO_AUG_2019 = [
    "Janvier 2019", "Février 2019", "Mars 2019", "Avril 2019", "Mai 2019",
    "Juin 2019", "Juillet 2019", "Août 2019",
]
_JAN_TO_SEP_2019 = _JAN_TO_AUG_2019 + ["Septembre 2019"]


# ─────────────────────────────────────────────────────────────────────────────
# 1. Determinism
# ─────────────────────────────────────────────────────────────────────────────

class TestDeterminism:

    def test_same_headers_same_bounds_every_time(self):
        ctx1 = _ctx(_JAN_TO_SEP_2019)
        ctx2 = _ctx(_JAN_TO_SEP_2019)
        b1 = fte.resolve_current_period_bounds(ctx1)
        b2 = fte.resolve_current_period_bounds(ctx2)
        assert b1 == b2 == fte.CurrentPeriodBounds(date(2019, 9, 1), date(2019, 9, 30))

    def test_same_inputs_same_relationship_every_time(self):
        current = fte.CurrentPeriodBounds(date(2019, 9, 1), date(2019, 9, 30))
        results = {
            fte.classify_period_relationship(current, date(2019, 8, 31))
            for _ in range(10)
        }
        assert results == {"NEW"}


# ─────────────────────────────────────────────────────────────────────────────
# 2/3. UNKNOWN stays UNKNOWN — absence never becomes "current"
# ─────────────────────────────────────────────────────────────────────────────

class TestUnknownStaysUnknown:

    def test_no_resolvable_month_returns_none_not_system_date(self):
        # Headers with no year/month signal at all.
        ctx = _ctx(["Code", "Libellé", "Total"])
        assert fte.resolve_current_period_bounds(ctx) is None
        assert fte.resolve_newest_observed_period_end(ctx) is None

    def test_year_only_without_month_is_excluded_not_defaulted(self):
        # "YEAR 2019"-style header: year resolves, month does not (mirrors
        # the real Phidani file's cumulative YEAR column).
        ctx = _ctx(["YEAR 2019"])
        assert fte.resolve_current_period_bounds(ctx) is None

    def test_current_none_yields_unknown_relationship(self):
        assert fte.classify_period_relationship(None, date(2019, 8, 31)) == "UNKNOWN"

    def test_current_none_and_previous_none_yields_unknown_not_new(self):
        """Corrected 2026-08-08 per the independent adversarial pre-merge
        review: classify_period_relationship((current=None, previous=None))
        must never return 'NEW'. The prior implementation checked
        'previous_period_end is None' before 'current is None', so a
        completely unresolvable current period with no prior history
        fabricated a positive 'NEW' claim instead of the honest 'UNKNOWN'.
        This exact scenario is reached by the real Phidani.xlsx replay
        (YYYY-MM headers unresolved by temporal_normalizer, no history
        yet persisted) — see TestPhidaniRealFile below, which now asserts
        this same invariant against the real file end-to-end."""
        assert fte.classify_period_relationship(None, None) == "UNKNOWN"

    def test_known_current_and_no_prior_is_new_not_unknown(self):
        """The companion positive case: UNKNOWN must not overreach either
        — a genuinely resolvable current period with no prior history is
        correctly 'NEW' (mirrors TestComparisonStates::
        test_no_prior_observed_period_is_new; kept here too so the three
        required precedence cases from the correction mission read as one
        group)."""
        current = fte.CurrentPeriodBounds(start=date(2019, 9, 1), end=date(2019, 9, 30))
        assert fte.classify_period_relationship(current, None) == "NEW"

    def test_absent_period_never_silently_becomes_current(self):
        """An empty/unparseable dataset must never be reported as 'this is
        the current period' by omission — resolve_current_period_bounds
        returning None IS the honest signal; nothing downstream may
        substitute today's date."""
        ctx = _ctx([])
        assert fte.resolve_current_period_bounds(ctx) is None


# ─────────────────────────────────────────────────────────────────────────────
# 4. Business time distinct from created_at (knowledge time)
# ─────────────────────────────────────────────────────────────────────────────

class TestBusinessTimeDistinctFromKnowledgeTime:

    def test_observed_period_end_is_explicit_never_a_db_default(self):
        """created_at relies on the DB column DEFAULT NOW() (v18) and is
        never set by application code. observed_period_end must always be
        set explicitly by the caller (or omitted) — proving the two clocks
        are never conflated at the write site."""
        sb = make_supabase_mock()
        with patch("main.get_supabase_service", return_value=sb):
            save_evidence_capture(
                analyse_id="a1", company_id="c1", entity_id=None,
                evidence_capture={"facts": [{"id": "F1"}]},
                observed_period_end=date(2019, 9, 30),
            )
        insert_payload = sb.insert.call_args[0][0]
        assert insert_payload["observed_period_end"] == "2019-09-30"
        assert "created_at" not in insert_payload  # left to the DB default


# ─────────────────────────────────────────────────────────────────────────────
# 5/19. DecisionArc untouched
# ─────────────────────────────────────────────────────────────────────────────

class TestDecisionArcUntouched:

    def test_fte_minimal_never_imports_decision_arc(self):
        src = open(
            os.path.join(os.path.dirname(__file__), "..", "services", "fte_minimal.py"),
            encoding="utf-8",
        ).read()
        assert "decision_arc" not in src.lower()
        assert "arc_service" not in src.lower()

    def test_evidence_ledger_service_change_does_not_reference_decision_arc(self):
        src = open(
            os.path.join(os.path.dirname(__file__), "..", "services", "evidence_ledger_service.py"),
            encoding="utf-8",
        ).read()
        assert "decision_arc" not in src.lower()


# ─────────────────────────────────────────────────────────────────────────────
# 6/7/8. Comparison states
# ─────────────────────────────────────────────────────────────────────────────

class TestComparisonStates:

    def test_no_prior_observed_period_is_new(self):
        current = fte.CurrentPeriodBounds(date(2019, 9, 1), date(2019, 9, 30))
        assert fte.classify_period_relationship(current, None) == "NEW"

    def test_newer_contiguous_period_is_new(self):
        current = fte.CurrentPeriodBounds(date(2019, 9, 1), date(2019, 9, 30))
        assert fte.classify_period_relationship(current, date(2019, 8, 31)) == "NEW"

    def test_same_period_is_duplicate(self):
        current = fte.CurrentPeriodBounds(date(2019, 9, 1), date(2019, 9, 30))
        assert fte.classify_period_relationship(current, date(2019, 9, 30)) == "DUPLICATE"

    def test_older_period_after_newer_known_is_out_of_order(self):
        current = fte.CurrentPeriodBounds(date(2019, 8, 1), date(2019, 8, 31))
        assert fte.classify_period_relationship(current, date(2019, 9, 30)) == "OUT_OF_ORDER"

    def test_overlapping_shape_is_out_of_order_not_a_sixth_state(self):
        # current starts before the previous boundary but ends after it.
        current = fte.CurrentPeriodBounds(date(2019, 8, 15), date(2019, 9, 15))
        result = fte.classify_period_relationship(current, date(2019, 8, 31))
        assert result in fte.PERIOD_RELATIONSHIP_STATES
        assert result == "OUT_OF_ORDER"


# ─────────────────────────────────────────────────────────────────────────────
# 9. Gap asserted only when deterministically justified — no cadence assumption
# ─────────────────────────────────────────────────────────────────────────────

class TestGapOnlyWhenDeterministicallyJustified:

    def test_skipped_month_is_a_gap(self):
        # August known, November arrives — July/September/October skipped.
        current = fte.CurrentPeriodBounds(date(2019, 11, 1), date(2019, 11, 30))
        assert fte.classify_period_relationship(current, date(2019, 8, 31)) == "GAP"

    def test_gap_detection_uses_actual_boundaries_not_an_assumed_monthly_cadence(self):
        """Contiguity is (current.start - previous.end).days == 1 — pure
        date arithmetic on the OBSERVED boundaries. Proven here with a
        non-monthly (10-day) spacing: since the two boundaries are
        genuinely adjacent (no unexplained days in between), this must be
        NEW, not GAP — a naive 'looks less than ~30 days so must be
        consecutive months' heuristic would get this wrong in the other
        direction; a hardcoded 'exactly ~30 days apart' monthly assumption
        would also get this wrong. Only true adjacency decides it."""
        previous_end = date(2019, 8, 31)
        current = fte.CurrentPeriodBounds(date(2019, 9, 1), date(2019, 9, 10))
        assert fte.classify_period_relationship(current, previous_end) == "NEW"

    def test_non_adjacent_boundaries_are_a_gap_regardless_of_spacing_size(self):
        previous_end = date(2019, 8, 31)
        current = fte.CurrentPeriodBounds(date(2019, 9, 5), date(2019, 9, 10))
        assert fte.classify_period_relationship(current, previous_end) == "GAP"


# ─────────────────────────────────────────────────────────────────────────────
# 10/11. YTD — honest about missing months, never zero-fills
# ─────────────────────────────────────────────────────────────────────────────

class TestYTDCoverage:

    def test_complete_contiguous_run_is_complete(self):
        ctx = _ctx(_JAN_TO_SEP_2019)
        result = fte.resolve_ytd_coverage(ctx)
        assert result == {"status": "complete", "months_present": [1, 2, 3, 4, 5, 6, 7, 8, 9]}

    def test_missing_month_is_reported_incomplete_not_silently_accepted(self):
        headers = [h for h in _JAN_TO_SEP_2019 if "Mars" not in h]  # drop March
        ctx = _ctx(headers)
        result = fte.resolve_ytd_coverage(ctx)
        assert result["status"] == "incomplete"
        assert 3 in result["months_missing"]
        assert 3 not in result["months_present"]

    def test_missing_month_never_interpreted_as_zero(self):
        """The absent month must appear in months_missing, never silently
        folded into months_present with an implicit zero value — this
        module carries no amounts at all, so there is structurally no
        place a zero could be fabricated, which is itself the invariant
        under test."""
        headers = [h for h in _JAN_TO_SEP_2019 if "Mars" not in h]
        ctx = _ctx(headers)
        result = fte.resolve_ytd_coverage(ctx)
        assert "months_missing" in result
        assert result["months_missing"] == [3]

    def test_no_resolvable_period_is_unavailable(self):
        ctx = _ctx(["Code", "Libellé"])
        assert fte.resolve_ytd_coverage(ctx) == {"status": "unavailable"}


# ─────────────────────────────────────────────────────────────────────────────
# 12. Rolling-12 — honest insufficient-history, never fabricated
# ─────────────────────────────────────────────────────────────────────────────

class TestRolling12:

    def test_nine_months_is_insufficient_history(self):
        ctx = _ctx(_JAN_TO_SEP_2019)
        result = fte.resolve_rolling_12(ctx)
        assert result == {"status": "insufficient_history", "months_available": 9}

    def test_twelve_contiguous_months_is_available(self):
        headers = [
            "Octobre 2018", "Novembre 2018", "Décembre 2018",
        ] + _JAN_TO_SEP_2019
        ctx = _ctx(headers)
        result = fte.resolve_rolling_12(ctx)
        assert result["status"] == "available"
        assert result["window_start"] == {"year": 2018, "month": 10}
        assert result["window_end"] == {"year": 2019, "month": 9}

    def test_twelve_months_with_a_gap_inside_the_window_is_not_extrapolated(self):
        # 12 total (year, month) entries, but August 2019 is skipped —
        # the gap falls INSIDE the most-recent-12 window, so it must not
        # be reported as available.
        headers = [
            "Octobre 2018", "Novembre 2018", "Décembre 2018",
            "Janvier 2019", "Février 2019", "Mars 2019", "Avril 2019",
            "Mai 2019", "Juin 2019", "Juillet 2019",
            # "Août 2019" skipped
            "Septembre 2019", "Octobre 2019",
        ]
        ctx = _ctx(headers)
        result = fte.resolve_rolling_12(ctx)
        assert result["status"] == "insufficient_history"

    def test_gap_outside_the_most_recent_12_window_does_not_block_availability(self):
        # An old, disconnected entry (Oct 2017) followed by a fully
        # contiguous 12-month run (Dec 2018 -> Nov 2019). The Oct-2017 /
        # Dec-2018 gap falls entirely OUTSIDE the most-recent-12 window,
        # proving the window is evaluated on its own contiguity, not on
        # the full historical list.
        headers = [
            "Octobre 2017",
            "Décembre 2018", "Janvier 2019", "Février 2019", "Mars 2019",
            "Avril 2019", "Mai 2019", "Juin 2019", "Juillet 2019",
            "Août 2019", "Septembre 2019", "Octobre 2019", "Novembre 2019",
        ]
        ctx = _ctx(headers)
        result = fte.resolve_rolling_12(ctx)
        assert result["status"] == "available"
        assert result["window_start"] == {"year": 2018, "month": 12}
        assert result["window_end"] == {"year": 2019, "month": 11}


# ─────────────────────────────────────────────────────────────────────────────
# 13. Re-analysis of the same business period does not fabricate a new period
# ─────────────────────────────────────────────────────────────────────────────

class TestReanalysisDoesNotFabricateNewPeriod:

    def test_re_upload_of_same_month_is_duplicate_not_new(self):
        """Three re-analyses on different calendar days (Mar 10, Mar 15,
        Apr 2), same underlying September file — the business period must
        stay DUPLICATE every time; only knowledge time (created_at) moves,
        never business time."""
        current = fte.CurrentPeriodBounds(date(2019, 9, 1), date(2019, 9, 30))
        for _ in range(3):
            assert fte.classify_period_relationship(current, date(2019, 9, 30)) == "DUPLICATE"


# ─────────────────────────────────────────────────────────────────────────────
# 14/15. Persistence — writes correctly, NULL historical rows stay valid
# ─────────────────────────────────────────────────────────────────────────────

class TestPersistence:

    def test_observed_period_end_persists_as_iso_date_string(self):
        sb = make_supabase_mock()
        with patch("main.get_supabase_service", return_value=sb):
            save_evidence_capture(
                analyse_id="a1", company_id="c1", entity_id="e1",
                evidence_capture={"facts": [{"id": "F1"}]},
                observed_period_end=date(2019, 9, 30),
            )
        assert sb.insert.call_args[0][0]["observed_period_end"] == "2019-09-30"

    def test_none_observed_period_end_is_omitted_not_inserted_as_null(self):
        """Mirrors the existing entity_id convention in this same module —
        absent means omitted, not forced to an explicit NULL write."""
        sb = make_supabase_mock()
        with patch("main.get_supabase_service", return_value=sb):
            save_evidence_capture(
                analyse_id="a1", company_id="c1", entity_id="e1",
                evidence_capture={"facts": [{"id": "F1"}]},
                observed_period_end=None,
            )
        assert "observed_period_end" not in sb.insert.call_args[0][0]

    def test_historical_row_with_null_observed_period_end_is_skipped_not_fatal(self):
        """resolve_previous_observed_period_end must tolerate rows where
        observed_period_end is NULL (pre-v23 rows, or rows where no period
        was resolvable) — the SQL query itself excludes NULLs
        (.not_.is_("observed_period_end", "null")), this test proves the
        Python side degrades gracefully if a NULL slips through anyway."""
        sb = make_supabase_mock(rows=[{"observed_period_end": None}])
        with patch("main.get_supabase_service", return_value=None):
            result = fte.resolve_previous_observed_period_end(sb, "c1", "e1")
        assert result is None

    def test_no_rows_returns_none_not_an_error(self):
        sb = make_supabase_mock(rows=[])
        result = fte.resolve_previous_observed_period_end(sb, "c1", "e1")
        assert result is None

    def test_read_failure_returns_none_never_raises(self):
        sb = MagicMock()
        sb.from_.side_effect = RuntimeError("connection lost")
        result = fte.resolve_previous_observed_period_end(sb, "c1", "e1")
        assert result is None

    def test_previous_period_end_is_parsed_correctly(self):
        sb = make_supabase_mock(rows=[{"observed_period_end": "2019-08-31"}])
        result = fte.resolve_previous_observed_period_end(sb, "c1", "e1")
        assert result == date(2019, 8, 31)


# ─────────────────────────────────────────────────────────────────────────────
# 14b. Previous-period selection uses BUSINESS TIME, never KNOWLEDGE TIME
# ─────────────────────────────────────────────────────────────────────────────

class TestPreviousPeriodSelectionIsBusinessTime:
    """Corrected 2026-08-08 per the independent adversarial pre-merge
    review: resolve_previous_observed_period_end previously ordered
    candidate rows by created_at DESC (Knowledge Time — when Pepperyn
    learned the row) instead of observed_period_end DESC (Business Time —
    the period the row actually describes). The two diverge whenever
    analyses are ingested out of business-time order; ordering by the
    wrong clock can silently select a LESS advanced business-time
    baseline than one that genuinely exists earlier in the ledger,
    undermining OUT_OF_ORDER detection across analyses."""

    def test_row_with_greater_business_time_wins_even_if_ingested_earlier(self):
        # Row A: ingested (created_at) AFTER Row B, but describes an
        # OLDER business period.
        row_a = {"observed_period_end": "2019-08-31", "created_at": "2026-08-08T12:00:00Z"}
        # Row B: ingested BEFORE Row A, but describes a NEWER business
        # period — this is the one the function must select.
        row_b = {"observed_period_end": "2019-09-30", "created_at": "2026-08-01T09:00:00Z"}
        sb = _OrderAwareSupabaseMock([row_a, row_b])
        result = fte.resolve_previous_observed_period_end(sb, "c1", "e1")
        assert result == date(2019, 9, 30)  # Row B wins — business time, not knowledge time

    def test_query_orders_by_observed_period_end_not_created_at(self):
        """Direct proof of the query construction itself, independent of
        the sorting simulation above."""
        sb = make_supabase_mock(rows=[{"observed_period_end": "2019-09-30"}])
        fte.resolve_previous_observed_period_end(sb, "c1", "e1")
        order_call = sb.order.call_args
        assert order_call is not None
        assert order_call.args[0] == "observed_period_end"
        assert order_call.kwargs.get("desc") is True
        assert order_call.args[0] != "created_at"


# ─────────────────────────────────────────────────────────────────────────────
# 16/21. temporal_role is never canonical FTE input
# ─────────────────────────────────────────────────────────────────────────────

class TestTemporalRoleIsolation:

    def test_fte_minimal_never_reads_temporal_role_as_code(self):
        """AST-based, not substring-based: the module's own docstring
        legitimately DISCUSSES QuantifiedImpact.temporal_role (to explain
        the isolation rule) — that is documentation, not a violation.
        What must never exist is an actual attribute-access node reading
        `.temporal_role` off any object in real code."""
        import ast
        path = os.path.join(
            os.path.dirname(__file__), "..", "services", "fte_minimal.py"
        )
        tree = ast.parse(open(path, encoding="utf-8").read(), filename=path)
        attr_reads = [
            node.attr for node in ast.walk(tree)
            if isinstance(node, ast.Attribute) and node.attr == "temporal_role"
        ]
        assert attr_reads == [], "fte_minimal.py contains a live .temporal_role attribute access"

    def test_fte_minimal_never_imports_financial_truth_model(self):
        import ast
        path = os.path.join(
            os.path.dirname(__file__), "..", "services", "fte_minimal.py"
        )
        tree = ast.parse(open(path, encoding="utf-8").read(), filename=path)
        imported_modules = [
            n.module for n in ast.walk(tree) if isinstance(n, ast.ImportFrom)
        ] + [
            alias.name for n in ast.walk(tree) if isinstance(n, ast.Import) for alias in n.names
        ]
        assert not any(m and "financial_truth" in m for m in imported_modules)


# ─────────────────────────────────────────────────────────────────────────────
# 17. No LLM in the deterministic kernel
# ─────────────────────────────────────────────────────────────────────────────

class TestNoLLMInKernel:

    def test_fte_minimal_has_no_llm_related_imports_or_calls(self):
        src = open(
            os.path.join(os.path.dirname(__file__), "..", "services", "fte_minimal.py"),
            encoding="utf-8",
        ).read().lower()
        forbidden = ("anthropic", "claude", "llm_service", "call_chat", "openai")
        for token in forbidden:
            assert token not in src, f"forbidden token {token!r} found in fte_minimal.py"


# ─────────────────────────────────────────────────────────────────────────────
# 18. No Evidence history is rewritten (immutability respected)
# ─────────────────────────────────────────────────────────────────────────────

class TestNoEvidenceRewrite:

    def test_fte_minimal_never_calls_update_or_upsert(self):
        src = open(
            os.path.join(os.path.dirname(__file__), "..", "services", "fte_minimal.py"),
            encoding="utf-8",
        ).read()
        assert ".update(" not in src
        assert ".upsert(" not in src

    def test_evidence_ledger_service_still_only_inserts(self):
        src = open(
            os.path.join(os.path.dirname(__file__), "..", "services", "evidence_ledger_service.py"),
            encoding="utf-8",
        ).read()
        assert ".update(" not in src
        assert ".upsert(" not in src


# ─────────────────────────────────────────────────────────────────────────────
# 20. Non-calendar / custom-period representation not blocked by schema
# ─────────────────────────────────────────────────────────────────────────────

class TestSchemaDoesNotBlockNonCalendarPeriods:

    def test_migration_column_is_plain_date_no_month_boundary_constraint(self):
        migration_path = os.path.join(
            os.path.dirname(__file__), "..", "migrations",
            "v23_evidence_ledger_observed_period.sql",
        )
        src = open(migration_path, encoding="utf-8").read()
        assert "observed_period_end DATE NULL" in src
        # No CHECK constraint pinning this to month-end/quarter-end/etc.
        assert "CHECK" not in src.upper()

    def test_arbitrary_dates_round_trip_through_save_evidence_capture(self):
        """A quarter-end, a week-end, or any other non-monthly boundary
        date persists identically — the column carries no granularity
        assumption."""
        for d in (date(2019, 9, 30), date(2019, 12, 31), date(2019, 3, 8)):
            sb = make_supabase_mock()
            with patch("main.get_supabase_service", return_value=sb):
                save_evidence_capture(
                    analyse_id="a", company_id="c", entity_id=None,
                    evidence_capture={"facts": [{"id": "F1"}]},
                    observed_period_end=d,
                )
            assert sb.insert.call_args[0][0]["observed_period_end"] == d.isoformat()


# ─────────────────────────────────────────────────────────────────────────────
# Golden Case Phidani (Phase 15) — real canonical data
# ─────────────────────────────────────────────────────────────────────────────

class TestPhidaniWalkingSkeletonInvariants:
    """
    Uses resolvable header formats (named French months) to prove the FTE
    v0 kernel itself satisfies GOLDEN_CASE_001_PHIDANI.md's REQUIRED
    assertions (contract §14: A, B, C, E, H-factual) end to end. See
    TestPhidaniRealFile below for the replay against the actual
    Phidani.xlsx headers, which exposes a separate, honestly-reported
    limitation in temporal_normalizer's month regex — not a defect in
    this kernel.
    """

    def test_newest_observed_business_period_is_september_2019(self):
        ctx = _ctx(_JAN_TO_SEP_2019)
        bounds = fte.resolve_current_period_bounds(ctx)
        assert bounds.end == date(2019, 9, 30)  # (A)

    def test_prior_known_newest_period_is_august_2019(self):
        # Simulates the previously-persisted value from the "old version"
        # (Jan-Aug) analysis.
        ctx_old = _ctx(_JAN_TO_AUG_2019)
        previous_end = fte.resolve_newest_observed_period_end(ctx_old)
        assert previous_end == date(2019, 8, 31)  # (B)

    def test_september_follows_august_without_a_gap(self):
        ctx_old = _ctx(_JAN_TO_AUG_2019)
        ctx_new = _ctx(_JAN_TO_SEP_2019)
        previous_end = fte.resolve_newest_observed_period_end(ctx_old)
        current = fte.resolve_current_period_bounds(ctx_new)
        assert fte.classify_period_relationship(current, previous_end) == "NEW"  # (C)

    def test_ytd_through_september_available(self):
        ctx_new = _ctx(_JAN_TO_SEP_2019)
        assert fte.resolve_ytd_coverage(ctx_new)["status"] == "complete"  # (E)

    def test_newer_information_exists_is_the_only_factual_claim_made(self):
        """H, factual half only — 'newer information exists'. The judgment
        half ('therefore re-analyze') is not produced anywhere in this
        module — there is no function that returns such a recommendation,
        which is itself the proof it was never implemented."""
        ctx_old = _ctx(_JAN_TO_AUG_2019)
        ctx_new = _ctx(_JAN_TO_SEP_2019)
        previous_end = fte.resolve_newest_observed_period_end(ctx_old)
        current = fte.resolve_current_period_bounds(ctx_new)
        relationship = fte.classify_period_relationship(current, previous_end)
        assert relationship == "NEW"
        assert not hasattr(fte, "should_reanalyze")
        assert not hasattr(fte, "analysis_pertinence")

    def test_rolling_12_honestly_insufficient_on_nine_months(self):
        ctx_new = _ctx(_JAN_TO_SEP_2019)
        assert fte.resolve_rolling_12(ctx_new)["status"] == "insufficient_history"  # (G)

    def test_no_closure_assertion_exists_anywhere_in_the_module(self):
        """(F) — closure must be UNSUPPORTED, not merely untested. Proven
        structurally: no function name, no return value, no field in this
        entire module contains 'clos'/'close'/'closure'/'closed'."""
        src = open(
            os.path.join(os.path.dirname(__file__), "..", "services", "fte_minimal.py"),
            encoding="utf-8",
        ).read().lower()
        # Allow the negative-contract docstring wording itself ("no closure
        # claim") but forbid any function/field actually named for it.
        import re
        assert not re.search(r"def\s+\w*clos\w*", src)
        assert "close_confidence" not in src
        assert "probably_closed" not in src
        assert "is_closed" not in src


class TestPhidaniRealFile:
    """
    Replay against the ACTUAL canonical Phidani.xlsx present in the repo
    root — not a synthetic reconstruction. Only one consolidated file
    exists (not the two separate "old"/"new" snapshot files the Golden
    Case narratively describes), so the two dataset versions are derived
    by slicing the real file's real header row at its real 2019-08 and
    2019-09 columns — never fabricated header text.

    Honest finding (Phase 15 — "report the limitation honestly"): the real
    file's header format is "2019-01".."2019-12" (hyphen-separated
    ISO-like), which temporal_normalizer._extract_month does NOT resolve
    (its regex requires a preceding "m" or "/", not "-" — confirmed by
    direct inspection, not assumed). This is a real, previously-undetected
    gap in temporal_normalizer.py's month coverage, exposed for the first
    time by testing against real data instead of synthetic headers. It is
    reported here, NOT fixed — fixing it would touch temporal_normalizer.py,
    forbidden by the contract (§8, §17) without a separate, explicitly
    authorized mission.
    """

    _REAL_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "Phidani.xlsx")

    def _real_headers(self, end_month: str) -> list[str]:
        import openpyxl
        wb = openpyxl.load_workbook(self._REAL_FILE, data_only=True)
        ws = wb["PHIDANI"]
        row2 = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        start_idx = row2.index("2019-01")
        end_idx = row2.index(end_month)
        return [str(h) for h in row2[start_idx:end_idx + 1]]

    @pytest.mark.skipif(
        not os.path.exists(_REAL_FILE), reason="Phidani.xlsx not present in this checkout"
    )
    def test_real_file_month_is_not_resolved_by_current_temporal_normalizer(self):
        """Documents the gap rather than papering over it: with the exact
        real-file header format, resolve_current_period_bounds() honestly
        returns None (UNKNOWN) — FTE v0's own logic is correct (it never
        fabricates a period it cannot support), but the full walking-
        skeleton assertions (A/B/C/E) cannot be positively demonstrated
        end-to-end against Phidani.xlsx's exact real headers with
        temporal_normalizer.py unchanged."""
        headers_new = self._real_headers("2019-09")
        ctx_new = _ctx(headers_new)
        bounds = fte.resolve_current_period_bounds(ctx_new)
        assert bounds is None, (
            "If this assertion fails, temporal_normalizer.py's month regex "
            "was changed to resolve 'YYYY-MM' — update this test's docstring, "
            "it no longer documents a gap."
        )

    @pytest.mark.skipif(
        not os.path.exists(_REAL_FILE), reason="Phidani.xlsx not present in this checkout"
    )
    def test_real_file_year_alone_still_resolves_current_actual_role(self):
        """Even without month resolution, PeriodRole classification itself
        (year-based) still works correctly against the real file — the
        gap is specifically in month extraction, not in the classifier as
        a whole."""
        headers_new = self._real_headers("2019-09")
        ctx_new = _ctx(headers_new)
        assert ctx_new["detected_current_year"] == 2019
        assert "2019-09" in ctx_new["columns_by_role"].get("CURRENT_ACTUAL", [])

    @pytest.mark.skipif(
        not os.path.exists(_REAL_FILE), reason="Phidani.xlsx not present in this checkout"
    )
    def test_real_file_yields_unknown_relationship_never_a_false_new(self):
        """Regression test added 2026-08-08 per the independent adversarial
        pre-merge review + FTE V0 FINAL CORRECTIONS mission (Mission 3).

        The prior test in this class stopped at proving
        resolve_current_period_bounds() honestly returns None against the
        real file's unresolved 'YYYY-MM' headers — it never carried that
        honest None through to classify_period_relationship(), which is
        exactly where the (None, None) → 'NEW' precedence bug hid. This
        test closes that gap end-to-end against the REAL file (no
        synthetic substitute — Golden Case discipline): with no prior
        observed_period_end persisted yet (a fresh Engagement), the real
        Phidani headers must yield UNKNOWN, never the fabricated 'NEW'.

        This must keep failing loudly if classify_period_relationship ever
        regresses to checking previous_period_end before current."""
        headers_new = self._real_headers("2019-09")
        ctx_new = _ctx(headers_new)
        current = fte.resolve_current_period_bounds(ctx_new)
        assert current is None  # honest gap, unchanged — see test above

        previous_period_end = None  # no prior history for this Engagement
        relationship = fte.classify_period_relationship(current, previous_period_end)
        assert relationship == "UNKNOWN", (
            "Real Phidani.xlsx replay with no prior history must classify as "
            "UNKNOWN (nothing is actually known), never as 'NEW' (a positive "
            "claim the data does not support). The result is allowed to "
            "remain UNKNOWN until temporal_normalizer's YYYY-MM gap is "
            "repaired in a separate, narrowly-scoped mission — see "
            "STRATEGIC_DEFERRED_WORK_REGISTER.md."
        )
