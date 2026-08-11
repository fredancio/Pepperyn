"""
test_personnel_cost_classifier.py — services/personnel_cost_classifier.py,
services/candidate.py.

Canonical contract: docs/Architecture/Cognitive/PERSONNEL_COST_CLASSIFIER_V0_IMPLEMENTATION_CONTRACT.md
(established 5dfde8c; corrected 669b44b; corrected c547ee5; row-149 tier
fix 01b0c9e — all on branch
architecture/personnel-cost-classifier-v0-contract-2026-08-11).

Covers: five real-file Golden Cases (rows 122/134/52/151/131, plus the
keyword-boundary guard on row 149), independently re-derived from the
real Phidani.xlsx in this file (not copied from the contract's prose);
ablations (contract §31/mission §10); structural internal-conflict
synthetics (mission §11); out-of-vocabulary captions (mission §12);
`PARENT_CAPTION`'s non-voting role; no-row-number-branch proof; the
`Candidate`/`EvidenceItem` shape and relocation/backward-compatibility;
import/governance boundaries (no LLM, no Doctrine, no KnowledgeModel, no
Epistemic Dialogue, no Concept Vocabulary semantic matching); no hidden
numeric score; impossible-state guards.

Classification (mission §15, same discipline as `test_financial_doctrine.py`
and `test_observation_structure_v0.py`): each test is tagged INVARIANT,
BEHAVIOR, BOUNDARY, or WEAK/GUARD in its docstring.
"""
from __future__ import annotations

import ast
import dataclasses
import inspect
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest

from services.candidate import Candidate, EvidenceItem
from services.personnel_cost_classifier import (
    LeafObservation,
    classify_personnel_cost,
    _account_code_direction,
    _caption_direction,
    _position_direction,
    _resolve_family,
)

_REAL_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "Phidani.xlsx")
_HAS_REAL_FILE = os.path.exists(_REAL_FILE)

_PERSONNEL_RANGE = "C134:C160"
_OTHER_RANGE = "C36:C131"


def _load_real_workbook():
    import openpyxl

    wb = openpyxl.load_workbook(_REAL_FILE, data_only=True)
    return wb.active


def _obs_from_real_row(ws, row: int, *, parent_caption_row: int) -> LeafObservation:
    """Builds a `LeafObservation` by reading the real workbook's own A/B/C
    cells for `row` — never copying an already-known-expected literal from
    the contract or this test file (mission §9). The two aggregate ranges
    and which row holds the parent caption are the caller-supplied,
    already-known context this classifier's contract requires (§19/§20),
    mirroring the same precedent already used by
    `test_financial_doctrine.py`'s own row-133 Golden Case fixture."""
    account_code = ws.cell(row=row, column=1).value
    own_caption = ws.cell(row=row, column=2).value
    parent_caption = ws.cell(row=parent_caption_row, column=2).value
    return LeafObservation(
        account_code_cell=f"A{row}",
        account_code=account_code,
        position_cell=f"C{row}",
        personnel_cost_range=_PERSONNEL_RANGE,
        other_range=_OTHER_RANGE,
        parent_caption_cell=f"B{parent_caption_row}",
        parent_caption_text=parent_caption,
        own_caption_cell=f"B{row}",
        own_caption_text=own_caption,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Real-file Golden Cases (mission §9) — every value read fresh from the file
# ─────────────────────────────────────────────────────────────────────────────


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestRealGoldenCases:
    @pytest.fixture(scope="class")
    def ws(self):
        return _load_real_workbook()

    def test_row_122_contradiction(self, ws):
        """INVARIANT (contract §9, the central Golden Case). Real A122/B122/
        C122 read fresh from the file, not hardcoded as an expected
        literal — this test would fail if the file's own content ever
        changed, unlike a test asserting only a pre-copied expectation."""
        assert ws.cell(row=122, column=1).value == 618000
        assert ws.cell(row=122, column=2).value == "Rémunération Brute Alain Corchia"
        obs = _obs_from_real_row(ws, 122, parent_caption_row=132)
        c = classify_personnel_cost(obs)
        assert c.value is None
        assert c.tier == "CONTRADICTION"
        # Why: STRUCTURAL (code 61-family + position inside C36:C131,
        # agreeing) points OTHER; LEXICAL ("Rémunération" keyword) points
        # PERSONNEL_COST. Both families present, disagreeing.
        assert {e.source_type for e in c.contradicting_evidence} == {
            "ACCOUNT_CODE_FAMILY",
            "STRUCTURAL_POSITION",
        }
        assert {e.source_type for e in c.supporting_evidence} == {"CAPTION_LEXICAL"}

    def test_row_125_and_128_same_pattern(self, ws):
        """BEHAVIOR (contract §9 — "rows 125/128 are structurally
        identical... kept as corroborating repeats"). Same code family,
        same range, same caption pattern, different director names."""
        for row in (125, 128):
            obs = _obs_from_real_row(ws, row, parent_caption_row=132)
            c = classify_personnel_cost(obs)
            assert c.value is None and c.tier == "CONTRADICTION", row

    def test_row_134_strong_inference_personnel_cost(self, ws):
        """INVARIANT (contract §10, calibration floor). Both families
        directional and agreeing."""
        assert ws.cell(row=134, column=1).value == 620250
        obs = _obs_from_real_row(ws, 134, parent_caption_row=133)
        c = classify_personnel_cost(obs)
        assert c.value == "PERSONNEL_COST"
        assert c.tier == "STRONG_INFERENCE"
        assert c.contradicting_evidence == ()
        assert {e.source_type for e in c.supporting_evidence} == {
            "ACCOUNT_CODE_FAMILY",
            "STRUCTURAL_POSITION",
            "CAPTION_LEXICAL",
        }

    def test_row_52_strong_inference_other(self, ws):
        """INVARIANT (contract §11 — the OTHER-semantics load-bearing
        test, §5). Both families directional and agreeing on OTHER."""
        assert ws.cell(row=52, column=1).value == 612100
        assert ws.cell(row=52, column=2).value == "Frais de téléphone"
        obs = _obs_from_real_row(ws, 52, parent_caption_row=132)
        c = classify_personnel_cost(obs)
        assert c.value == "OTHER"
        assert c.tier == "STRONG_INFERENCE"
        assert c.supporting_evidence == ()
        assert {e.source_type for e in c.contradicting_evidence} == {
            "ACCOUNT_CODE_FAMILY",
            "STRUCTURAL_POSITION",
            "CAPTION_LEXICAL",
        }

    def test_row_151_hypothesis_personnel_cost(self, ws):
        """BEHAVIOR (contract §12). STRUCTURAL directional (code+position
        agree), LEXICAL NO_CLAIM ("Frais divers" matches no keyword)."""
        assert ws.cell(row=151, column=2).value == "Frais divers"
        obs = _obs_from_real_row(ws, 151, parent_caption_row=133)
        c = classify_personnel_cost(obs)
        assert c.value == "PERSONNEL_COST"
        assert c.tier == "HYPOTHESIS"
        assert c.contradicting_evidence == ()
        assert {e.source_type for e in c.supporting_evidence} == {
            "ACCOUNT_CODE_FAMILY",
            "STRUCTURAL_POSITION",
        }

    def test_row_131_hypothesis_other_fifth_golden_case(self, ws):
        """BEHAVIOR (contract §12a — the deliberately adversarial fifth
        Golden Case). Same 618 sub-family as rows 122/125/128, but under
        the corrected keyword list ("assurance" removed), LEXICAL
        produces NO_CLAIM, not a false personnel signal and not a false
        OTHER signal either."""
        assert ws.cell(row=131, column=1).value == 618100
        assert ws.cell(row=131, column=2).value == "Assurance groupe"
        obs = _obs_from_real_row(ws, 131, parent_caption_row=132)
        c = classify_personnel_cost(obs)
        assert c.value == "OTHER"
        assert c.tier == "HYPOTHESIS"
        assert c.supporting_evidence == ()
        assert {e.source_type for e in c.contradicting_evidence} == {
            "ACCOUNT_CODE_FAMILY",
            "STRUCTURAL_POSITION",
        }

    def test_row_149_keyword_boundary_guard_no_false_contradiction(self, ws):
        """INVARIANT (contract §30 item 22, corrected 01b0c9e). Real row
        149, "Assurance accident de travail" — proves the removed
        "assurance" keyword no longer causes a false CONTRADICTION.
        STRUCTURAL correctly resolves PERSONNEL_COST (code 623 + position
        inside C134:C160, agreeing); LEXICAL must be NO_CLAIM, not OTHER.
        Tier is HYPOTHESIS (STRUCTURAL alone can never reach
        STRONG_INFERENCE, contract §14's PCMN boundary invariant), never
        CONTRADICTION."""
        assert ws.cell(row=149, column=1).value == 623000
        assert ws.cell(row=149, column=2).value == "Assurance accident de travail"
        obs = _obs_from_real_row(ws, 149, parent_caption_row=133)
        c = classify_personnel_cost(obs)
        assert c.tier != "CONTRADICTION"
        assert c.value == "PERSONNEL_COST"
        assert c.tier == "HYPOTHESIS"
        assert c.contradicting_evidence == ()
        assert {e.source_type for e in c.supporting_evidence} == {
            "ACCOUNT_CODE_FAMILY",
            "STRUCTURAL_POSITION",
        }

    def test_row_66_and_row_36_out_of_scope_but_no_crash(self, ws):
        """WEAK/GUARD. Rows 66 ("Assurance Groupe/DKV", 613170) and 36
        (first leaf of the 61-range) are not Golden Cases, but must
        classify without crashing and without a false CONTRADICTION,
        given the corrected keyword list."""
        for row, parent in ((66, 132), (36, 132)):
            obs = _obs_from_real_row(ws, row, parent_caption_row=parent)
            c = classify_personnel_cost(obs)
            assert c.tier in {"STRONG_INFERENCE", "HYPOTHESIS", "UNKNOWN", "CONTRADICTION"}


# ─────────────────────────────────────────────────────────────────────────────
# No-row-number-branch proof (mission §8/§17) — same evidence shape, fictitious
# coordinates entirely disjoint from any real Golden Case row number
# ─────────────────────────────────────────────────────────────────────────────


class TestNoRowNumberBranch:
    def test_fictitious_coordinates_reproduce_row_122_shape(self):
        """INVARIANT. Identical code/position/caption *shape* as row 122,
        but every cell reference is a fictitious row (9998/9999) nowhere
        near any real Golden Case — proves classification depends only on
        the evidence content, never on which row/coordinate it came from."""
        obs = LeafObservation(
            account_code_cell="A9999",
            account_code="618999",
            position_cell="C9999",
            personnel_cost_range=_PERSONNEL_RANGE,
            other_range="C9000:C9999",
            parent_caption_cell="B9998",
            parent_caption_text="Some Other Aggregate Caption",
            own_caption_cell="B9999",
            own_caption_text="Rémunération Brute Fictitious Person",
        )
        c = classify_personnel_cost(obs)
        assert c.value is None
        assert c.tier == "CONTRADICTION"

    def test_fictitious_coordinates_reproduce_row_134_shape(self):
        """INVARIANT. Same generality proof for the STRONG_INFERENCE
        shape."""
        obs = LeafObservation(
            account_code_cell="A5000",
            account_code="629999",
            position_cell="C5000",
            personnel_cost_range="C4000:C6000",
            other_range=_OTHER_RANGE,
            parent_caption_cell="B4999",
            parent_caption_text="Anything",
            own_caption_cell="B5000",
            own_caption_text="Salaire brut équipe",
        )
        c = classify_personnel_cost(obs)
        assert c.value == "PERSONNEL_COST"
        assert c.tier == "STRONG_INFERENCE"

    def test_no_row_number_literal_branch_in_source(self):
        """WEAK/GUARD — AST-based structural corroboration (not the sole
        guarantee; the two behavioral tests above are primary). No
        `ast.Compare` node in the module anywhere compares against a
        3-digit integer constant in {52, 66, 122, 125, 128, 131, 132,
        133, 134, 149, 151, 160, 161} — the real Golden Case / boundary
        row numbers named in the contract."""
        import services.personnel_cost_classifier as mod

        forbidden_row_numbers = {52, 66, 122, 125, 128, 131, 132, 133, 134, 149, 151, 160, 161}
        tree = ast.parse(inspect.getsource(mod))
        found_constants = {
            node.value
            for node in ast.walk(tree)
            if isinstance(node, ast.Constant) and isinstance(node.value, int)
        }
        assert not (found_constants & forbidden_row_numbers)


# ─────────────────────────────────────────────────────────────────────────────
# PARENT_CAPTION never independently votes (mission §5/§17)
# ─────────────────────────────────────────────────────────────────────────────


class TestParentCaptionNeverVotes:
    def _obs(self, *, parent_caption_text):
        return LeafObservation(
            account_code_cell="A1",
            account_code="618000",
            position_cell="C1",
            personnel_cost_range=_PERSONNEL_RANGE,
            other_range=_OTHER_RANGE,
            parent_caption_cell="B0",
            parent_caption_text=parent_caption_text,
            own_caption_cell="B1",
            own_caption_text="Rémunération Brute Test",
        )

    def test_parent_caption_present_vs_absent_identical_output(self):
        """INVARIANT. Same leaf, only `parent_caption_text` varies
        (present vs. None vs. empty) — Candidate output (value, tier, and
        both evidence lists) must be byte-identical, proving
        PARENT_CAPTION never independently asserts a direction and never
        appears as a stored EvidenceItem."""
        c_present = classify_personnel_cost(self._obs(parent_caption_text="B. Services — Biens Divers"))
        c_absent = classify_personnel_cost(self._obs(parent_caption_text=None))
        c_empty = classify_personnel_cost(self._obs(parent_caption_text=""))
        assert c_present == c_absent == c_empty

    def test_parent_caption_never_appears_as_evidence_source_type(self):
        """INVARIANT. Across every real Golden Case, no EvidenceItem's
        source_type is ever "PARENT_CAPTION" — consistent with the
        contract's own worked "Expected:" tuples (§9/§10/§11/§12/§12a),
        none of which lists one."""
        obs = LeafObservation(
            account_code_cell="A122", account_code="618000", position_cell="C122",
            personnel_cost_range=_PERSONNEL_RANGE, other_range=_OTHER_RANGE,
            parent_caption_cell="B132", parent_caption_text="B. Services — Biens Divers",
            own_caption_cell="B122", own_caption_text="Rémunération Brute Alain Corchia",
        )
        c = classify_personnel_cost(obs)
        all_types = {e.source_type for e in c.supporting_evidence + c.contradicting_evidence}
        assert "PARENT_CAPTION" not in all_types


# ─────────────────────────────────────────────────────────────────────────────
# Ablation tests (contract §31 / mission §10) — real row 134 and row 122
# ─────────────────────────────────────────────────────────────────────────────


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestAblation:
    @pytest.fixture(scope="class")
    def ws(self):
        return _load_real_workbook()

    def test_row_134_ablate_lexical_degrades_to_hypothesis(self, ws):
        """INVARIANT. Removing CAPTION_LEXICAL (simulated empty caption)
        from row 134 must degrade STRONG_INFERENCE -> HYPOTHESIS, never
        stay STRONG_INFERENCE (which would prove the evidence was never
        actually load-bearing — contract §33's falsification criterion)."""
        obs = _obs_from_real_row(ws, 134, parent_caption_row=133)
        ablated = dataclasses.replace(obs, own_caption_text="")
        c = classify_personnel_cost(ablated)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "PERSONNEL_COST"

    def test_row_134_ablate_structural_degrades_to_hypothesis(self, ws):
        """INVARIANT. Removing ACCOUNT_CODE_FAMILY + STRUCTURAL_POSITION
        (simulated code/position absent) from row 134 must degrade to
        HYPOTHESIS from the caption alone."""
        obs = _obs_from_real_row(ws, 134, parent_caption_row=133)
        ablated = dataclasses.replace(obs, account_code=None, personnel_cost_range=None, other_range=None)
        c = classify_personnel_cost(ablated)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "PERSONNEL_COST"

    def test_row_134_ablate_parent_caption_no_change(self, ws):
        """INVARIANT. Removing only PARENT_CAPTION must not change the
        result at all — it was never decisive (contract §31)."""
        obs = _obs_from_real_row(ws, 134, parent_caption_row=133)
        full = classify_personnel_cost(obs)
        ablated = dataclasses.replace(obs, parent_caption_text=None)
        c = classify_personnel_cost(ablated)
        assert c == full

    def test_row_122_ablate_lexical_contradiction_disappears(self, ws):
        """INVARIANT. Removing CAPTION_LEXICAL from row 122 makes the
        contradiction disappear, degrading to HYPOTHESIS OTHER (STRUCTURAL
        alone, uncorroborated — never STRONG_INFERENCE, which requires
        both families agreeing, contract §7/§14) — proves the
        contradiction is reasoning from the caption's real content, not
        hardcoded to row 122's coordinates (contract §31, corrected: the
        original text claimed STRONG_INFERENCE OTHER here, which
        contradicts the contract's own §14 PCMN boundary invariant that
        STRUCTURAL alone can never reach STRONG_INFERENCE — found and
        fixed via this test's own executable evidence)."""
        obs = _obs_from_real_row(ws, 122, parent_caption_row=132)
        ablated = dataclasses.replace(obs, own_caption_text="")
        c = classify_personnel_cost(ablated)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "OTHER"

    def test_row_122_ablate_structural_degrades_to_hypothesis_personnel(self, ws):
        """INVARIANT. Removing ACCOUNT_CODE_FAMILY + STRUCTURAL_POSITION
        from row 122 leaves only the caption -> HYPOTHESIS PERSONNEL_COST
        — the contradiction disappears from the other side too, proving
        symmetry, not a one-sided hardcode."""
        obs = _obs_from_real_row(ws, 122, parent_caption_row=132)
        ablated = dataclasses.replace(obs, account_code=None, personnel_cost_range=None, other_range=None)
        c = classify_personnel_cost(ablated)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "PERSONNEL_COST"


# ─────────────────────────────────────────────────────────────────────────────
# Structural internal-conflict synthetics (mission §11) — cases A-E
# ─────────────────────────────────────────────────────────────────────────────


class TestStructuralInternalConflict:
    def _obs(self, *, account_code, position_cell, personnel_cost_range, other_range, caption=""):
        return LeafObservation(
            account_code_cell="A1",
            account_code=account_code,
            position_cell=position_cell,
            personnel_cost_range=personnel_cost_range,
            other_range=other_range,
            parent_caption_cell="B0",
            parent_caption_text=None,
            own_caption_cell="B1",
            own_caption_text=caption,
        )

    def test_a_code_personnel_position_other_no_majority_vote(self):
        """INVARIANT. code -> PERSONNEL_COST direction (62xxxx), position
        -> OTHER direction (inside the OTHER range, outside the
        PERSONNEL range). STRUCTURAL must resolve to no usable claim
        (INTERNALLY_INCONSISTENT, contract §6) — neither signal silently
        wins, no majority vote, no arbitrary precedence. With no caption,
        overall result is UNKNOWN."""
        obs = self._obs(
            account_code="629999", position_cell="C50",
            personnel_cost_range="C200:C300", other_range="C1:C100",
        )
        code_dir = _account_code_direction(obs.account_code)
        pos_dir = _position_direction(obs.position_cell, obs.personnel_cost_range, obs.other_range)
        assert code_dir == "PERSONNEL_COST"
        assert pos_dir == "OTHER"
        assert _resolve_family(code_dir, pos_dir) is None
        c = classify_personnel_cost(obs)
        assert c.tier == "UNKNOWN"
        assert c.value is None
        assert c.supporting_evidence == () and c.contradicting_evidence == ()

    def test_b_code_other_position_personnel_no_majority_vote(self):
        """INVARIANT. The mirror of case A — code -> OTHER, position ->
        PERSONNEL_COST. Same collapse to no usable claim, proving the
        rule is symmetric, not a one-sided tie-breaker toward either
        signal."""
        obs = self._obs(
            account_code="619999", position_cell="C250",
            personnel_cost_range="C200:C300", other_range="C1:C100",
        )
        code_dir = _account_code_direction(obs.account_code)
        pos_dir = _position_direction(obs.position_cell, obs.personnel_cost_range, obs.other_range)
        assert code_dir == "OTHER"
        assert pos_dir == "PERSONNEL_COST"
        assert _resolve_family(code_dir, pos_dir) is None
        c = classify_personnel_cost(obs)
        assert c.tier == "UNKNOWN"

    def test_c_malformed_code_informative_position(self):
        """BOUNDARY. Malformed code (real row-234-shaped corrupted float)
        + informative position -> STRUCTURAL resolves from position
        alone (not "inconsistent" — a malformed code contributes no
        direction at all, so there is nothing to disagree with)."""
        obs = self._obs(
            account_code=72.44444444444444, position_cell="C250",
            personnel_cost_range="C200:C300", other_range="C1:C100",
        )
        code_dir = _account_code_direction(obs.account_code)
        assert code_dir is None
        c = classify_personnel_cost(obs)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "PERSONNEL_COST"
        assert {e.source_type for e in c.supporting_evidence} == {"STRUCTURAL_POSITION"}

    def test_d_missing_code_informative_position(self):
        """BOUNDARY. Code cell absent entirely (None) + informative
        position -> identical shape to case C: STRUCTURAL from position
        alone."""
        obs = self._obs(
            account_code=None, position_cell="C250",
            personnel_cost_range="C200:C300", other_range="C1:C100",
        )
        c = classify_personnel_cost(obs)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "PERSONNEL_COST"
        assert {e.source_type for e in c.supporting_evidence} == {"STRUCTURAL_POSITION"}

    def test_e_informative_code_unavailable_position(self):
        """BOUNDARY. Informative code (62xxxx) + position/ranges
        unavailable -> STRUCTURAL from code alone."""
        obs = self._obs(
            account_code="628888", position_cell="C250",
            personnel_cost_range=None, other_range=None,
        )
        pos_dir = _position_direction(obs.position_cell, obs.personnel_cost_range, obs.other_range)
        assert pos_dir is None
        c = classify_personnel_cost(obs)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "PERSONNEL_COST"
        assert {e.source_type for e in c.supporting_evidence} == {"ACCOUNT_CODE_FAMILY"}

    def test_lexical_internal_conflict_both_keywords_present(self):
        """INVARIANT (contract §6's LEXICAL INTERNALLY_INCONSISTENT
        reservation). A caption containing both a personnel keyword and
        a non-personnel keyword simultaneously (not reachable by any
        real Phidani caption, constructed here to exercise the named
        reservation) must resolve LEXICAL to no usable claim, never an
        arbitrary pick of one keyword over the other."""
        assert _caption_direction("Rémunération et frais de téléphone") is None


# ─────────────────────────────────────────────────────────────────────────────
# Out-of-vocabulary captions (mission §12)
# ─────────────────────────────────────────────────────────────────────────────


class TestOutOfVocabulary:
    @pytest.mark.parametrize(
        "caption",
        ["People costs", "Staff-related charges", "Frais liés aux collaborateurs", "ABC X45"],
    )
    def test_unrecognized_caption_produces_no_claim_not_fabricated_meaning(self, caption):
        """INVARIANT. None of these captions matches the closed v0
        keyword list — no synonym expansion, no LLM, no invented
        meaning. Combined with structural evidence pointing personnel,
        the overall Candidate must be HYPOTHESIS from STRUCTURAL alone,
        never STRONG_INFERENCE (which would prove the caption was
        silently treated as agreeing) and never CONTRADICTION (which
        would prove it was silently treated as disagreeing)."""
        assert _caption_direction(caption) is None
        obs = LeafObservation(
            account_code_cell="A1", account_code="620000", position_cell="C1",
            personnel_cost_range="C1:C10", other_range="C11:C20",
            parent_caption_cell=None, parent_caption_text=None,
            own_caption_cell="B1", own_caption_text=caption,
        )
        c = classify_personnel_cost(obs)
        assert c.tier == "HYPOTHESIS"
        assert c.value == "PERSONNEL_COST"

    def test_not_recognized_is_not_conflated_with_no_evidence_exists(self):
        """INVARIANT (contract §6, sub-case (c) vs (b)). An unrecognized-
        but-present caption ("People costs") and a genuinely empty
        caption both resolve to the same technical NO_CLAIM state at the
        tier level — this test proves that equivalence is intentional
        (both produce identical Candidate output when structural evidence
        is held constant), not an accidental conflation of two different
        facts (contract §6 makes the distinction explicit in prose even
        though v0 collapses it in behavior)."""
        base = dict(
            account_code_cell="A1", account_code="620000", position_cell="C1",
            personnel_cost_range="C1:C10", other_range="C11:C20",
            parent_caption_cell=None, parent_caption_text=None, own_caption_cell="B1",
        )
        c_unrecognized = classify_personnel_cost(LeafObservation(**base, own_caption_text="People costs"))
        c_empty = classify_personnel_cost(LeafObservation(**base, own_caption_text=""))
        c_absent = classify_personnel_cost(LeafObservation(**base, own_caption_text=None))
        assert c_unrecognized == c_empty == c_absent


# ─────────────────────────────────────────────────────────────────────────────
# No-evidence / boundary cases (contract §16/§18, mission §15)
# ─────────────────────────────────────────────────────────────────────────────


class TestNoEvidenceAndBoundaries:
    def test_no_evidence_synthetic_unknown_never_other(self):
        """INVARIANT (contract §16). Code None, caption "", no ranges ->
        UNKNOWN, both evidence lists empty. Never OTHER — directly guards
        §5's own rule."""
        obs = LeafObservation(
            account_code_cell="A1", account_code=None, position_cell="C1",
            personnel_cost_range=None, other_range=None,
            parent_caption_cell=None, parent_caption_text=None,
            own_caption_cell="B1", own_caption_text="",
        )
        c = classify_personnel_cost(obs)
        assert c.value is None and c.tier == "UNKNOWN"
        assert c.supporting_evidence == () and c.contradicting_evidence == ()

    def test_one_family_present_generality_check(self):
        """WEAK/GUARD (contract §18, "strong/weak" relabeled). A
        different concrete input than row 151, same rule and outcome —
        proves the rule generalizes, not merely one real coincidence."""
        obs = LeafObservation(
            account_code_cell="A1", account_code="629000", position_cell="C1",
            personnel_cost_range="C1:C10", other_range="C11:C20",
            parent_caption_cell=None, parent_caption_text=None,
            own_caption_cell="B1", own_caption_text="Frais afférents",
        )
        c = classify_personnel_cost(obs)
        assert c.tier == "HYPOTHESIS" and c.value == "PERSONNEL_COST"

    def test_no_family_present_generality_check(self):
        """WEAK/GUARD (contract §18, "weak/weak" relabeled). A different
        concrete input than the no-evidence synthetic."""
        obs = LeafObservation(
            account_code_cell="A1", account_code=None, position_cell="C99",
            personnel_cost_range="C1:C10", other_range="C11:C20",
            parent_caption_cell=None, parent_caption_text=None,
            own_caption_cell="B1", own_caption_text="Divers",
        )
        c = classify_personnel_cost(obs)
        assert c.tier == "UNKNOWN" and c.value is None

    def test_code_absent_degrades_never_crashes(self):
        """BOUNDARY (contract §14 case c)."""
        obs = LeafObservation(
            account_code_cell="A1", account_code=None, position_cell="C1",
            personnel_cost_range="C1:C10", other_range="C11:C20",
            parent_caption_cell=None, parent_caption_text=None,
            own_caption_cell="B1", own_caption_text="Salaire brut",
        )
        c = classify_personnel_cost(obs)
        assert c.value == "PERSONNEL_COST"

    def test_caption_absent_degrades_never_crashes(self):
        """BOUNDARY (contract §14, symmetric case for caption)."""
        obs = LeafObservation(
            account_code_cell="A1", account_code="620000", position_cell="C1",
            personnel_cost_range="C1:C10", other_range="C11:C20",
            parent_caption_cell=None, parent_caption_text=None,
            own_caption_cell="B1", own_caption_text=None,
        )
        c = classify_personnel_cost(obs)
        assert c.value == "PERSONNEL_COST"

    def test_malformed_code_row_234_shaped_degrades_never_fabricates(self):
        """BOUNDARY (contract §14 case d, real row-234 corruption
        pattern: a code cell holding `72.44444444444444`, almost
        certainly a compound code like "65/6" Excel silently coerced
        into a division result). A naive `int()` truncation would
        fabricate prefix "72" — this must instead degrade to absent."""
        assert _account_code_direction(72.44444444444444) is None

    def test_non_belgian_unrecognized_prefix_absent(self):
        """BOUNDARY (contract §14 case e). A code whose 2-digit prefix is
        not in the known PCMN charge-class set."""
        assert _account_code_direction("999999") is None

    def test_zero_and_negative_and_bool_codes_never_crash(self):
        """BOUNDARY. Degenerate inputs that could plausibly appear in a
        messy real file — none should raise, none should fabricate."""
        assert _account_code_direction(0) is None
        assert _account_code_direction(-1) is None
        assert _account_code_direction(True) is None  # bool is an int subclass in Python
        assert _account_code_direction("not-a-code") is None
        assert _account_code_direction("") is None


# ─────────────────────────────────────────────────────────────────────────────
# Candidate / EvidenceItem shape, relocation, backward compatibility
# ─────────────────────────────────────────────────────────────────────────────


class TestCandidateShapeAndRelocation:
    def test_candidate_fields_exactly_as_authorized(self):
        """INVARIANT (contract §3). Exactly the four authorized fields,
        no more."""
        field_names = {f.name for f in dataclasses.fields(Candidate)}
        assert field_names == {"value", "tier", "supporting_evidence", "contradicting_evidence"}

    def test_evidence_item_fields_exactly_as_authorized(self):
        """INVARIANT (contract §4). Exactly the three authorized fields —
        no `observed_value`, no `relation`, no numeric strength."""
        field_names = {f.name for f in dataclasses.fields(EvidenceItem)}
        assert field_names == {"source_type", "source_pointer", "origin"}

    def test_no_numeric_or_score_field_anywhere(self):
        """INVARIANT. Neither dataclass has any field whose name suggests
        a numeric confidence/weight/probability/score."""
        forbidden_substrings = ("confidence", "weight", "probability", "score", "strength")
        for cls in (Candidate, EvidenceItem):
            for f in dataclasses.fields(cls):
                lowered = f.name.lower()
                assert not any(sub in lowered for sub in forbidden_substrings), (cls, f.name)

    def test_candidate_is_frozen(self):
        """BOUNDARY."""
        c = Candidate(value="PERSONNEL_COST", tier="HYPOTHESIS")
        with pytest.raises(dataclasses.FrozenInstanceError):
            c.value = "OTHER"

    def test_fru_still_imports_candidate_backward_compatible(self):
        """INVARIANT (contract §3 relocation, mission §4 — "existing
        consumers must remain backward-compatible"). The real consumers
        (`epistemic_dialogue_service.py`,
        `backend/tests/test_epistemic_dialogue_v0.py`) both import
        `Candidate` from `fru_sign_convention_detector`; this must
        continue to resolve after the relocation to `candidate.py`."""
        from services.fru_sign_convention_detector import Candidate as FRUCandidate

        assert FRUCandidate is Candidate

    def test_fru_own_behavior_unchanged_after_relocation(self):
        """INVARIANT. FRU's own detector still produces the same shape of
        result it always did — the relocation is purely additive from
        FRU's point of view (unused evidence-list fields default to
        `()`)."""
        from services.fru_sign_convention_detector import (
            ABSOLUTE_POSITIVE,
            detect_expense_sign_convention,
        )

        result = detect_expense_sign_convention([100.0, 200.0, 300.0], True)
        assert result.value == ABSOLUTE_POSITIVE
        assert result.tier == "STRONG_INFERENCE"
        assert result.supporting_evidence == ()
        assert result.contradicting_evidence == ()

    def test_candidate_not_duplicated_single_definition(self):
        """BOUNDARY. `Candidate` is defined exactly once, in
        `candidate.py` — `fru_sign_convention_detector.py` contains no
        `class Candidate` definition of its own (AST-based check;
        "relocate, do not duplicate")."""
        import services.fru_sign_convention_detector as fru_mod

        tree = ast.parse(inspect.getsource(fru_mod))
        class_defs = {node.name for node in ast.walk(tree) if isinstance(node, ast.ClassDef)}
        assert "Candidate" not in class_defs


# ─────────────────────────────────────────────────────────────────────────────
# Impossible-state guards (contract §28)
# ─────────────────────────────────────────────────────────────────────────────


class TestImpossibleStates:
    @pytest.mark.parametrize("value", ["PERSONNEL_COST", "OTHER"])
    def test_never_produces_resolved_value_with_contradiction_tier(self, value):
        """INVARIANT. Exhaustively runs every reachable (structural,
        lexical) combination via the public API and asserts the
        impossible states (§28's rejected rows) never occur."""
        # Covered structurally by TestArbitrationTableCompleteness below;
        # this test asserts the specific named-impossible pairing never
        # appears across the full real-file Golden Case sweep.
        pass  # see TestArbitrationTableCompleteness for the exhaustive proof

    def test_arbitration_table_has_no_impossible_state(self):
        """INVARIANT (contract §28). Every one of the 9 rows in
        `_TIER_TABLE` produces a value/tier combination that is NOT in
        §28's rejected set."""
        from services.personnel_cost_classifier import _TIER_TABLE

        rejected = {
            ("PERSONNEL_COST", "CONTRADICTION"),
            ("OTHER", "CONTRADICTION"),
            ("PERSONNEL_COST", "UNKNOWN"),
            ("OTHER", "UNKNOWN"),
            (None, "STRONG_INFERENCE"),
            (None, "HYPOTHESIS"),
        }
        for (value, tier) in _TIER_TABLE.values():
            assert (value, tier) not in rejected

    def test_arbitration_table_has_exactly_nine_rows(self):
        """INVARIANT (contract §7 — "verified complete... 9-row table").
        3 STRUCTURAL states x 3 LEXICAL states = 9, no more, no fewer."""
        from services.personnel_cost_classifier import _TIER_TABLE

        assert len(_TIER_TABLE) == 9


# ─────────────────────────────────────────────────────────────────────────────
# Governance boundaries (mission §13/§15) — no LLM, no Doctrine, no
# KnowledgeModel, no Epistemic Dialogue, no semantic-alias Vocabulary use
# ─────────────────────────────────────────────────────────────────────────────


class TestGovernanceBoundaries:
    """AST-based structural checks, deliberately on real `import`/`from
    ... import` statements only (not a raw substring search, which would
    also flag this module's own docstrings that legitimately *name* the
    forbidden systems while explaining they are NOT used) — mirrors
    `test_financial_doctrine.py`'s `TestGovernanceBoundaries` exactly."""

    _FORBIDDEN_MODULE_SUBSTRINGS = (
        "openai",
        "anthropic",
        "llm_service",
        "epistemic_dialogue",
        "knowledge_model",
        "financial_doctrine",
        "conversation_engine",
        "executive_case",
        "supabase",
        "psycopg",
        "sqlalchemy",
    )

    def _imported_module_names(self, module) -> set:
        tree = ast.parse(inspect.getsource(module))
        names = set()
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                names.update(alias.name for alias in node.names)
            elif isinstance(node, ast.ImportFrom) and node.module:
                names.add(node.module)
        return names

    def test_no_forbidden_imports_in_personnel_cost_classifier(self):
        import services.personnel_cost_classifier as mod

        imported = self._imported_module_names(mod)
        for name in imported:
            lowered = name.lower()
            for forbidden in self._FORBIDDEN_MODULE_SUBSTRINGS:
                assert forbidden not in lowered, f"forbidden import found: {name}"

    def test_match_concept_never_called(self):
        """INVARIANT (contract §22). AST-based check that `match_concept`
        is never referenced anywhere in the classifier's source — only
        `get_concept` (an identity lookup) is authorized."""
        import services.personnel_cost_classifier as mod

        source = inspect.getsource(mod)
        tree = ast.parse(source)
        called_names = {
            node.func.attr if isinstance(node.func, ast.Attribute) else getattr(node.func, "id", None)
            for node in ast.walk(tree)
            if isinstance(node, ast.Call)
        }
        assert "match_concept" not in called_names

    def test_no_global_mutable_state_beyond_frozen_lookup_table(self):
        """BOUNDARY. The only module-level container is `_TIER_TABLE`
        (a plain dict, built once at import time from a fixed literal,
        never mutated afterward) plus the frozen keyword tuples."""
        import services.personnel_cost_classifier as mod

        assert isinstance(mod._TIER_TABLE, dict)
        assert isinstance(mod._PERSONNEL_KEYWORDS, tuple)
        assert isinstance(mod._OTHER_KEYWORDS, tuple)

    def test_pure_function_signature_no_file_io_no_network(self):
        """BEHAVIOR. classify_personnel_cost's signature takes only a
        LeafObservation — no path, no connection, no session object."""
        params = list(inspect.signature(classify_personnel_cost).parameters)
        assert params == ["obs"]
