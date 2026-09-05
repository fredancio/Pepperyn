from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader

from sandbox.governed_exports import generate_governed_excel, generate_governed_pdf
from services.v1_analysis_contract import GovernedAnalysisEnvelope


def _envelope(*, action: str = "Valider le tableau des flux.") -> GovernedAnalysisEnvelope:
    fact_id = "FABCDEF123456"
    source_hash = "A" * 64
    return GovernedAnalysisEnvelope.model_validate({
        "source_facts": {
            "status": "UNDERSTOOD", "current_period": "2025",
            "source_representation_sha256": source_hash,
            "facts": [{"fact_id": fact_id, "metric": "EBITDA", "value": -145000,
                       "unit": "EUR", "period": "2025", "source_sheet_ref": "SABCDEF123456",
                       "source_field": "R123456ABCDEF"}], "unknowns": [],
        },
        "governed_analysis": {
            "source_representation_sha256": source_hash, "invocation_nonce": "B" * 32,
            "executive_diagnosis": "La rentabilite operationnelle est negative.",
            "diagnosis_fact_ids": [fact_id],
            "observations": [{"fact_id": fact_id, "metric": "EBITDA", "observed_value": -145000,
                              "severity": "HIGH"}],
            "dimension_assessments": [{"scope": "PROFITABILITY", "score": 2,
                "rationale": "L'EBITDA degrade la rentabilite.", "fact_ids": [fact_id],
                "confidence": 90, "validation_required": ["Confirmer les retraitements."]}],
            "inferences": [{"statement": "Une tension de tresorerie est possible.",
                            "fact_ids": [fact_id], "confidence": 60,
                            "validation_required": ["Valider le tableau des flux."]}],
            "unknowns": [{"question": "Quel est le flux de tresorerie ?", "materiality": "HIGH"}],
            "contradictions": [],
            "recommendations": [{"priority": "P1", "action": action,
                "rationale": "La traduction en tresorerie n'est pas etablie.", "fact_ids": [fact_id],
                "prerequisite_validation": ["Obtenir le tableau des flux."]}],
        },
    })


def test_excel_is_structured_auditable_and_contains_no_formulas():
    workbook = load_workbook(BytesIO(generate_governed_excel(_envelope())), data_only=False)
    assert workbook.sheetnames == ["Synthese", "Faits sources", "Inferences", "Recommandations", "UNKNOWN"]
    values = [cell.value for sheet in workbook for row in sheet.iter_rows() for cell in row if cell.value is not None]
    text = "\n".join(map(str, values))
    for required in ("Diagnostic (inference)", "Faits sources",
                     "Observation source-matched - severite inferentielle HIGH", "EBITDA = -145000",
                     "Validations requises", "Prerequis",
                     "Les recommandations IA ne constituent pas des decisions confirmees."):
        assert required in text or required in workbook.sheetnames
    assert not [cell for sheet in workbook for row in sheet.iter_rows() for cell in row if cell.data_type == "f"]


def test_provider_text_cannot_become_an_excel_formula():
    workbook = load_workbook(BytesIO(generate_governed_excel(
        _envelope(action='=HYPERLINK("https://invalid.example","click")')
    )), data_only=False)
    action = workbook["Recommandations"]["B2"]
    assert action.data_type == "s"
    assert action.value.startswith("'=HYPERLINK")


def test_pdf_contains_governed_sections_and_no_confirmed_decision():
    reader = PdfReader(BytesIO(generate_governed_pdf(_envelope())))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    for required in ("Diagnostic - inference", "Faits sources", "Observations gouvernees",
                     "Observation source-matched: EBITDA = -145000", "Severite inferentielle: HIGH",
                     "Inferences et validations",
                     "UNKNOWN et contradictions", "Recommandations proposees",
                     "ne constituent pas des decisions confirmees"):
        assert required in text
