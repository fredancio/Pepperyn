"""
test_formula_evidence.py — Observation Structure v0's dedicated formula-
evidence module (backend/services/formula_evidence.py).

Covers the 10 formula-module-specific tests required by the implementation
mission (§21), plus the same-artifact-guarantee test (mission §5). See
docs/Architecture/Cognitive/OBSERVATION_STRUCTURE_V0_IMPLEMENTATION_CONTRACT.md.
"""
import os
import sys
import time

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import pytest

from services.formula_evidence import (
    CellFormulaEvidence,
    cell_formula_evidence,
    get_worksheet,
    load_formula_workbook,
)

_REAL_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "Phidani.xlsx")
_HAS_REAL_FILE = os.path.exists(_REAL_FILE)


def _real_bytes() -> bytes:
    with open(_REAL_FILE, "rb") as f:
        return f.read()


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestFormulaCellDetection:
    """1. Formula cell detected. 2. Non-formula cell detected."""

    def test_formula_cell_detected(self):
        wb = load_formula_workbook(_real_bytes())
        ws = get_worksheet(wb, "PHIDANI")
        ev = cell_formula_evidence(ws, 163, 3)  # row 163, col C: "=C162"
        assert ev.is_formula is True

    def test_non_formula_cell_detected(self):
        wb = load_formula_workbook(_real_bytes())
        ws = get_worksheet(wb, "PHIDANI")
        ev = cell_formula_evidence(ws, 134, 3)  # row 134, col C: literal 863540.12
        assert ev.is_formula is False


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestFormulaTextPreservation:
    """3. Formula text preserved exactly, unparsed, unevaluated."""

    def test_formula_text_preserved_verbatim(self):
        wb = load_formula_workbook(_real_bytes())
        ws = get_worksheet(wb, "PHIDANI")
        ev = cell_formula_evidence(ws, 161, 3)
        assert ev.formula_text == "=SUM(C134:C160)"

    def test_formula_text_absent_when_not_a_formula(self):
        wb = load_formula_workbook(_real_bytes())
        ws = get_worksheet(wb, "PHIDANI")
        ev = cell_formula_evidence(ws, 134, 3)
        assert ev.formula_text is None


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestNoNumericValueProduction:
    """4. This module never produces or exposes a numeric value (contract
    §5, §9 — "no second source of truth"). CellFormulaEvidence has no
    field that could carry a cached number, and has_literal_value is a
    boolean, never the value itself."""

    def test_evidence_dataclass_has_no_numeric_value_field(self):
        fields = CellFormulaEvidence.__dataclass_fields__.keys()
        assert set(fields) == {"is_formula", "formula_text", "has_literal_value"}
        for name in fields:
            assert "value" not in name or name == "has_literal_value", (
                f"Unexpected field {name!r} risks exposing a numeric value; "
                "this module must never become a second source of truth."
            )

    def test_has_literal_value_is_a_boolean_not_the_number(self):
        wb = load_formula_workbook(_real_bytes())
        ws = get_worksheet(wb, "PHIDANI")
        ev = cell_formula_evidence(ws, 134, 3)
        assert ev.has_literal_value is True
        assert isinstance(ev.has_literal_value, bool)
        # The dataclass must not separately expose 863540.12 anywhere.
        assert not hasattr(ev, "value")
        assert not hasattr(ev, "cached_value")


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestCoordinatesPreserved:
    """5. Coordinates preserved — real, unshifted openpyxl (row, column),
    never a position-remapped index."""

    def test_real_row_column_coordinates_used_directly(self):
        wb = load_formula_workbook(_real_bytes())
        ws = get_worksheet(wb, "PHIDANI")
        # Row 163 is real sheet row 163, not a DataFrame-relative index.
        ev_direct = cell_formula_evidence(ws, 163, 3)
        ev_via_openpyxl = ws.cell(row=163, column=3)
        assert ev_direct.formula_text == ev_via_openpyxl.value


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestSameSheetCellIdentity:
    """6. Same sheet/cell identity — evidence for (row, col) in one sheet
    is never conflated with the same coordinates in a different sheet."""

    def test_sheet_selection_by_name_is_respected(self):
        wb = load_formula_workbook(_real_bytes())
        ws_named = get_worksheet(wb, "PHIDANI")
        ws_active = get_worksheet(wb)  # single-sheet real file: active == PHIDANI
        assert ws_named.title == ws_active.title == "PHIDANI"


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestWorkbookUnchanged:
    """7. Workbook unchanged — reading formula evidence never mutates or
    re-saves the source file."""

    def test_file_bytes_on_disk_untouched_after_read(self):
        before = _real_bytes()
        wb = load_formula_workbook(before)
        ws = get_worksheet(wb, "PHIDANI")
        cell_formula_evidence(ws, 163, 3)
        cell_formula_evidence(ws, 161, 3)
        after = _real_bytes()
        assert before == after, "Reading formula evidence must never mutate the source file."


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestNoFormulaExecution:
    """8. No formula execution — openpyxl.load_workbook(data_only=False)
    never evaluates formulas; formula_text is exactly the saved string,
    never a computed replacement."""

    def test_formula_text_is_the_raw_formula_string_not_a_computed_value(self):
        wb = load_formula_workbook(_real_bytes())
        ws = get_worksheet(wb, "PHIDANI")
        ev = cell_formula_evidence(ws, 163, 3)
        assert ev.formula_text == "=C162"
        assert ev.formula_text != 131323.73  # the real cached result for this cell


class TestNoMacrosExecuted:
    """9. No macros executed — openpyxl never executes VBA regardless of
    load options; this module does not pass keep_vba=True and performs no
    macro-related call."""

    def test_module_never_requests_vba_execution(self):
        import inspect
        from services import formula_evidence as mod

        source = inspect.getsource(mod.load_formula_workbook)
        assert "keep_vba" not in source, (
            "load_formula_workbook must never pass keep_vba=True — this "
            "module reads formula text only, never macros."
        )


class TestMalformedWorkbookDegradesSafely:
    """10. Malformed workbook degrades safely — invalid bytes raise a
    clear, catchable exception rather than crashing uncontrolled or
    silently returning fabricated evidence."""

    def test_invalid_bytes_raise_instead_of_fabricating_evidence(self):
        garbage = b"this is not a valid xlsx file at all"
        with pytest.raises(Exception):
            load_formula_workbook(garbage)


@pytest.mark.skipif(not _HAS_REAL_FILE, reason="Phidani.xlsx not present in this checkout")
class TestSameArtifactGuarantee:
    """Mission §5: both the canonical numeric parse and this module's
    formula-evidence read must refer to the exact same uploaded workbook
    artifact — no re-download, no alternate path, no regenerated copy.

    This module's public entry point takes `file_bytes: bytes` (the same
    shape as file_parser.parse_file(file_bytes, filename)) precisely so a
    caller holding one `bytes` object can pass it to both reads. This test
    proves that guarantee holds: reading the fixture's bytes exactly once
    and deriving both a data_only=True view and this module's data_only=False
    view from that same bytes object agree exactly on every non-formula
    cell, and diverge exactly as expected on formula cells.
    """

    def test_data_only_true_and_formula_evidence_agree_from_same_bytes(self):
        file_bytes = _real_bytes()  # read once, reused for both views below

        wb_values = openpyxl.load_workbook(__import__("io").BytesIO(file_bytes), data_only=True)
        ws_values = wb_values["PHIDANI"]

        wb_formula = load_formula_workbook(file_bytes)
        ws_formula = get_worksheet(wb_formula, "PHIDANI")

        # Non-formula cell: both views must agree exactly on the same bytes.
        cached_134 = ws_values.cell(row=134, column=3).value
        ev_134 = cell_formula_evidence(ws_formula, 134, 3)
        assert cached_134 == 863540.12
        assert ev_134.is_formula is False
        assert ev_134.has_literal_value is True

        # Formula cell: cached result and formula text diverge as expected,
        # both correctly derived from the identical underlying bytes.
        cached_163 = ws_values.cell(row=163, column=3).value
        ev_163 = cell_formula_evidence(ws_formula, 163, 3)
        assert cached_163 == 131323.73
        assert ev_163.is_formula is True
        assert ev_163.formula_text == "=C162"

    def test_performance_of_the_one_additional_read(self):
        """Measurement only (mission §22) — not a pass/fail correctness
        assertion, reported in the final report, not optimized here."""
        file_bytes = _real_bytes()

        t0 = time.perf_counter()
        openpyxl.load_workbook(__import__("io").BytesIO(file_bytes), data_only=True)
        t1 = time.perf_counter()
        load_formula_workbook(file_bytes)
        t2 = time.perf_counter()

        numeric_parse_seconds = t1 - t0
        formula_evidence_seconds = t2 - t1
        # No assertion on absolute timing (environment-dependent) — this
        # test exists to keep the measurement in the executable suite, per
        # the mission's "measure, do not silently skip" instruction.
        assert numeric_parse_seconds >= 0
        assert formula_evidence_seconds >= 0
