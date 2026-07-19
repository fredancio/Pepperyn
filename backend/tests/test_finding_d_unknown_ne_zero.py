"""
test_finding_d_unknown_ne_zero.py — Invariant test: ABSENCE DE DONNÉE ≠ ZÉRO FINANCIER

Vérifie que "Données insuffisantes" ne produit JAMAIS un 0 financier visible
dans aucun livrable Excel (EDM + Hypothèses).

Chaîne testée :
  source "Données insuffisantes"
  → _parse_eur() → None
  → cellule EDM (EBITDA, Cash) → None (vide)          ← ne doit JAMAIS être 0
  → cellule HYPO (CA estimé) → None (vide)             ← ne doit JAMAIS être 0
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import io
import pytest
from openpyxl import load_workbook

# ── Constants (dupliqués ici pour ne pas importer le module complet) ──────────
EDM_R_EBITDA = 5     # ligne EDM EBITDA
EDM_R_CASH   = 6     # ligne EDM Cash
HYPO_R_CA    = 11    # ligne Hypothèses CA


@pytest.fixture(scope="module")
def edm_ws_and_hypo_ws():
    """
    Génère un Excel complet avec EBITDA = Cash = CA = "Données insuffisantes"
    et retourne les feuilles EDM et Hypothèses pour assertions.
    """
    from services.excel_export import generate_excel_report
    from models.schemas import AnalysisResult

    raw_dict = {
        "ceo_dashboard": [
            {"label": "EBITDA",              "value": "Données insuffisantes"},
            {"label": "Trésorerie nette",    "value": "Données insuffisantes"},
            {"label": "CA total",            "value": "Données insuffisantes"},
        ],
        "value_destroyers": [],
        "quick_wins": [],
        "executive_summary": "",
        "risk_inaction": "",
    }

    xlsx_bytes = generate_excel_report(AnalysisResult(), raw_dict, filename="test")
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    return wb["EDM"], wb["⚙ Hypothèses"]


# ─── Test 1: _parse_eur contract ─────────────────────────────────────────────

class TestParseEurContract:
    """Unit tests for _parse_eur() return type and sentinel values."""

    def test_donnees_insuffisantes_returns_none(self):
        from services.excel_export import _parse_eur
        result = _parse_eur("Données insuffisantes")
        assert result is None, (
            f"_parse_eur('Données insuffisantes') should return None, got {result!r}. "
            "UNKNOWN financial data must never become 0."
        )

    def test_none_input_returns_none(self):
        from services.excel_export import _parse_eur
        assert _parse_eur(None) is None

    def test_empty_string_returns_none(self):
        from services.excel_export import _parse_eur
        assert _parse_eur("") is None

    def test_valid_euro_amount_returns_float(self):
        from services.excel_export import _parse_eur
        assert _parse_eur("1 800 000 €") == pytest.approx(1_800_000.0)

    def test_valid_negative_amount_returns_float(self):
        from services.excel_export import _parse_eur
        assert _parse_eur("-107 817 €") == pytest.approx(-107_817.0)

    def test_zero_euro_returns_zero(self):
        from services.excel_export import _parse_eur
        # "0 €" IS a valid financial observation (true zero) — must not become None
        assert _parse_eur("0 €") == 0.0

    def test_non_parsable_text_returns_none(self):
        from services.excel_export import _parse_eur
        assert _parse_eur("N/A") is None
        assert _parse_eur("n.c.") is None
        assert _parse_eur("En cours") is None


# ─── Test 2: EDM sheet — EBITDA cell ────────────────────────────────────────

class TestEDMSheetUnknownIsNotZero:
    """
    End-to-end: when EBITDA source is "Données insuffisantes",
    the EDM cell must NOT contain 0.
    """

    def test_ebitda_cell_is_not_zero_when_data_unavailable(self, edm_ws_and_hypo_ws):
        edm_ws, _ = edm_ws_and_hypo_ws
        cell_value = edm_ws.cell(row=EDM_R_EBITDA, column=2).value
        assert cell_value != 0 and cell_value != 0.0, (
            f"EDM!B{EDM_R_EBITDA} = {cell_value!r}. "
            "VIOLATION: 'Données insuffisantes' as EBITDA source must NOT produce 0 "
            "in the EDM sheet. UNKNOWN ≠ 0."
        )

    def test_ebitda_cell_is_none_when_data_unavailable(self, edm_ws_and_hypo_ws):
        edm_ws, _ = edm_ws_and_hypo_ws
        cell_value = edm_ws.cell(row=EDM_R_EBITDA, column=2).value
        assert cell_value is None, (
            f"EDM!B{EDM_R_EBITDA} = {cell_value!r}, expected None (blank cell). "
            "Absent data must produce a blank cell, not a zero."
        )

    def test_cash_cell_is_not_zero_when_data_unavailable(self, edm_ws_and_hypo_ws):
        edm_ws, _ = edm_ws_and_hypo_ws
        cell_value = edm_ws.cell(row=EDM_R_CASH, column=2).value
        assert cell_value != 0 and cell_value != 0.0, (
            f"EDM!B{EDM_R_CASH} = {cell_value!r}. "
            "VIOLATION: 'Données insuffisantes' as Cash source must NOT produce 0."
        )

    def test_cash_cell_is_none_when_data_unavailable(self, edm_ws_and_hypo_ws):
        edm_ws, _ = edm_ws_and_hypo_ws
        cell_value = edm_ws.cell(row=EDM_R_CASH, column=2).value
        assert cell_value is None


# ─── Test 3: Hypothèses sheet — CA estimé cell ───────────────────────────────

class TestHypothesesSheetUnknownIsNotZero:
    """
    End-to-end: when EBITDA AND CA are both "Données insuffisantes",
    the CA estimé cell in the Hypothèses sheet must NOT contain 0.
    The fallback ebitda * 2.5 must NOT fire when EBITDA is unknown.
    """

    def test_ca_cell_is_not_zero_when_both_ebitda_and_ca_unavailable(self, edm_ws_and_hypo_ws):
        _, hypo_ws = edm_ws_and_hypo_ws
        cell_value = hypo_ws.cell(row=HYPO_R_CA, column=2).value
        # cell_value will be None (blank) OR a formula string (=EDM!B...)
        # If it's a numeric 0, that's a violation
        if isinstance(cell_value, (int, float)):
            assert cell_value != 0, (
                f"Hypothèses!B{HYPO_R_CA} = {cell_value!r}. "
                "VIOLATION: CA estimé must not be 0 when both EBITDA and CA are unknown. "
                "The ebitda_num * 2.5 fallback must not fire when EBITDA is None."
            )


# ─── Test 4: Regression — known good values still work ───────────────────────

class TestKnownGoodValuesStillParsed:
    """
    Regression: fixing UNKNOWN ≠ 0 must not break parsing of real financial values.
    When EBITDA = "-107 817 €", the EDM cell must contain -107817.0, not None.
    """

    def test_real_ebitda_value_appears_in_edm(self):
        from services.excel_export import generate_excel_report
        from models.schemas import AnalysisResult

        raw_dict = {
            "ceo_dashboard": [
                {"label": "EBITDA",       "value": "-107 817 €"},
                {"label": "Trésorerie",   "value": "320 000 €"},
            ],
            "value_destroyers": [],
            "quick_wins": [],
        }

        xlsx_bytes = generate_excel_report(AnalysisResult(), raw_dict, filename="test_known")
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        edm_ws = wb["EDM"]

        ebitda_cell = edm_ws.cell(row=EDM_R_EBITDA, column=2).value
        assert ebitda_cell == pytest.approx(-107_817.0), (
            f"EDM!B{EDM_R_EBITDA} = {ebitda_cell!r}, expected -107817.0. "
            "Fixing UNKNOWN ≠ 0 must not break parsing of known EBITDA values."
        )

        cash_cell = edm_ws.cell(row=EDM_R_CASH, column=2).value
        assert cash_cell == pytest.approx(320_000.0), (
            f"EDM!B{EDM_R_CASH} = {cash_cell!r}, expected 320000.0."
        )
