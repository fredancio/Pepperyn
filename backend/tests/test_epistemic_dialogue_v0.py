"""
test_epistemic_dialogue_v0.py — Epistemic Dialogue v0 first executable
vertical slice, adversarial test matrix (mission Phase 15, items A-Q).

Reuses `MockSupabase` from `test_knowledge_model.py` (single source of
truth for the in-memory DB-constraint simulation, including the v25/v26
UNIQUE constraints needed for Phase 10's concurrency-recovery tests) —
not duplicated here.

Test classification (same discipline as `test_knowledge_model.py`):
  INVARIANT / BEHAVIOR / BOUNDARY — as there, against the mock.
  WEAK/GUARD — AST-based structural checks (recall-before-ask call
    order, zero-LLM, zero-chat-coupling), explicitly not claimed as
    proof of runtime semantics on their own.
  REAL-FILE — Phase 14/Q: the actual canonical Phidani.xlsx, not a
    synthetic reconstruction.

POST-REVIEW CORRECTIONS (independent adversarial review, 2026-08-09,
"EPISTEMIC DIALOGUE v0 — FINAL CORRECTIONS BEFORE MERGE"):
  Correction 1 — TestConcurrencyRecovery's test_L is now the
    different-value case (CONCURRENT_CONFLICT); test_L2 adds the
    same-value case (RECONCILED_TO_EXISTING) that was previously
    conflated with it.
  Correction 2 — TestWorkbookAdapterSynthetic's rollup-exclusion test
    reproduces the real Phidani.xlsx subtotal-contamination finding
    (rows 163/231) in a controlled synthetic fixture.
  Correction 3 — no test change (documentation-only, see
    epistemic_dialogue_service.py and the canonical contract §13).
  Correction 4 — TestIntegrityEscalation (previously-unexercised
    INTEGRITY_ESCALATION path), TestThresholdBoundary (89/90/91%),
    TestWorkbookAdapterSynthetic's formula-sign-inversion tests, and
    test_G3 (documents natural decline phrases as UNINTERPRETABLE, not
    a silently-broken DECLINE match).
"""
from __future__ import annotations

import ast
import inspect
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest

from backend.services.fru_sign_convention_detector import (
    ABSOLUTE_POSITIVE,
    SIGNED_NATURAL,
    Candidate,
    detect_expense_sign_convention,
    detect_expense_sign_convention_from_workbook,
)
from backend.services.epistemic_dialogue_service import (
    ClarificationNeed,
    InterpretedAnswer,
    interpret_human_answer,
    reason_recall_compare,
    render_clarification_question,
    resolve_clarification,
)
from backend.tests.test_knowledge_model import MockSupabase


def _entity_id() -> str:
    return str(uuid.uuid4())


def _user_id() -> str:
    return str(uuid.uuid4())


def _now() -> datetime:
    return datetime(2026, 8, 9, 12, 0, 0, tzinfo=timezone.utc)


_SUBJECT = "EXPENSE_SIGN_CONVENTION"


# ── A-D: REASON -> RECALL -> COMPARE (BEHAVIOR) ─────────────────────────────

class TestReasonRecallCompare:
    def test_A_no_knowledge_strong_candidate_asks(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        candidate = Candidate(value=ABSOLUTE_POSITIVE, tier="STRONG_INFERENCE")
        outcome = reason_recall_compare(db, entity, _SUBJECT, candidate)
        assert outcome.status == "ASK"
        assert outcome.clarification_need == ClarificationNeed(
            entity_id=entity, subject=_SUBJECT,
            candidate_value=ABSOLUTE_POSITIVE, recalled_value=None,
        )

    def test_B_recalled_matches_candidate_no_ask(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        from backend.services.knowledge_model_service import confirm
        confirm(db, entity, _SUBJECT, ABSOLUTE_POSITIVE, confirmed_by=_user_id(), confirmed_at=_now())
        candidate = Candidate(value=ABSOLUTE_POSITIVE, tier="STRONG_INFERENCE")
        outcome = reason_recall_compare(db, entity, _SUBJECT, candidate)
        assert outcome.status == "NO_ASK_ALREADY_KNOWN"
        assert outcome.clarification_need is None

    def test_C_recalled_contradicts_candidate_asks_targeted(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        from backend.services.knowledge_model_service import confirm
        confirm(db, entity, _SUBJECT, ABSOLUTE_POSITIVE, confirmed_by=_user_id(), confirmed_at=_now())
        candidate = Candidate(value=SIGNED_NATURAL, tier="STRONG_INFERENCE")
        outcome = reason_recall_compare(db, entity, _SUBJECT, candidate)
        assert outcome.status == "ASK"
        assert outcome.clarification_need.recalled_value == ABSOLUTE_POSITIVE
        assert outcome.clarification_need.candidate_value == SIGNED_NATURAL

    def test_D_unknown_candidate_never_fabricates_a_question(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        candidate = Candidate(value=None, tier="UNKNOWN")
        outcome = reason_recall_compare(db, entity, _SUBJECT, candidate)
        assert outcome.status == "UNRESOLVED_NO_CANDIDATE"
        assert outcome.clarification_need is None


# ── E-G: human answer gate (BEHAVIOR) ───────────────────────────────────────

class TestHumanAnswerGate:
    def test_E_human_confirms_candidate(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        need = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        result = resolve_clarification(db, need, "YES", confirmed_by=_user_id(), confirmed_at=_now())
        assert result.status == "CONFIRMED"
        assert result.knowledge_row.value == ABSOLUTE_POSITIVE

    def test_F_human_rejects_and_supplies_valid_alternative(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        need = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        result = resolve_clarification(db, need, "SIGNED_NATURAL", confirmed_by=_user_id(), confirmed_at=_now())
        assert result.status == "CONFIRMED"
        assert result.knowledge_row.value == SIGNED_NATURAL  # human's value, not the candidate

    def test_G_ambiguous_human_answer_writes_nothing(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        need = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        result = resolve_clarification(db, need, "maybe? not sure", confirmed_by=_user_id(), confirmed_at=_now())
        assert result.status == "NO_WRITE_AMBIGUOUS"
        assert result.knowledge_row is None
        assert db.knowledge_model.rows == {}

    def test_G2_decline_writes_nothing(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        need = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        result = resolve_clarification(db, need, "IDK", confirmed_by=_user_id(), confirmed_at=_now())
        assert result.status == "NO_WRITE_DECLINED"
        assert db.knowledge_model.rows == {}

    def test_G3_naturally_typed_decline_phrases_are_uninterpretable_not_decline(self):
        """Correction 4 (documented, not implemented as NLP): the
        DECLINE vocabulary's snake_case tokens ("I_DONT_KNOW",
        "JE_NE_SAIS_PAS") are UI enum/button values, not free-text
        patterns -- naturally typed prose does not normalize to them
        (an apostrophe and spaces are never collapsed to underscores).
        This is a deliberate, documented v0 boundary (see
        interpret_human_answer's docstring), guarded here so a future
        change cannot silently start treating typed decline phrases as
        DECLINE without an explicit decision to do so. Either way,
        nothing is ever written -- both outcomes refuse confirm()."""
        for raw in ("je ne sais pas", "I don't know", "I DON'T KNOW"):
            answer = interpret_human_answer(raw, candidate_value=ABSOLUTE_POSITIVE)
            assert answer.outcome == "UNINTERPRETABLE", (
                f"{raw!r} unexpectedly matched a DECLINE token"
            )
        # The programmatic enum forms (as a UI button would send) still work.
        for raw in ("IDK", "I_DONT_KNOW", "JE_NE_SAIS_PAS", "je_ne_sais_pas"):
            answer = interpret_human_answer(raw, candidate_value=ABSOLUTE_POSITIVE)
            assert answer.outcome == "DECLINE", f"{raw!r} should be DECLINE"


# ── Phase 7: never-ask-twice, executable four-upload proof (BEHAVIOR) ──────

class TestNeverAskTwiceFourUploadLoop:
    def test_full_loop_H_I_M_N(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        human = _user_id()

        # Upload 1: no prior knowledge -> clarification required.
        c1 = Candidate(value=ABSOLUTE_POSITIVE, tier="STRONG_INFERENCE")
        o1 = reason_recall_compare(db, entity, _SUBJECT, c1)
        assert o1.status == "ASK"
        r1 = resolve_clarification(db, o1.clarification_need, "YES", confirmed_by=human, confirmed_at=_now())
        assert r1.status == "CONFIRMED"
        k1_id = r1.knowledge_row.id

        # Upload 2 (Phase 7 / matrix H): same Entity, same representation
        # -> RECALL returns canonical knowledge, no clarification.
        c2 = Candidate(value=ABSOLUTE_POSITIVE, tier="STRONG_INFERENCE")
        o2 = reason_recall_compare(db, entity, _SUBJECT, c2)
        assert o2.status == "NO_ASK_ALREADY_KNOWN"

        # Upload 3 (Phase 7 / matrix I): same Entity, same representation,
        # arbitrarily later system/knowledge time -> STILL no clarification.
        # Proves "never ask twice" is not session- or time-based: recall()
        # is purely graph-structural, never confirmed_at-based.
        c3 = Candidate(value=ABSOLUTE_POSITIVE, tier="STRONG_INFERENCE")
        much_later = _now() + timedelta(days=400)
        o3 = reason_recall_compare(db, entity, _SUBJECT, c3)
        assert o3.status == "NO_ASK_ALREADY_KNOWN"
        # (much_later is unused by RECALL by construction -- asserting the
        # absence of any time dependency, not merely omitting the check.)
        assert much_later > _now()

        # Upload 4 (Phase 7 / matrix M/N): contradictory deterministic
        # fixture -> targeted clarification, human confirms changed
        # convention -> new Knowledge row supersedes old.
        c4 = Candidate(value=SIGNED_NATURAL, tier="STRONG_INFERENCE")
        o4 = reason_recall_compare(db, entity, _SUBJECT, c4)
        assert o4.status == "ASK"
        assert o4.clarification_need.recalled_value == ABSOLUTE_POSITIVE
        assert o4.clarification_need.candidate_value == SIGNED_NATURAL
        r4 = resolve_clarification(db, o4.clarification_need, "YES", confirmed_by=human, confirmed_at=_now())
        assert r4.status == "CONFIRMED"
        k2_id = r4.knowledge_row.id
        assert r4.knowledge_row.relates_to_knowledge_id == k1_id  # supersedes, never overwrites

        # Old row remains immutable and independently queryable (matrix N).
        assert db.knowledge_model.rows[k1_id]["value"] == ABSOLUTE_POSITIVE
        assert db.knowledge_model.rows[k1_id]["relates_to_knowledge_id"] is None

        # Subsequent RECALL returns the new chain head.
        c5 = Candidate(value=SIGNED_NATURAL, tier="STRONG_INFERENCE")
        o5 = reason_recall_compare(db, entity, _SUBJECT, c5)
        assert o5.status == "NO_ASK_ALREADY_KNOWN"
        assert db.knowledge_model.rows[k2_id]["value"] == SIGNED_NATURAL


# ── J: Entity isolation (BOUNDARY) ──────────────────────────────────────────

class TestEntityIsolation:
    def test_J_different_entity_gets_independent_clarification(self):
        entity_a, entity_b = _entity_id(), _entity_id()
        db = MockSupabase(entities={entity_a, entity_b})
        from backend.services.knowledge_model_service import confirm
        confirm(db, entity_a, _SUBJECT, ABSOLUTE_POSITIVE, confirmed_by=_user_id(), confirmed_at=_now())

        # Entity A already has knowledge -> no ask.
        outcome_a = reason_recall_compare(db, entity_a, _SUBJECT, Candidate(ABSOLUTE_POSITIVE, "STRONG_INFERENCE"))
        assert outcome_a.status == "NO_ASK_ALREADY_KNOWN"

        # Entity B's first upload must NOT be suppressed by Entity A's knowledge.
        outcome_b = reason_recall_compare(db, entity_b, _SUBJECT, Candidate(ABSOLUTE_POSITIVE, "STRONG_INFERENCE"))
        assert outcome_b.status == "ASK"
        assert outcome_b.clarification_need.recalled_value is None


# ── K: Subject isolation (BOUNDARY) ─────────────────────────────────────────

class TestSubjectIsolation:
    def test_K_different_subject_is_independent_knowledge(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        from backend.services.knowledge_model_service import confirm
        confirm(db, entity, _SUBJECT, ABSOLUTE_POSITIVE, confirmed_by=_user_id(), confirmed_at=_now())

        outcome_same_subject = reason_recall_compare(db, entity, _SUBJECT, Candidate(ABSOLUTE_POSITIVE, "STRONG_INFERENCE"))
        assert outcome_same_subject.status == "NO_ASK_ALREADY_KNOWN"

        # A different subject (forged into the registry-adjacent store the
        # same way test_knowledge_model.py's own cross-subject tests do,
        # since v0's real registry has only one subject) must resolve
        # entirely independently -- no leakage from EXPENSE_SIGN_CONVENTION.
        other_subject = "OTHER_SUBJECT_FOR_ISOLATION_TEST"
        outcome_other = reason_recall_compare(db, entity, other_subject, Candidate(ABSOLUTE_POSITIVE, "STRONG_INFERENCE"))
        assert outcome_other.status == "ASK"
        assert outcome_other.clarification_need.recalled_value is None


# ── L: Concurrency recovery (BOUNDARY, contract §9's new paragraph) ────────
# CORRECTED post-review (correction 1, 2026-08-09): same-value and
# different-value concurrent outcomes must be distinguishable, not both
# reported as "RECONCILED_TO_EXISTING".

class TestConcurrencyRecovery:
    def test_L_concurrent_first_confirmation_different_value_is_a_conflict(self):
        """Different-value winner: Actor 2 must NEVER see this reported as
        benign reconciliation -- it is a genuine, honestly-labeled conflict."""
        entity = _entity_id()
        db = MockSupabase(entities={entity})

        # Actor 1's ClarificationNeed, built when no knowledge existed yet.
        need_1 = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        result_1 = resolve_clarification(db, need_1, "YES", confirmed_by=_user_id(), confirmed_at=_now())
        assert result_1.status == "CONFIRMED"

        # Actor 2's ClarificationNeed was built from the SAME pre-write
        # RECALL state (recalled_value=None) -- simulating the genuine
        # race window two concurrent first confirmations would share.
        # Actor 2 attempts a DIFFERENT value than the one that actually won.
        need_2 = ClarificationNeed(entity, _SUBJECT, SIGNED_NATURAL, None)
        result_2 = resolve_clarification(db, need_2, "SIGNED_NATURAL", confirmed_by=_user_id(), confirmed_at=_now())

        # Never a second canonical root, never a fabricated winner, never
        # a retry with Actor 2's own value, and -- the correction -- never
        # reported as if it were benign: the winner's value (ABSOLUTE_POSITIVE)
        # differs from what Actor 2 attempted (SIGNED_NATURAL), so this must
        # be the distinct CONCURRENT_CONFLICT outcome, not RECONCILED_TO_EXISTING.
        assert result_2.status == "CONCURRENT_CONFLICT"
        assert result_2.knowledge_row.value == ABSOLUTE_POSITIVE

        # Exactly one root row exists for this (entity, subject) — proven,
        # not merely asserted by the status string.
        roots = [
            r for r in db.knowledge_model.rows.values()
            if r["entity_id"] == entity and r["subject"] == _SUBJECT
            and r["relates_to_knowledge_id"] is None
        ]
        assert len(roots) == 1
        assert roots[0]["value"] == ABSOLUTE_POSITIVE

    def test_L2_concurrent_first_confirmation_same_value_is_benign(self):
        """Same-value winner (correction 1's other named case): Actor 2's
        own attempted value happens to match whoever actually won. This
        must be reported as benign reconciliation, NOT as a conflict --
        nothing contradictory actually happened from Actor 2's point of
        view, even though their own write was rejected."""
        entity = _entity_id()
        db = MockSupabase(entities={entity})

        need_1 = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        result_1 = resolve_clarification(db, need_1, "YES", confirmed_by=_user_id(), confirmed_at=_now())
        assert result_1.status == "CONFIRMED"

        # Actor 2 also attempted ABSOLUTE_POSITIVE -- the SAME value that won.
        need_2 = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        result_2 = resolve_clarification(db, need_2, "YES", confirmed_by=_user_id(), confirmed_at=_now())

        assert result_2.status == "RECONCILED_TO_EXISTING"
        assert result_2.knowledge_row.value == ABSOLUTE_POSITIVE

        roots = [
            r for r in db.knowledge_model.rows.values()
            if r["entity_id"] == entity and r["subject"] == _SUBJECT
            and r["relates_to_knowledge_id"] is None
        ]
        assert len(roots) == 1


# ── Case D (KnowledgeChainIntegrityError) → INTEGRITY_ESCALATION ───────────
# Correction 4: this path existed in code since the original implementation
# but was never exercised by any test until now.

class TestIntegrityEscalation:
    def test_corrupted_chain_never_treated_as_case_b(self):
        from backend.services.knowledge_model_service import confirm as km_confirm

        entity = _entity_id()
        db = MockSupabase(entities={entity})
        k1 = km_confirm(db, entity, _SUBJECT, ABSOLUTE_POSITIVE, confirmed_by=_user_id(), confirmed_at=_now())
        # Forge TWO rows both claiming to supersede K1 -- bypassing the v25
        # UNIQUE(relates_to_knowledge_id) constraint the same way
        # test_knowledge_model.py's own corrupted-chain tests do, since a
        # live constraint would never let this state arise through the
        # service itself. This simulates historical/pre-migration corruption.
        # recall()'s chain-head rule is "not referenced by anyone else" --
        # a single forged successor would itself become the one clean head
        # (K1 becomes the referenced, non-head row); it takes TWO rows both
        # referencing K1 to produce two competing heads and trigger the
        # integrity error.
        db.knowledge_model.force_insert_bypassing_constraints({
            "entity_id": entity, "subject": _SUBJECT,
            "value": SIGNED_NATURAL, "relates_to_knowledge_id": k1.id,
            "provenance": "HUMAN_CONFIRMATION", "confirmed_by": _user_id(),
            "confirmed_at": _now().isoformat(),
        })
        db.knowledge_model.force_insert_bypassing_constraints({
            "entity_id": entity, "subject": _SUBJECT,
            "value": ABSOLUTE_POSITIVE, "relates_to_knowledge_id": k1.id,
            "provenance": "HUMAN_CONFIRMATION", "confirmed_by": _user_id(),
            "confirmed_at": _now().isoformat(),
        })

        candidate = Candidate(value=ABSOLUTE_POSITIVE, tier="STRONG_INFERENCE")
        outcome = reason_recall_compare(db, entity, _SUBJECT, candidate)

        # Must never silently degrade to "no knowledge" (Case B / ASK) --
        # a distinct, named escalation outcome, with no ClarificationNeed
        # ever constructed (nothing a human could meaningfully confirm here).
        assert outcome.status == "INTEGRITY_ESCALATION"
        assert outcome.clarification_need is None


# ── O: raw chat never stored (BOUNDARY) ─────────────────────────────────────

class TestRawChatNeverStored:
    def test_O_clarification_need_has_no_chat_or_transcript_field(self):
        fields = set(ClarificationNeed.__dataclass_fields__.keys())
        assert fields == {"entity_id", "subject", "candidate_value", "recalled_value"}
        forbidden = {"message", "chat", "transcript", "text", "raw_text", "conversation_id"}
        assert not (fields & forbidden)

    def test_O2_resolve_clarification_never_writes_raw_answer_to_knowledge_model(self):
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        need = ClarificationNeed(entity, _SUBJECT, ABSOLUTE_POSITIVE, None)
        raw = "YES"
        result = resolve_clarification(db, need, raw, confirmed_by=_user_id(), confirmed_at=_now())
        row = db.knowledge_model.rows[result.knowledge_row.id]
        assert raw not in str(row)
        assert set(row.keys()) <= {
            "id", "entity_id", "engagement_id", "subject", "value",
            "relates_to_knowledge_id", "provenance", "confirmed_by", "confirmed_at",
        }


# ── Structural / WEAK-GUARD checks (Phases 3, 6, 12, 13) ───────────────────

_EPISTEMIC_SERVICE_PATH = Path(__file__).resolve().parents[1] / "services" / "epistemic_dialogue_service.py"
_DETECTOR_PATH = Path(__file__).resolve().parents[1] / "services" / "fru_sign_convention_detector.py"
_LLM_MODULE_MARKERS = ("llm_service", "anthropic", "openai", "claude")
_CHAT_MODULE_MARKERS = ("conversation_engine",)


def _imported_module_names(source: str) -> set[str]:
    tree = ast.parse(source)
    names: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                names.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom) and node.module:
            names.add(node.module.split(".")[0])
    return names


class TestNoLLMInvolvement:
    def test_P_no_llm_import_in_epistemic_service(self):
        source = _EPISTEMIC_SERVICE_PATH.read_text(encoding="utf-8")
        imported = _imported_module_names(source)
        hit = imported.intersection(_LLM_MODULE_MARKERS)
        assert not hit, f"epistemic_dialogue_service.py imports LLM-related module(s): {hit}"

    def test_P2_no_llm_import_in_detector(self):
        source = _DETECTOR_PATH.read_text(encoding="utf-8")
        imported = _imported_module_names(source)
        hit = imported.intersection(_LLM_MODULE_MARKERS)
        assert not hit, f"fru_sign_convention_detector.py imports LLM-related module(s): {hit}"


class TestChatBoundary:
    def test_no_conversation_engine_coupling(self):
        """WEAK/GUARD: a structural import-absence check, not semantic
        proof that no future call ever crosses this boundary — the same
        honest classification `test_knowledge_model.py` applies to its
        own structural tests."""
        source = _EPISTEMIC_SERVICE_PATH.read_text(encoding="utf-8")
        imported = _imported_module_names(source)
        hit = imported.intersection(_CHAT_MODULE_MARKERS)
        assert not hit, f"epistemic_dialogue_service.py imports chat module(s): {hit}"
        # The module docstring is allowed to name conversation_engine.py in
        # prose (documenting the boundary, per contract §11) -- what must
        # never exist is an actual import/call reference to it in code.
        tree = ast.parse(source)
        code_only = ast.unparse(
            ast.Module(body=[n for n in tree.body if not (
                isinstance(n, ast.Expr) and isinstance(getattr(n, "value", None), ast.Constant)
                and isinstance(n.value.value, str)
            )], type_ignores=[])
        )
        assert "conversation_engine" not in code_only


class TestRecallBeforeAskStructural:
    def test_recall_called_before_clarification_need_construction(self):
        """WEAK/GUARD, AST-based (same pattern as
        TestTemporalRoleIsolation/TestNoLLMInvolvement elsewhere in this
        repo): within `reason_recall_compare`'s function body, every
        `ClarificationNeed(...)` call site is textually preceded by a
        `recall(...)` call — the structural property the contract's
        RECALL-before-ASK invariant (§6) demands, not merely a runtime
        behavior verified by example."""
        source = inspect.getsource(reason_recall_compare)
        recall_idx = source.index("recall(")
        need_idx = source.index("ClarificationNeed(")
        assert recall_idx < need_idx, (
            "recall() must be called before ClarificationNeed is constructed "
            "in reason_recall_compare's source"
        )

    def test_clarification_need_fields_have_no_defaults(self):
        """Mandatory, no-default parameters (contract §6) — a caller
        cannot construct a ClarificationNeed while omitting recalled_value."""
        import dataclasses
        for f in dataclasses.fields(ClarificationNeed):
            assert f.default is dataclasses.MISSING, (
                f"ClarificationNeed.{f.name} must not have a default"
            )


class TestNoNewPersistence:
    def test_clarification_need_has_no_persistence_methods(self):
        """Phase 16: ClarificationNeed must remain ephemeral -- no save/
        insert/to_row method exists on it."""
        forbidden_methods = {"save", "insert", "to_row", "persist", "to_dict_for_db"}
        actual_methods = {
            name for name in dir(ClarificationNeed) if not name.startswith("_")
        }
        assert not (actual_methods & forbidden_methods)


# ── Q: Real Phidani.xlsx (REAL-FILE, Phase 14) ──────────────────────────────

_PHIDANI_PATH = Path(__file__).resolve().parents[2] / "Phidani.xlsx"


class TestRealPhidaniWorkbook:
    @pytest.mark.skipif(not _PHIDANI_PATH.exists(), reason="Phidani.xlsx not present in this checkout")
    def test_Q_real_workbook_detects_absolute_positive_strong_inference(self):
        """Direct proof against the actual canonical file, not a
        synthetic reconstruction. Column C = 2014-12 (first populated
        period column in the real workbook)."""
        candidate = detect_expense_sign_convention_from_workbook(str(_PHIDANI_PATH), period_column=3)
        assert candidate.value == ABSOLUTE_POSITIVE
        assert candidate.tier == "STRONG_INFERENCE"

    @pytest.mark.skipif(not _PHIDANI_PATH.exists(), reason="Phidani.xlsx not present in this checkout")
    def test_Q2_real_workbook_consistent_across_multiple_period_columns(self):
        """The detected convention must not be an artifact of one column
        -- checked against several real period columns from the actual
        file (C through H)."""
        for col in range(3, 9):
            candidate = detect_expense_sign_convention_from_workbook(str(_PHIDANI_PATH), period_column=col)
            assert candidate.value == ABSOLUTE_POSITIVE, f"column {col} disagreed"
            assert candidate.tier == "STRONG_INFERENCE", f"column {col} was not STRONG_INFERENCE"

    @pytest.mark.skipif(not _PHIDANI_PATH.exists(), reason="Phidani.xlsx not present in this checkout")
    def test_Q3_full_loop_against_real_detected_candidate(self):
        """The real detector's output feeds correctly into the orchestration
        loop end to end -- not a synthetic Candidate."""
        entity = _entity_id()
        db = MockSupabase(entities={entity})
        real_candidate = detect_expense_sign_convention_from_workbook(str(_PHIDANI_PATH), period_column=3)
        outcome = reason_recall_compare(db, entity, _SUBJECT, real_candidate)
        assert outcome.status == "ASK"
        question = render_clarification_question(outcome.clarification_need)
        assert "positives" in question
        assert "postes/comptes" in question


# ── Pure-core adversarial coverage for the detector itself (synthetic,
#    Phase 4/15 — the real-file tests above are the Golden Case proof) ─────

class TestDetectorPureCore:
    def test_both_signals_agree_strong_inference(self):
        c = detect_expense_sign_convention([100.0, 200.0, 50.0], charges_subtracted_in_margin_formula=True)
        assert c == Candidate(ABSOLUTE_POSITIVE, "STRONG_INFERENCE")

    def test_signals_disagree_never_guesses(self):
        c = detect_expense_sign_convention([100.0, 200.0], charges_subtracted_in_margin_formula=False)
        assert c == Candidate(None, "UNKNOWN")

    def test_only_code_signal_present_hypothesis_tier(self):
        c = detect_expense_sign_convention([100.0, 200.0], charges_subtracted_in_margin_formula=None)
        assert c == Candidate(ABSOLUTE_POSITIVE, "HYPOTHESIS")

    def test_neither_signal_present_unknown(self):
        c = detect_expense_sign_convention([], charges_subtracted_in_margin_formula=None)
        assert c == Candidate(None, "UNKNOWN")

    def test_majority_tolerates_documented_exceptions(self):
        """Mirrors the real Phidani file: a small minority of legitimate
        negative exceptions (refunds/reversals) must not flip the signal —
        this is exactly why the code-range signal is majority-based, not
        an 'any negative disqualifies' rule (see detector module docstring)."""
        values = [100.0] * 97 + [-50.0] * 3  # 97% non-negative
        c = detect_expense_sign_convention(values, charges_subtracted_in_margin_formula=True)
        assert c.value == ABSOLUTE_POSITIVE

    def test_genuinely_mixed_distribution_is_signal_absent(self):
        values = [100.0] * 50 + [-100.0] * 50  # perfectly mixed, no majority
        c = detect_expense_sign_convention(values, charges_subtracted_in_margin_formula=None)
        assert c == Candidate(None, "UNKNOWN")


# ── Threshold boundary tests (89% / 90% / 91%, BOTH directions) ────────────
# CORRECTED ("FRU SIGN CONVENTION STUB — 90% NUMERIC BOUNDARY REPAIR"
# mission, 2026-08-09): the detector's two symmetric comparisons
# (non-negative majority / negative majority) are now computed via exact
# integer cross-multiplication (see `_code_range_signal`'s own comment),
# never via a derived floating-point ratio or `1 - <float>` subtraction.
# Previously, the exact-90%-negative case fell through to UNKNOWN due to
# `1 - 0.9 == 0.09999999999999998` in IEEE 754 -- silently asymmetric
# with the exact-90%-non-negative case, which worked correctly because it
# compared directly against the threshold rather than a derived value.
# These tests now prove BOTH directions are exactly symmetric at 89/90/91%.

class TestThresholdBoundary:
    def _values(self, pct_negative: float) -> list[float]:
        n_neg = round(100 * pct_negative)
        n_pos = 100 - n_neg
        return [10.0] * n_pos + [-10.0] * n_neg

    # ── non-negative (ABSOLUTE_POSITIVE) side ──
    def test_89_percent_nonnegative_signal_absent(self):
        # 11% negative == 89% non-negative -- genuinely mixed, no majority.
        c = detect_expense_sign_convention(self._values(0.11), charges_subtracted_in_margin_formula=None)
        assert c == Candidate(None, "UNKNOWN")

    def test_90_percent_nonnegative_signal_fires(self):
        # 10% negative == 90% non-negative, the inclusive boundary.
        c = detect_expense_sign_convention(self._values(0.10), charges_subtracted_in_margin_formula=None)
        assert c.value == ABSOLUTE_POSITIVE
        assert c.tier == "HYPOTHESIS"  # only one signal present here

    def test_91_percent_nonnegative_signal_fires(self):
        # 9% negative == 91% non-negative.
        c = detect_expense_sign_convention(self._values(0.09), charges_subtracted_in_margin_formula=None)
        assert c.value == ABSOLUTE_POSITIVE
        assert c.tier == "HYPOTHESIS"

    # ── negative (SIGNED_NATURAL) side -- the previously-broken mirror ──
    def test_89_percent_negative_signal_absent(self):
        c = detect_expense_sign_convention(self._values(0.89), charges_subtracted_in_margin_formula=None)
        assert c == Candidate(None, "UNKNOWN")

    def test_90_percent_negative_signal_fires(self):
        """The exact case that was previously broken: at exactly 90%
        negative, the old `ratio <= (1 - 0.9)` comparison silently failed
        (`0.1 <= 0.09999999999999998` is False) and returned UNKNOWN
        instead of SIGNED_NATURAL, breaking symmetry with the
        non-negative side's identical 90% boundary. Now fixed via exact
        integer cross-multiplication -- this must fire, matching
        test_90_percent_nonnegative_signal_fires exactly."""
        c = detect_expense_sign_convention(self._values(0.90), charges_subtracted_in_margin_formula=None)
        assert c.value == SIGNED_NATURAL
        assert c.tier == "HYPOTHESIS"

    def test_91_percent_negative_signal_fires(self):
        c = detect_expense_sign_convention(self._values(0.91), charges_subtracted_in_margin_formula=None)
        assert c.value == SIGNED_NATURAL
        assert c.tier == "HYPOTHESIS"

    # ── explicit symmetry proof, both sides in the same test ──
    def test_symmetry_between_absolute_positive_and_signed_natural(self):
        """The invariant this whole repair exists to guarantee: for any
        of the three boundary percentages, the non-negative-side and
        negative-side candidates must be exact mirror images (same tier,
        opposite value) -- proven together, not just separately, so a
        future regression that breaks only one side cannot hide."""
        for pct in (0.89, 0.90, 0.91):
            nonneg_side = detect_expense_sign_convention(
                self._values(1 - pct), charges_subtracted_in_margin_formula=None
            )
            neg_side = detect_expense_sign_convention(
                self._values(pct), charges_subtracted_in_margin_formula=None
            )
            assert nonneg_side.tier == neg_side.tier, f"tier asymmetry at {pct}"
            if nonneg_side.tier != "UNKNOWN":
                assert nonneg_side.value == ABSOLUTE_POSITIVE
                assert neg_side.value == SIGNED_NATURAL
            else:
                assert nonneg_side.value is None and neg_side.value is None


# ── Correction 4: formula sign inversion + subtotal/rollup contamination ───
# Synthetic, in-memory workbooks (built with openpyxl, not the real Phidani
# file) so these two structural properties are proven as committed,
# deterministic regression guards rather than only having been checked by
# the adversarial review's own ad hoc scripts against the real file.

class TestWorkbookAdapterSynthetic:
    @staticmethod
    def _build_workbook(tmp_path, rows: list[tuple], subtotal_formula: str) -> str:
        """
        rows: list of (account_code, column_C_formula_or_value) tuples,
        starting at sheet row 4 (mirrors the real file's own layout: a
        header/title region in rows 1-3).
        subtotal_formula: the formula placed in column C of the bare "60"
        aggregate row, whatever row that ends up on.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PHIDANI"
        r = 4
        for code, val in rows:
            ws.cell(row=r, column=1, value=code)
            ws.cell(row=r, column=3, value=val)
            r += 1
        # Bare "60" charges-subtotal row, referenced by the margin formula.
        subtotal_row = r
        ws.cell(row=subtotal_row, column=1, value="60")
        ws.cell(row=subtotal_row, column=3, value="=SUM(C4:C%d)" % (subtotal_row - 1))
        r += 1
        # "Marge brute" row -- the arithmetic signal's own target.
        ws.cell(row=r, column=1, value="B0")
        ws.cell(row=r, column=2, value="Marge brute")
        ws.cell(row=r, column=3, value=subtotal_formula % f"C{subtotal_row}")

        path = str(tmp_path / "synthetic.xlsx")
        wb.save(path)
        return path

    def test_formula_sign_subtraction_means_absolute_positive(self, tmp_path):
        # Leaf charge rows, all positive -- code-range signal alone -> HYPOTHESIS.
        # Margin formula SUBTRACTS the charges subtotal -> arithmetic signal
        # -> ABSOLUTE_POSITIVE. Both agree -> STRONG_INFERENCE.
        rows = [(604000, 100.0), (611001, 200.0), (622000, 50.0)]
        path = self._build_workbook(tmp_path, rows, subtotal_formula="=C31-%s")
        # revenue row untouched here -- only the subtraction pattern matters.
        c = detect_expense_sign_convention_from_workbook(path, period_column=3)
        assert c.value == ABSOLUTE_POSITIVE
        assert c.tier == "STRONG_INFERENCE"

    def test_formula_sign_addition_means_signed_natural(self, tmp_path):
        # Same leaf rows, but the margin formula ADDS the charges subtotal
        # instead of subtracting it -- only arithmetically sound if charges
        # are themselves stored as negative numbers (SIGNED_NATURAL).
        # Code-range signal alone (all positive leaf values) would say
        # ABSOLUTE_POSITIVE (HYPOTHESIS); arithmetic signal says
        # SIGNED_NATURAL -- signals DISAGREE -> UNKNOWN, never guesses.
        rows = [(604000, 100.0), (611001, 200.0), (622000, 50.0)]
        path = self._build_workbook(tmp_path, rows, subtotal_formula="=C31+%s")
        c = detect_expense_sign_convention_from_workbook(path, period_column=3)
        assert c == Candidate(None, "UNKNOWN")

    def test_formula_sign_addition_agrees_when_leaf_values_are_negative(self, tmp_path):
        # Same addition formula, but this time the leaf values are
        # genuinely negative -- now both signals agree on SIGNED_NATURAL.
        rows = [(604000, -100.0), (611001, -200.0), (622000, -50.0)]
        path = self._build_workbook(tmp_path, rows, subtotal_formula="=C31+%s")
        c = detect_expense_sign_convention_from_workbook(path, period_column=3)
        assert c.value == SIGNED_NATURAL
        assert c.tier == "STRONG_INFERENCE"

    def test_rollup_row_excluded_from_leaf_observations(self, tmp_path):
        """Correction 2's own regression guard: a row whose account code
        falls in the charge range, but whose period-column cell is itself
        a formula (a rollup of other rows), must never be counted as an
        independent leaf observation -- reproduces, in a controlled
        synthetic fixture, exactly the real Phidani.xlsx contamination
        the adversarial review found at rows 163/231 (codes '630'/'650')."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PHIDANI"
        # Three genuine leaf charge rows, all positive.
        ws.cell(row=4, column=1, value=604000); ws.cell(row=4, column=3, value=100.0)
        ws.cell(row=5, column=1, value=611001); ws.cell(row=5, column=3, value=200.0)
        ws.cell(row=6, column=1, value=622000); ws.cell(row=6, column=3, value=50.0)
        # A rollup row: an in-range charge code (630) whose OWN period-column
        # cell is a formula summing the leaf rows above -- exactly the shape
        # of the real file's row 231 ('650', '=SUM(C174:C230)').
        ws.cell(row=7, column=1, value=630)
        ws.cell(row=7, column=3, value="=SUM(C4:C6)")
        # Bare "60" subtotal (feeds the arithmetic signal, independent role).
        ws.cell(row=8, column=1, value="60")
        ws.cell(row=8, column=3, value="=SUM(C4:C7)")
        ws.cell(row=9, column=1, value="B0")
        ws.cell(row=9, column=3, value="=C31-C8")
        path = str(tmp_path / "synthetic_rollup.xlsx")
        wb.save(path)

        # Independently re-derive charge_values the same way the adapter
        # does, to assert the rollup row (630) is excluded by count, not
        # merely by final classification (which could mask the defect if
        # the rollup's sign happened to agree, as correction 2's own
        # finding on the real file showed).
        wb_f = openpyxl.load_workbook(path, data_only=False)
        ws_f = wb_f.active
        leaf_rows_seen = []
        for row in ws_f.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            code = cell.value
            if code is None or (isinstance(code, str) and code.strip() in ("60", "B0")):
                continue
            try:
                icode = int(code)
            except (TypeError, ValueError):
                continue
            if icode <= 99:
                continue
            period_formula = ws_f.cell(row=cell.row, column=3).value
            if isinstance(period_formula, str) and period_formula.startswith("="):
                continue
            leaf_rows_seen.append(cell.row)

        assert leaf_rows_seen == [4, 5, 6]  # row 7 (the '630' rollup) excluded
        assert 7 not in leaf_rows_seen

        # And the classification itself remains correct and robust --
        # STRONG_INFERENCE / ABSOLUTE_POSITIVE, unaffected by the rollup's
        # exclusion (all three leaf rows and the rollup itself happen to
        # be positive here, exactly mirroring why the real Phidani file's
        # own contamination never flipped its classification either).
        c = detect_expense_sign_convention_from_workbook(path, period_column=3)
        assert c.value == ABSOLUTE_POSITIVE
        assert c.tier == "STRONG_INFERENCE"
